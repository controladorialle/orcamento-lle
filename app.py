"""
LLE Orçamento — Acompanhamento orçado x realizado (Despesas)
Streamlit + Supabase. Segredos: SUPABASE_URL e SUPABASE_ANON_KEY.
Repaginação: menu horizontal (tabs), filtros no corpo, resumo em 2 colunas
(Mês x Acumulado YTD) e drill-down por CR -> conta para os desvios.
"""
import re
import unicodedata
import pandas as pd
import streamlit as st
from supabase import create_client

# ---------------------------------------------------------------- identidade LLE
AZUL_PROFUNDO = "#071639"; AZUL_CORP = "#183F78"; AMARELO = "#F8B11E"
VERDE = "#0F8C3B"; VERMELHO = "#C0392B"; CINZA_TXT = "#6B7583"
CINZA_BG = "#F5F7FA"; LINHA = "#E4E8F0"

st.set_page_config(page_title="Sistema de Acompanhamento Orçamentário", page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")
URL = st.secrets.get("SUPABASE_URL", ""); ANON = st.secrets.get("SUPABASE_ANON_KEY", "")

# Fragmento: isola reruns da área editada (não recarrega o app inteiro a cada célula).
# Usa st.fragment quando disponível; senão, cai para identidade (comportamento antigo).
fragment = getattr(st, "fragment", None) or getattr(st, "experimental_fragment", None) or (lambda f: f)

MESES = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
MABREV = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
STATUS_LABEL = {"PENDENTE": "Pendente", "JUSTIFICADO": "Justificado", "EM_REVISAO": "Em revisão", "DEVOLVIDO": "Devolvido", "APROVADO": "Aprovado", "ISENTA": "Isenta de justificativa"}
CATEGORIAS = [("SALARIOS", "Salário"), ("OUTROS_VENCIMENTOS", "Outros vencimentos"), ("ENCARGOS", "Encargos"), ("BENEFICIOS", "Benefícios"), ("OUTROS", "Outros de pessoal")]
CAT_LABEL = dict(CATEGORIAS)
DEDUCOES = ["Devolução de Vendas", "COFINS", "ICMS", "ICMS - Bonificação", "ICMS - ST",
            "ICMS ST - Bonificação", "ICMS Subvenção", "IPI", "PIS"]
DRE_GRUPOS = [("Receitas Financeiras", "rev"), ("Outras Receitas Não Operacionais", "rev"),
              ("Despesas Variáveis", "cost"),
              ("Despesas Comerciais", "cost"), ("Despesas Administrativas", "cost"),
              ("Depreciação e Amortização", "cost"),
              ("Despesas Financeiras", "cost"), ("Outras Despesas Operacionais", "cost"),
              ("Impostos (IRPJ/CSLL)", "cost")]
DRE_VAR_COST = ["Despesas Variáveis"]
DRE_OP_COST = ["Despesas Comerciais", "Despesas Administrativas", "Depreciação e Amortização", "Outras Despesas Operacionais"]
DRE_PRE_ADD = ["Receitas Financeiras", "Outras Receitas Não Operacionais"]
DRE_PRE_SUB = ["Despesas Financeiras"]
DRE_IMPOSTO = "Impostos (IRPJ/CSLL)"
# Grupos de RECEITA vindos do orçamento: o razão guarda receita como crédito (negativo).
# Na DRE invertemos o sinal para exibi-las positivas, como a Receita Bruta.
DRE_REV = {g for g, t in DRE_GRUPOS if t == "rev"}
DRE_LINHAS_OPC = [g for g, _ in DRE_GRUPOS]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# Mapeamento das rubricas do template Treasy nas 4 categorias
HC_MAP = {
    "SALARIOS": ["SALARIO_TOTAL", "PRO_LABORE", "JOVEM_APRENDIZ"],
    "OUTROS_VENCIMENTOS": ["HE", "ADICIONAIS", "BONIFICACOES_GRATIFICACOES",
                 "BOLSA_ESTAGIO", "QUEBRA_DE_CAIXA", "PROVISAO_FERIAS",
                 "PROVISAO_13_SALARIO", "AJUDA_DE_CUSTO", "AUXILIO_EDUCACAO", "COMISSAO"],
    "ENCARGOS": ["INSS_SALARIOS", "FGTS_SALARIOS", "FGTS_RESCISORIO", "INSS_SOBRE_PROVISAO_FERIAS",
                 "FGTS_SOBRE_PROVISAO_FERIAS", "INSS_SOBRE_PROVISAO_13_SALARIO", "FGTS_SOBRE_PROVISAO_13_SALARIO"],
    "BENEFICIOS": ["PROGRAMA_ALIMENTACAO", "ASSISTENCIA_MEDICA_ODONTO", "VALE_TRANSPORTE",
                   "VALE_COMBUSTIVEL", "QUALIDADE_DE_VIDA", "CESTA_BASICA"],
    "OUTROS": ["INDENIZACOES_AVISO", "TREINAMENTO", "UNIFORME_EPI", "EXAMES_MEDICOS", "CONTRIBUICAO_SINDICAL",
               "ABONO_PECUNIARIO_SINDICATO", "PLR", "EVENTOS_INTERNOS", "CONDUCOES_TRANSPORTES",
               "SEGURO_DE_VIDA", "DESPESAS_ENDOMARKETING"],
}
HC_COLS_DIM = ["ANO", "MES", "CODIGO_UNIDADE_NEGOCIO", "DESCRICAO_UNIDADE_NEGOCIO", "CODIGO_CENTRO_RESULTADO",
               "DESCRICAO_CENTRO_RESULTADO", "CODIGO_CARGO_FUNCIONARIO", "DESCRICAO_CARGO_FUNCIONARIO", "QUANTIDADE_FUNCIONARIOS"]

LOGO = f"""<svg width="42" height="54" viewBox="0 0 32 44" xmlns="http://www.w3.org/2000/svg">
<polygon points="16,2 23,9 16,16 9,9" fill="{AMARELO}"/><polygon points="16,10 23,17 16,24 9,17" fill="{AMARELO}"/>
<polygon points="16,18 23,25 16,32 9,25" fill="{VERDE}"/><polygon points="16,26 23,33 16,40 9,33" fill="{AZUL_CORP}"/></svg>"""

def inject_css():
    st.markdown(f"""<style>
      @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap');
      html, body, [class*="css"], .stApp {{ font-family:'Montserrat',sans-serif; }}
      .stApp {{ background:{CINZA_BG}; }}
      #MainMenu, footer, header[data-testid="stHeader"] {{ visibility:hidden; }}
      .block-container {{ padding-top:1.1rem; max-width:1360px; }}
      div.stButton>button[kind="primary"] {{ background:{AZUL_CORP}; border-color:{AZUL_CORP}; }}
      div.stButton>button[kind="primary"]:hover {{ background:{AZUL_PROFUNDO}; border-color:{AZUL_PROFUNDO}; }}

      /* header */
      .lle-header {{ background:linear-gradient(135deg,{AZUL_PROFUNDO} 0%,{AZUL_CORP} 100%);
        border-bottom:3px solid {AMARELO}; border-radius:12px; padding:15px 24px; display:flex;
        align-items:center; justify-content:space-between; margin-bottom:10px;
        box-shadow:0 4px 16px rgba(7,22,57,.15); }}
      .lle-hl {{ display:flex; align-items:center; gap:15px; }}
      .lle-header h1 {{ color:#fff; font-size:20px; font-weight:700; margin:0; letter-spacing:-.3px; }}
      .lle-header p {{ color:rgba(255,255,255,.72); font-size:12px; margin:3px 0 0; font-weight:500; }}
      .lle-badge {{ color:#fff; font-size:12px; font-weight:600; background:rgba(255,255,255,.12);
        border:1px solid rgba(255,255,255,.22); border-radius:20px; padding:6px 14px; white-space:nowrap; }}
      .lle-badge span {{ color:{AMARELO}; }}
      .modtag {{ font-size:18px; font-weight:700; color:{AZUL_PROFUNDO}; margin:2px 0 2px; }}
      .modsub {{ font-size:13px; color:{CINZA_TXT}; margin:0 0 8px; }}

      /* ----- barra lateral (navegação setorizada) ----- */
      [data-testid="stSidebar"] {{ background:#fff; border-right:1px solid {LINHA}; }}
      [data-testid="stSidebar"] .block-container {{ padding-top:1rem; }}
      .side-logo {{ background:linear-gradient(135deg,{AZUL_PROFUNDO} 0%,{AZUL_CORP} 100%);
        border-radius:12px; padding:14px 16px; display:flex; align-items:center; gap:12px;
        margin-bottom:16px; box-shadow:0 3px 10px rgba(7,22,57,.18); }}
      .side-logo-txt {{ color:#fff; font-weight:700; font-size:22px; letter-spacing:.5px; line-height:1; }}
      .navsec {{ color:{CINZA_TXT}; font-size:13px; font-weight:700; letter-spacing:1.2px;
        text-transform:uppercase; margin:16px 4px 6px; }}
      [data-testid="stSidebar"] .stButton>button {{ justify-content:flex-start !important; text-align:left !important;
        font-size:16px; font-weight:600; border:none; background:transparent; color:#1f2b45;
        padding:9px 12px; border-radius:8px; box-shadow:none; }}
      [data-testid="stSidebar"] .stButton>button > div {{ justify-content:flex-start !important; width:100%; }}
      [data-testid="stSidebar"] .stButton>button p {{ text-align:left !important; width:100%; margin:0; }}
      [data-testid="stSidebar"] .stButton>button:hover {{ background:{CINZA_BG}; color:{AZUL_PROFUNDO}; }}
      [data-testid="stSidebar"] .stButton>button[kind="primary"] {{ background:{AZUL_CORP}; color:#fff; }}
      [data-testid="stSidebar"] .stButton>button[kind="primary"]:hover {{ background:{AZUL_PROFUNDO}; }}
      .side-user {{ font-size:15px; color:#1f2b45; margin:4px 4px; line-height:1.35; }}
      .side-user span {{ color:{CINZA_TXT}; font-size:13px; }}

      /* tabs horizontais (barra navy + sublinhado dourado) */
      .stTabs [data-baseweb="tab-list"] {{ gap:2px; background:{AZUL_PROFUNDO}; padding:0 10px;
        border-radius:10px; margin-bottom:18px; box-shadow:0 3px 10px rgba(7,22,57,.14); }}
      .stTabs [data-baseweb="tab"] {{ background:transparent; color:rgba(255,255,255,.72);
        padding:13px 22px; font-weight:600; font-size:14px; border-bottom:3px solid transparent; }}
      .stTabs [data-baseweb="tab"]:hover {{ color:#fff; background:rgba(255,255,255,.05); }}
      .stTabs [aria-selected="true"] {{ color:{AMARELO} !important; background:transparent !important;
        border-bottom:3px solid {AMARELO} !important; }}
      .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] {{ display:none; }}

      /* resumo em 2 colunas */
      .resumo {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin:4px 0 6px; }}
      @media (max-width:820px) {{ .resumo {{ grid-template-columns:1fr; }} }}
      .rpanel {{ background:#fff; border:1px solid {LINHA}; border-radius:12px; overflow:hidden;
        box-shadow:0 2px 8px rgba(7,22,57,.05); }}
      .rphead {{ background:{AZUL_PROFUNDO}; color:#fff; font-weight:700; font-size:13px;
        letter-spacing:.04em; padding:10px 16px; display:flex; justify-content:space-between; align-items:center; }}
      .rphead .tag {{ font-size:10px; font-weight:600; color:{AMARELO}; }}
      .rprow {{ display:flex; justify-content:space-between; align-items:center; padding:9px 16px;
        border-bottom:1px solid #F0F2F6; font-size:13px; }}
      .rprow:last-child {{ border-bottom:none; }}
      .rprow span {{ color:{CINZA_TXT}; }}
      .rprow b {{ color:{AZUL_PROFUNDO}; font-variant-numeric:tabular-nums; font-size:14px; }}
      .rprow.big b {{ font-size:16px; }}

      /* chips de status */
      .chip {{ display:inline-block; font-size:11px; font-weight:700; padding:3px 11px; border-radius:20px; }}

      /* cards contadores */
      .cards {{ display:flex; gap:12px; flex-wrap:wrap; margin:6px 0 4px; }}
      .card {{ flex:1; min-width:150px; background:#fff; border:1px solid {LINHA}; border-radius:12px;
        padding:13px 16px; box-shadow:0 2px 8px rgba(7,22,57,.05); text-align:center; }}
      .card .lab {{ font-size:11px; color:{CINZA_TXT}; text-transform:uppercase; letter-spacing:.04em; }}
      .card .val {{ font-size:21px; color:{AZUL_PROFUNDO}; font-weight:700; margin-top:3px; }}

      /* tabelas */
      table.lle {{ border-collapse:collapse; width:100%; font-size:13px; background:#fff;
        border-radius:10px; overflow:hidden; }}
      table.lle th {{ background:{AZUL_CORP}; color:#fff; padding:9px 12px; text-align:right; font-weight:600; }}
      table.lle th:first-child {{ text-align:left; }}
      table.lle td {{ padding:7px 12px; text-align:right; border-bottom:1px solid {LINHA}; font-variant-numeric:tabular-nums; }}
      table.lle td:first-child {{ text-align:left; }}
      table.lle tr.total td {{ font-weight:700; border-top:2px solid {AZUL_CORP}; background:#EEF2F8; }}
      table.lle tr.mark td {{ background:#FFF7E6; }}
      .scroll {{ overflow-x:auto; }}
      table.lle.matrix {{ width:auto; min-width:100%; }}
      table.lle.matrix th, table.lle.matrix td {{ white-space:nowrap; }}

      /* linha-cabeça do drill (CR) */
      .drow {{ display:grid; grid-template-columns:3.1fr 1.5fr 1.5fr 1.5fr 0.9fr 1.5fr; gap:0;
        align-items:center; background:#fff; border:1px solid {LINHA}; border-radius:9px;
        padding:9px 14px; font-size:13px; margin-bottom:2px; }}
      .drow.head {{ background:{AZUL_PROFUNDO}; color:#fff; font-weight:600; border:none; }}
      .drow .r {{ text-align:right; font-variant-numeric:tabular-nums; }}
      .drow .nm {{ font-weight:600; color:{AZUL_PROFUNDO}; }}
      div[data-testid="column"] div.stButton>button {{ padding:2px 0; border:1px solid {LINHA};
        background:#fff; color:{AZUL_CORP}; font-weight:700; border-radius:7px; }}

      /* rodapé */
      .lle-foot {{ margin-top:34px; padding:16px 22px; background:linear-gradient(135deg,{AZUL_PROFUNDO} 0%,{AZUL_CORP} 100%);
        border-radius:12px; display:flex; align-items:center; justify-content:space-between; }}
      .lle-foot .t {{ color:#fff; font-weight:600; font-size:13px; }}
      .lle-foot .s {{ color:#B8C5D9; font-size:11px; margin-top:3px; }}
      .lle-foot .v {{ color:#B8C5D9; font-size:11px; text-align:right; }}
    </style>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- helpers
def norm(s): return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode().lower().strip()
def brl(n): return "R$ " + f"{round(float(n or 0)):,.0f}".replace(",", ".")
def pct_txt(n): return f"{n:+.1f}".replace(".", ",") + "%"
def _upsert_soft(c, table, row, on_conflict):
    """Upsert que não quebra se a coluna 'ativo' ainda não existir no banco (tenta sem ela)."""
    try:
        c.table(table).upsert(row, on_conflict=on_conflict).execute()
    except Exception:
        c.table(table).upsert({k: v for k, v in row.items() if k != "ativo"}, on_conflict=on_conflict).execute()

def chunks(seq, n=500):
    for i in range(0, len(seq), n): yield seq[i:i + n]

def chip(label, cor):
    return f"<span class='chip' style='color:{cor}; background:{cor}1A; border:1px solid {cor}55;'>{label}</span>"

def _toggle(k):
    st.session_state[k] = not st.session_state.get(k, False)

def classifica(raw, pct, conta_cod, banda):
    """Retorna (label, cor) respeitando a convenção de sinal e a faixa neutra."""
    invert = str(conta_cod)[:1] in ("3", "6")     # receita/dedução: sinal invertido
    impacto = -raw if invert else raw               # >0 = pior resultado
    if abs(pct) <= banda:
        return "Neutro", CINZA_TXT
    return ("Desfavorável", VERMELHO) if impacto > 0 else ("Favorável", VERDE)

def var_de(vp, vr):
    raw = (vr or 0) - (vp or 0)
    pct = raw / vp * 100 if vp else (100 if vr else 0)
    return raw, pct

def hist_n(total, key, default=20):
    """Controle compacto para os históricos: mostra as últimas N (padrão 20), com opção
    de expandir (50/100/Tudo). Devolve quantas linhas exibir. Reduz o render e mantém o drill-down."""
    if total <= default:
        return total
    opts = [20, 50, 100, "Tudo"]
    opts = [o for o in opts if (o == "Tudo") or o <= max(total, default)]
    if default in opts:
        idx = opts.index(default)
    else:
        idx = 0
    esc = st.selectbox("Exibir", opts, index=idx, key=key,
                       format_func=lambda o: (f"últimas {o}" if isinstance(o, int) else f"tudo ({total})"))
    n = total if esc == "Tudo" else min(int(esc), total)
    st.caption(f"Mostrando {n} de {total} alteração(ões) — as mais recentes primeiro.")
    return n

def _exige_justif_orc(planned, histv):
    """Justificativa do orçamento é obrigatória quando o orçado da conta excede em mais de
    5% o realizado do ano anterior (sem histórico e com orçado > 0 também exige)."""
    planned = float(planned or 0); histv = float(histv or 0)
    if planned <= 0:
        return False
    if histv <= 0:
        return True
    return planned > histv * 1.05

# ---------------------------------------------------------------- supabase
def client():
    c = create_client(URL, ANON)
    tok, rtok = st.session_state.get("access_token"), st.session_state.get("refresh_token")
    if tok and rtok:
        try: c.auth.set_session(tok, rtok)
        except Exception: pass
    return c

def get_faixa(c):
    try:
        r = c.table("config").select("valor").eq("chave", "faixa_neutra_pct").execute()
        return float(r.data[0]["valor"]) if r.data else 2.0
    except Exception:
        return 2.0

def set_faixa(c, valor):
    c.table("config").upsert({"chave": "faixa_neutra_pct", "valor": str(valor)}, on_conflict="chave").execute()

def get_cobranca(c):
    try:
        r = c.table("config").select("valor").eq("chave", "mes_inicio_cobranca").execute()
        return int(float(r.data[0]["valor"])) if r.data else 1
    except Exception:
        return 1

def set_cobranca(c, mes):
    c.table("config").upsert({"chave": "mes_inicio_cobranca", "valor": str(int(mes))}, on_conflict="chave").execute()

def get_janela(c):
    """Janela de justificativas aberta? (default: aberta)"""
    try:
        r = c.table("config").select("valor").eq("chave", "justif_aberta").execute()
        return (str(r.data[0]["valor"]) != "0") if r.data else True
    except Exception:
        return True

def set_janela(c, aberta):
    c.table("config").upsert({"chave": "justif_aberta", "valor": "1" if aberta else "0"}, on_conflict="chave").execute()

def get_isentas(c):
    """Códigos de conta isentos de justificativa (regra LLE: depreciação, contingências, impostos)."""
    try:
        r = c.table("config").select("valor").eq("chave", "justif_isentas").execute()
        if not r.data:
            return set()
        return {int(x) for x in re.findall(r"\d+", str(r.data[0]["valor"] or ""))}
    except Exception:
        return set()

def set_isentas(c, texto):
    codigos = sorted({int(x) for x in re.findall(r"\d+", texto or "")})
    c.table("config").upsert({"chave": "justif_isentas", "valor": ",".join(str(x) for x in codigos)}, on_conflict="chave").execute()
    return codigos

def get_cr_corporativos(c):
    """Códigos de CR que compõem a unidade de negócio 'Corporativo' (só afeta a visão da DRE por unidade)."""
    try:
        r = c.table("config").select("valor").eq("chave", "cr_corporativos").execute()
        if not r.data:
            return set()
        return {int(x) for x in re.findall(r"\d+", str(r.data[0]["valor"] or ""))}
    except Exception:
        return set()

def set_cr_corporativos(c, cods):
    cods = sorted({int(x) for x in cods})
    c.table("config").upsert({"chave": "cr_corporativos", "valor": ",".join(str(x) for x in cods)}, on_conflict="chave").execute()
    return cods

def get_plan_anos(c):
    """Anos habilitados para o gestor preencher o orçamento (padrão: 2027)."""
    try:
        r = c.table("config").select("valor").eq("chave", "plan_anos").execute()
        if r.data and str(r.data[0]["valor"] or "").strip():
            anos = sorted({int(x) for x in re.findall(r"\d{4}", str(r.data[0]["valor"]))})
            if anos:
                return anos
    except Exception:
        pass
    return [2027]

def set_plan_anos(c, texto):
    anos = sorted({int(x) for x in re.findall(r"\d{4}", texto or "")})
    c.table("config").upsert({"chave": "plan_anos", "valor": ",".join(str(a) for a in anos)}, on_conflict="chave").execute()
    return anos

def ler_tudo(c, tabela, ano):
    """Le todas as linhas de uma tabela para o ano, paginando de 1000 em 1000."""
    linhas, passo, ini = [], 1000, 0
    while True:
        lote = c.table(tabela).select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
        linhas.extend(lote)
        if len(lote) < passo:
            break
        ini += passo
    return linhas

def perfil(c, email):
    r = c.table("gestor_usuario").select("gestor_codigo, senha_provisoria, gestor(nome, papel)").eq("email", email).execute()
    if not r.data: return None
    g = r.data[0].get("gestor") or {}
    return {"gestor_codigo": r.data[0]["gestor_codigo"], "nome": g.get("nome", email),
            "papel": g.get("papel", "gestor"), "senha_provisoria": bool(r.data[0].get("senha_provisoria", False))}

# ---------------------------------------------------------------- leituras cacheadas
# Cache por TOKEN do usuário: respeita o RLS (cada perfil só cacheia o que pode ver)
# e é limpo em TODA escrita via limpar_cache(); TTL curto só como respaldo.
def _cli_tok(tok, rtok):
    cc = create_client(URL, ANON)
    if tok and rtok:
        try: cc.auth.set_session(tok, rtok)
        except Exception: pass
    return cc

@st.cache_data(ttl=300, show_spinner=False)
def _q_orc(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    cols = "ano,mes,uni_cod,unidade,cr_cod,cr_nome,cr_grupo,conta_cod,conta_desc,valor_planejado,valor_realizado,tipo_conta,classificacao"
    linhas, passo, ini = [], 1000, 0
    while True:
        lote = cc.table("orc_realizado").select(cols).eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
        linhas.extend(lote)
        if len(lote) < passo: break
        ini += passo
    return linhas

@st.cache_data(ttl=300, show_spinner=False)
def _q_cr_gestor(tok, rtok):
    cc = _cli_tok(tok, rtok)
    return cc.table("cr_gestor").select("uni_cod, cr_cod, cr_nome, gestor_codigo, gestor(nome)").execute().data or []

@st.cache_data(ttl=120, show_spinner=False)
def _q_justif(tok, rtok, ano, mes):
    cc = _cli_tok(tok, rtok)
    return cc.table("justificativa").select("*").eq("ano", ano).eq("mes", mes).execute().data or []

@st.cache_data(ttl=300, show_spinner=False)
def _q_oper_mes(tok, rtok, ano, mes):
    cc = _cli_tok(tok, rtok)
    linhas, passo, ini = [], 1000, 0
    while True:
        lote = (cc.table("operacional_detalhe").select("uni_cod,cr_cod,conta_cod,num_doc,valor,historico")
                .eq("ano", ano).eq("mes", mes).range(ini, ini + passo - 1).execute().data or [])
        linhas.extend(lote)
        if len(lote) < passo: break
        ini += passo
    return linhas

def _tok(): return st.session_state.get("access_token"), st.session_state.get("refresh_token")
def carregar_orc(ano): return _q_orc(*_tok(), ano)
def carregar_cr_gestor(): return _q_cr_gestor(*_tok())
def carregar_justificativas(ano, mes): return _q_justif(*_tok(), ano, mes)
def carregar_operacional(ano, mes): return _q_oper_mes(*_tok(), ano, mes)

@st.cache_data(ttl=120, show_spinner=False)
def _q_justif_ano(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    linhas, passo, ini = [], 1000, 0
    while True:
        lote = cc.table("justificativa").select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
        linhas.extend(lote)
        if len(lote) < passo: break
        ini += passo
    return linhas
def carregar_justificativas_ano(ano): return _q_justif_ano(*_tok(), ano)

@st.cache_data(ttl=300, show_spinner=False)
def _q_hc_quadro(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try:
        linhas, passo, ini = [], 1000, 0
        while True:
            lote = cc.table("hc_quadro").select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
            linhas.extend(lote)
            if len(lote) < passo: break
            ini += passo
        return linhas
    except Exception:
        return []

@st.cache_data(ttl=300, show_spinner=False)
def _q_hc_custo(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try:
        linhas, passo, ini = [], 1000, 0
        while True:
            lote = cc.table("hc_custo").select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
            linhas.extend(lote)
            if len(lote) < passo: break
            ini += passo
        return linhas
    except Exception:
        return []

def carregar_hc_quadro(ano): return _q_hc_quadro(*_tok(), ano)
def carregar_hc_custo(ano): return _q_hc_custo(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_orc_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("orc_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_orc_log(limite=200): return _q_orc_log(*_tok(), limite)

@st.cache_data(ttl=60, show_spinner=False)
def _q_hc_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("hc_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_hc_log(limite=200): return _q_hc_log(*_tok(), limite)

@st.cache_data(ttl=120, show_spinner=False)
def _q_receita(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("receita_venda").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_receita(ano=2026): return _q_receita(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_receita_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("receita_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_receita_log(limite=200): return _q_receita_log(*_tok(), limite)

@st.cache_data(ttl=120, show_spinner=False)
def _q_cmv(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("cmv_valor").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_cmv(ano=2026): return _q_cmv(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_cmv_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("cmv_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_cmv_log(limite=200): return _q_cmv_log(*_tok(), limite)

@st.cache_data(ttl=120, show_spinner=False)
def _q_deducao(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("deducao_valor").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_deducao(ano=2026): return _q_deducao(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_deducao_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("deducao_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_deducao_log(limite=200): return _q_deducao_log(*_tok(), limite)

@st.cache_data(ttl=120, show_spinner=False)
def _q_dre_mapa(tok, rtok):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("dre_mapa").select("*").execute().data or []
    except Exception: return []
def carregar_dre_mapa(): return _q_dre_mapa(*_tok())

@st.cache_data(ttl=120, show_spinner=False)
def _q_investimento(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("investimento_valor").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_investimento(ano=2026): return _q_investimento(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_investimento_log(tok, rtok, limite):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("investimento_log").select("*").order("alterado_em", desc=True).limit(limite).execute().data or []
    except Exception: return []
def carregar_investimento_log(limite=200): return _q_investimento_log(*_tok(), limite)

@st.cache_data(ttl=60, show_spinner=False)
def _q_orc_plan(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try:
        linhas, passo, ini = [], 1000, 0
        while True:
            lote = cc.table("orc_plan").select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
            linhas.extend(lote)
            if len(lote) < passo: break
            ini += passo
        return linhas
    except Exception:
        return []
def carregar_orc_plan(ano): return _q_orc_plan(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_orc_plan_status(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("orc_plan_status").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_orc_plan_status(ano): return _q_orc_plan_status(*_tok(), ano)

@st.cache_data(ttl=300, show_spinner=False)
def _q_plano_contas(tok, rtok):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("plano_contas").select("*").order("conta_cod").execute().data or []
    except Exception: return []
def carregar_plano_contas(): return _q_plano_contas(*_tok())

@st.cache_data(ttl=60, show_spinner=False)
def _q_qlp_plan(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try:
        linhas, passo, ini = [], 1000, 0
        while True:
            lote = cc.table("qlp_plan").select("*").eq("ano", ano).range(ini, ini + passo - 1).execute().data or []
            linhas.extend(lote)
            if len(lote) < passo: break
            ini += passo
        return linhas
    except Exception:
        return []
def carregar_qlp_plan(ano): return _q_qlp_plan(*_tok(), ano)

@st.cache_data(ttl=60, show_spinner=False)
def _q_qlp_status(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("qlp_plan_status").select("*").eq("ano", ano).execute().data or []
    except Exception: return []
def carregar_qlp_status(ano): return _q_qlp_status(*_tok(), ano)

@st.cache_data(ttl=300, show_spinner=False)
def _q_hc_cargo(tok, rtok):
    cc = _cli_tok(tok, rtok)
    try: return cc.table("hc_cargo").select("*").order("cargo_cod").execute().data or []
    except Exception: return []
def carregar_hc_cargo(): return _q_hc_cargo(*_tok())

@st.cache_data(ttl=120, show_spinner=False)
def _q_cargos_todos(tok, rtok):
    """Todos os cargos já vistos no headcount (hc_custo + hc_quadro), de QUALQUER ano — para o
    catálogo do QLP não ficar limitado a ano/ano-1. Paginado (lotes de 1000)."""
    cc = _cli_tok(tok, rtok)
    out = {}
    for tbl in ("hc_custo", "hc_quadro"):
        try:
            passo, ini = 1000, 0
            while True:
                lote = cc.table(tbl).select("cargo_cod,cargo_nome").range(ini, ini + passo - 1).execute().data or []
                for r in lote:
                    k = r.get("cargo_cod")
                    if k is not None:
                        out.setdefault(str(k), (r.get("cargo_nome") or str(k)))
                if len(lote) < passo:
                    break
                ini += passo
        except Exception:
            pass
    return out
def carregar_cargos_todos(): return _q_cargos_todos(*_tok())

@st.cache_data(ttl=120, show_spinner=False)
def _q_cargos_por_cr(tok, rtok, uni, cr):
    """Cargos que já existiram NESTE centro de resultado (hc_custo + hc_quadro), de qualquer ano.
    Escopo por CR: o seletor de adicionar cargo não mistura cargos de outras áreas."""
    cc = _cli_tok(tok, rtok)
    out = {}
    for tbl in ("hc_custo", "hc_quadro"):
        try:
            passo, ini = 1000, 0
            while True:
                lote = (cc.table(tbl).select("cargo_cod,cargo_nome").eq("uni_cod", uni).eq("cr_cod", cr)
                        .range(ini, ini + passo - 1).execute().data or [])
                for r in lote:
                    k = r.get("cargo_cod")
                    if k is not None:
                        out.setdefault(str(k), (r.get("cargo_nome") or str(k)))
                if len(lote) < passo:
                    break
                ini += passo
        except Exception:
            pass
    return out
def carregar_cargos_por_cr(uni, cr): return _q_cargos_por_cr(*_tok(), uni, cr)

def get_plan_janela(c, ano):
    """Janela de preenchimento do orçamento do ano-alvo aberta? (default: fechada)"""
    try:
        r = c.table("config").select("valor").eq("chave", f"plan_aberta_{ano}").execute()
        return (str(r.data[0]["valor"]) == "1") if r.data else False
    except Exception:
        return False
def set_plan_janela(c, ano, aberta):
    c.table("config").upsert({"chave": f"plan_aberta_{ano}", "valor": "1" if aberta else "0"}, on_conflict="chave").execute()

def get_qlp_janela(c, ano):
    """Janela de preenchimento do QLP (headcount) do ano-alvo aberta? (default: fechada)"""
    try:
        r = c.table("config").select("valor").eq("chave", f"qlp_aberta_{ano}").execute()
        return (str(r.data[0]["valor"]) == "1") if r.data else False
    except Exception:
        return False
def set_qlp_janela(c, ano, aberta):
    c.table("config").upsert({"chave": f"qlp_aberta_{ano}", "valor": "1" if aberta else "0"}, on_conflict="chave").execute()

def limpar_cache():
    """Limpeza total — usar em importações (mudam orçado e operacional)."""
    try: st.cache_data.clear()
    except Exception: pass

def limpar_cache_justif():
    """Limpeza cirúrgica após ações de justificativa: NÃO derruba orçado/operacional (que são pesados)."""
    for f in (_q_justif, _q_justif_ano):
        try: f.clear()
        except Exception: pass

# ---------------------------------------------------------------- login / senha
def tela_trocar_senha(c, email):
    inject_css()
    _, c2, _ = st.columns([1, 1.3, 1])
    with c2:
        st.markdown(f"<div style='text-align:center; padding:20px 0 6px;'>{LOGO}</div>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("##### Defina sua senha")
            st.caption("Este é seu primeiro acesso. Crie uma senha pessoal para continuar.")
            with st.form("trocar"):
                s1 = st.text_input("Nova senha (mínimo 6 caracteres)", type="password")
                s2 = st.text_input("Confirme a nova senha", type="password")
                ok = st.form_submit_button("Salvar e entrar", type="primary", use_container_width=True)
            if ok:
                if len(s1) < 6:
                    st.error("A senha precisa ter ao menos 6 caracteres.")
                elif s1 != s2:
                    st.error("As duas senhas não conferem.")
                else:
                    try:
                        c.auth.update_user({"password": s1})
                        c.table("gestor_usuario").update({"senha_provisoria": False}).eq("email", email).execute()
                        st.success("Senha definida! Recarregando...")
                        st.rerun()
                    except Exception:
                        st.error("Não consegui salvar a senha. Saia e entre novamente para tentar.")


def tela_login():
    inject_css()
    _, c2, _ = st.columns([1, 1.3, 1])
    with c2:
        st.markdown(f"""<div style="text-align:center; padding:26px 0 6px;"><div>{LOGO}</div>
            <h1 style="color:{AZUL_PROFUNDO}; margin:12px 0 2px;">Sistema de Acompanhamento Orçamentário</h1>
            <p style="color:{AZUL_CORP}; font-size:13px; margin:0;">LLE Ferragens · Controladoria</p></div>""", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("##### Acesso ao sistema")
            with st.form("login"):
                email = st.text_input("E-mail corporativo", placeholder="seu.email@grupolle.com.br")
                senha = st.text_input("Senha", type="password")
                ok = st.form_submit_button("Entrar", type="primary", use_container_width=True)
            if ok:
                try:
                    cli = create_client(URL, ANON)
                    res = cli.auth.sign_in_with_password({"email": email.strip(), "password": senha})
                    st.session_state.access_token = res.session.access_token
                    st.session_state.refresh_token = res.session.refresh_token
                    st.session_state.email = email.strip(); st.rerun()
                except Exception:
                    st.error("E-mail ou senha inválidos.")
            st.caption("Use o e-mail cadastrado pela controladoria.")

def header(prof):
    papel = "Controladoria" if prof["papel"] == "controladoria" else "Gestor"
    st.markdown(f"""<div class="lle-header">
        <div class="lle-hl">{LOGO}<div>
          <h1>Sistema de Acompanhamento Orçamentário — LLE Ferragens</h1>
          <p>{prof['nome']}</p></div></div>
        <div class="lle-badge">GRUPO LLE <span>—</span> {papel}</div></div>""", unsafe_allow_html=True)

def rodape():
    st.markdown(f"""<div class="lle-foot">
        <div style="display:flex; align-items:center; gap:12px;">{LOGO}
          <div><div class="t">Sistema de Acompanhamento Orçamentário</div>
          <div class="s">Controladoria · Grupo LLE Ferragens</div></div></div>
        <div class="v"><div>Módulo Despesas · Orçado x Realizado</div>
          <div style="opacity:.7; margin-top:3px;">Todos os direitos reservados</div></div></div>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- blocos de insight
def resumo_colunas(d_mes, d_ytd, banda, mes, ano):
    vp, vr = d_mes["valor_planejado"].sum(), d_mes["valor_realizado"].sum()
    raw, pct = var_de(vp, vr); lab, cor = classifica(raw, pct, "5", banda)
    yp, yr = d_ytd["valor_planejado"].sum(), d_ytd["valor_realizado"].sum()
    yraw, ypct = var_de(yp, yr); ylab, ycor = classifica(yraw, ypct, "5", banda)

    def panel(titulo, tag, o, r, va, p, lab, cor):
        return f"""<div class='rpanel'>
          <div class='rphead'><span>{titulo}</span><span class='tag'>{tag}</span></div>
          <div class='rprow'><span>Orçado</span><b>{brl(o)}</b></div>
          <div class='rprow'><span>Realizado</span><b>{brl(r)}</b></div>
          <div class='rprow'><span>Variação (R$)</span><b style='color:{cor}'>{brl(va)}</b></div>
          <div class='rprow'><span>Variação (%)</span><b style='color:{cor}'>{pct_txt(p)}</b></div>
          <div class='rprow big'><span>Status</span>{chip(lab, cor)}</div>
        </div>"""

    st.markdown("<div class='resumo'>"
        + panel("MÊS", MESES[mes] + f"/{ano}", vp, vr, raw, pct, lab, cor)
        + panel("ACUMULADO YTD", f"Jan–{MABREV[mes]}/{ano}", yp, yr, yraw, ypct, ylab, ycor)
        + "</div>", unsafe_allow_html=True)

def contadores(df_mes, banda):
    fav = desf = neu = 0
    for _, r in df_mes.iterrows():
        raw, pct = var_de(r["valor_planejado"], r["valor_realizado"])
        lab, _ = classifica(raw, pct, r["conta_cod"], banda)
        fav += lab == "Favorável"; desf += lab == "Desfavorável"; neu += lab == "Neutro"
    html = "<div class='cards'>"
    for lab, val, cor in [("Lançamentos", len(df_mes), AZUL_PROFUNDO), ("Favoráveis", fav, VERDE),
                          ("Desfavoráveis", desf, VERMELHO), ("Neutros", neu, CINZA_TXT)]:
        html += f"<div class='card'><div class='lab'>{lab}</div><div class='val' style='color:{cor}'>{val}</div></div>"
    st.markdown(html + "</div>", unsafe_allow_html=True)

def tabela_evolucao(df, banda, mes_sel, hist_m=None, hist_label=""):
    g = df.groupby("mes")[["valor_planejado", "valor_realizado"]].sum().reindex(range(1, 13), fill_value=0)
    cum = g.cumsum()
    st.caption("Clique em \u25b6 para abrir o orçado por conta e empresa do mês. YTD = acumulado até o mês.")
    hist_on = hist_m is not None
    gtc = "3.1fr 1.4fr 1.4fr 1.35fr 1.4fr 0.9fr 1.4fr"
    style_grid = f' style="grid-template-columns:{gtc}"' if hist_on else ""
    head_hist = f'<div class="r">{hist_label}</div>' if hist_on else ""
    ch1, ch2 = st.columns([0.05, 0.95])
    ch2.markdown(f"""<div class="drow head"{style_grid}><div class="nm">Mês</div>
        <div class="r">Orçado</div><div class="r">Realizado</div>{head_hist}<div class="r">Var. (R$)</div>
        <div class="r">Var. (%)</div><div class="r">Status</div></div>""", unsafe_allow_html=True)

    for m in range(1, 13):
        vp = float(g.loc[m, "valor_planejado"]); vr = float(g.loc[m, "valor_realizado"])
        raw, pct = var_de(vp, vr)
        if vr == 0:
            cor = CINZA_TXT; vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem realizado", CINZA_TXT)
        else:
            lab, cor = classifica(raw, pct, "5", banda)
            vr_txt = brl(vr); var_txt = brl(raw); pctv = pct_txt(pct); status = chip(lab, cor)
        key = f"evo_open_{m}"
        exp = st.session_state.get(key, False)
        cbtn, cbody = st.columns([0.05, 0.95])
        with cbtn:
            st.button("\u25bc" if exp else "\u25b6", key=f"btn_{key}", on_click=_toggle, args=(key,))
        exp = st.session_state.get(key, False)
        with cbody:
            _sty = []
            if m == mes_sel: _sty.append("background:#FFF7E6")
            if hist_on: _sty.append(f"grid-template-columns:{gtc}")
            row_style = f' style="{";".join(_sty)}"' if _sty else ""
            if hist_on:
                hv = float((hist_m or {}).get(m, 0.0))
                cell_hist = f'<div class="r" style="color:{CINZA_TXT}">{brl(hv) if hv else "\u2014"}</div>'
            else:
                cell_hist = ""
            st.markdown(f"""<div class="drow"{row_style}><div class="nm">{MESES[m]}</div>
                <div class="r">{brl(vp)}</div><div class="r">{vr_txt}</div>{cell_hist}
                <div class="r" style="color:{cor};font-weight:600">{var_txt}</div>
                <div class="r" style="color:{cor}">{pctv}</div>
                <div class="r">{status}</div></div>""", unsafe_allow_html=True)
            if exp:
                yvp = float(cum.loc[m, "valor_planejado"]); yvr = float(cum.loc[m, "valor_realizado"])
                yraw, ypct = var_de(yvp, yvr)
                ycor = CINZA_TXT if yvr == 0 else (VERMELHO if yraw > 0 else VERDE)
                st.markdown(f"<div style='margin:2px 0 6px 8px;font-size:.85rem;color:{CINZA_TXT}'>"
                            f"Acumulado YTD (Jan\u2013{MESES[m]}): Orçado <b>{brl(yvp)}</b> · "
                            f"Realizado <b>{brl(yvr) if yvr else '\u2014'}</b> · "
                            f"Var. <b style='color:{ycor}'>{brl(yraw) if yvr else '\u2014'}</b> "
                            f"({pct_txt(ypct) if yvr else '\u2014'})</div>", unsafe_allow_html=True)
                det = df[df["mes"] == m].copy()
                if det.empty:
                    st.caption("Sem contas para os filtros selecionados neste mês.")
                else:
                    det = det.sort_values(["cr_nome", "conta_cod", "uni_cod"])
                    to = float(det["valor_planejado"].sum()); tr = float(det["valor_realizado"].sum()); tv = tr - to
                    linhas = ""
                    for r in det.itertuples():
                        p = float(r.valor_planejado or 0); q = float(r.valor_realizado or 0); rw = q - p
                        co = CINZA_TXT if q == 0 else (VERMELHO if rw > 0 else VERDE)
                        linhas += (f"<tr><td style='text-align:left'>{int(r.cr_cod)} · {r.cr_nome}</td>"
                                   f"<td style='text-align:left'>{int(r.conta_cod)} · {r.conta_desc}</td>"
                                   f"<td style='text-align:center'>{r.unidade}</td>"
                                   f"<td>{brl(p)}</td><td>{brl(q) if q else '\u2014'}</td>"
                                   f"<td style='color:{co}'>{brl(rw) if q else '\u2014'}</td></tr>")
                    tvcor = CINZA_TXT if tr == 0 else (VERMELHO if tv > 0 else VERDE)
                    total = (f"<tr class='mark'><td style='text-align:left' colspan='3'><b>Total ({len(det)} linha(s))</b></td>"
                             f"<td><b>{brl(to)}</b></td><td><b>{brl(tr) if tr else '\u2014'}</b></td>"
                             f"<td style='color:{tvcor}'><b>{brl(tv) if tr else '\u2014'}</b></td></tr>")
                    st.markdown(f"""<div class='scroll' style='margin:2px 0 12px 8px;'><table class="lle"><tr>
                        <th style='text-align:left'>Centro de resultado</th><th style='text-align:left'>Conta</th>
                        <th style='text-align:center'>Empresa</th><th>Orçado</th><th>Realizado</th><th>Var. (R$)</th></tr>
                        {linhas}{total}</table></div>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- DRILL-DOWN por CR -> conta
def drill_desvios(d_mes, banda, mes):
    if d_mes.empty:
        st.info("Sem lançamentos para os filtros atuais."); return
    only_desf = st.checkbox("Mostrar apenas centros com desvio desfavorável", value=False, key=f"drill_only_{mes}")

    g = d_mes.groupby(["cr_cod", "cr_nome"], as_index=False)[["valor_planejado", "valor_realizado"]].sum()
    rows = []
    for _, r in g.iterrows():
        raw, pct = var_de(r["valor_planejado"], r["valor_realizado"])
        lab, cor = classifica(raw, pct, "5", banda)
        rows.append((int(r["cr_cod"]), r["cr_nome"], r["valor_planejado"], r["valor_realizado"], raw, pct, lab, cor))
    if only_desf:
        rows = [x for x in rows if x[6] == "Desfavorável"]
    rows.sort(key=lambda x: x[4], reverse=True)
    if not rows:
        st.success("Nenhum centro de resultado com desvio desfavorável neste recorte."); return

    st.caption("Clique em ▶ para abrir as contas do centro de resultado. Ordenado do maior gasto acima do previsto para o menor.")
    # cabeçalho alinhado (mesma grade do corpo, com espaço para o botão)
    ch1, ch2 = st.columns([0.05, 0.95])
    ch2.markdown("""<div class="drow head"><div class="nm">Centro de resultado</div>
        <div class="r">Orçado</div><div class="r">Realizado</div><div class="r">Var. (R$)</div>
        <div class="r">Var. (%)</div><div class="r">Status</div></div>""", unsafe_allow_html=True)

    for cr_cod, cr_nome, vp, vr, raw, pct, lab, cor in rows:
        key = f"drill_{mes}_{cr_cod}"
        exp = st.session_state.get(key, False)
        cbtn, cbody = st.columns([0.05, 0.95])
        with cbtn:
            st.button("▼" if exp else "▶", key=f"btn_{key}", on_click=_toggle, args=(key,))
        exp = st.session_state.get(key, False)
        with cbody:
            st.markdown(f"""<div class="drow"><div class="nm">{cr_nome}</div>
                <div class="r">{brl(vp)}</div><div class="r">{brl(vr)}</div>
                <div class="r" style="color:{cor};font-weight:600">{brl(raw)}</div>
                <div class="r" style="color:{cor}">{pct_txt(pct)}</div>
                <div class="r">{chip(lab, cor)}</div></div>""", unsafe_allow_html=True)
            if exp:
                sub = d_mes[d_mes["cr_cod"] == cr_cod]
                det = []
                for _, v in sub.iterrows():
                    rw, pc = var_de(v["valor_planejado"], v["valor_realizado"])
                    lb, co = classifica(rw, pc, v["conta_cod"], banda)
                    _emp = v.get("unidade", "") or ""
                    _cta = f"{v['conta_cod']} · {v.get('conta_desc','')}" + (f" ({_emp})" if _emp else "")
                    det.append((_cta, v["valor_planejado"], v["valor_realizado"], rw, pc, lb, co))
                det.sort(key=lambda x: x[3], reverse=True)
                linhas = ""
                for nome, o, r, rw, pc, lb, co in det:
                    linhas += (f"<tr><td>{nome}</td><td>{brl(o)}</td><td>{brl(r)}</td>"
                               f"<td style='color:{co}'>{brl(rw)}</td><td style='color:{co}'>{pct_txt(pc)}</td>"
                               f"<td>{chip(lb, co)}</td></tr>")
                st.markdown(f"""<div class='scroll' style='margin:2px 0 10px 8px;'><table class="lle"><tr>
                    <th>Conta</th><th>Orçado</th><th>Realizado</th><th>Var. (R$)</th><th>Var. (%)</th><th>Status</th>
                    </tr>{linhas}</table></div>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- justificativas
def secao_justificativas(c, prof, df_mes, mes, is_ctrl, banda, ano):
    if mes < get_cobranca(c):
        st.info(f"{MESES[mes]}/{ano} não está sujeito à cobrança de justificativa.")
        return
    js = carregar_justificativas(ano, mes)
    isentas = get_isentas(c)
    jmap = {(int(j["uni_cod"]), int(j["cr_cod"]), int(j["conta_cod"])): j for j in js}
    itens = []
    for _, v in df_mes.iterrows():
        raw, pct = var_de(v["valor_planejado"], v["valor_realizado"])
        lab, _ = classifica(raw, pct, v["conta_cod"], banda)
        if lab != "Desfavorável":
            continue
        j = jmap.get((int(v["uni_cod"]), int(v["cr_cod"]), int(v["conta_cod"])), {"status": "PENDENTE", "texto": "", "comentario_controladoria": ""})
        itens.append((v, raw, pct, j))
    itens.sort(key=lambda t: t[1], reverse=True)
    st_cor = {"APROVADO": VERDE, "DEVOLVIDO": VERMELHO, "PENDENTE": CINZA_TXT, "JUSTIFICADO": AZUL_CORP, "EM_REVISAO": AZUL_CORP, "ISENTA": "#8A6D1F"}

    if not itens:
        st.success("Nenhum desvio desfavorável a justificar com os filtros atuais.")
        return

    # ----- filtro por situação (o gestor escolhe o que ver) -----
    GRUPOS = {"A responder": {"PENDENTE", "DEVOLVIDO"},
              "Aguardando controladoria": {"JUSTIFICADO", "EM_REVISAO"},
              "Aprovadas": {"APROVADO"},
              "Isentas": {"ISENTA"}}
    def _st(it):
        if int(it[0]["conta_cod"]) in isentas: return "ISENTA"
        return it[3].get("status", "PENDENTE")
    n_resp = sum(1 for it in itens if _st(it) in GRUPOS["A responder"])
    n_agu = sum(1 for it in itens if _st(it) in GRUPOS["Aguardando controladoria"])
    n_apr = sum(1 for it in itens if _st(it) in GRUPOS["Aprovadas"])
    n_isen = sum(1 for it in itens if _st(it) == "ISENTA")
    opcoes = [f"Todas ({len(itens)})", f"A responder ({n_resp})",
              f"Aguardando controladoria ({n_agu})", f"Aprovadas ({n_apr})"]
    if n_isen:
        opcoes.append(f"Isentas ({n_isen})")
    escolha = st.radio("Situação", opcoes, horizontal=True, key=f"just_filtro_{mes}")
    alvo = escolha.split(" (")[0]
    vis = itens if alvo == "Todas" else [it for it in itens if _st(it) in GRUPOS[alvo]]
    if not vis:
        st.caption(f"Exibindo 0 de {len(itens)} desvio(s) desfavorável(is) em {MESES[mes]}/{ano}")
        st.info("Nenhuma conta nesta situação.")
        return

    # paginação: desenhar centenas de expanders de uma vez deixava a página lenta
    PAGE = 25
    total = len(vis)
    npag = (total + PAGE - 1) // PAGE
    if npag > 1:
        pagina = int(st.number_input("Página", min_value=1, max_value=npag, value=1, step=1, key=f"just_pag_{mes}_{alvo}"))
        pagina = max(1, min(pagina, npag))
        ini = (pagina - 1) * PAGE
        page_vis = vis[ini:ini + PAGE]
        st.caption(f"Exibindo {ini + 1}–{ini + len(page_vis)} de {total} · página {pagina} de {npag} — dica: filtre por Gestor ou Centro de resultado acima para reduzir a lista.")
    else:
        page_vis = vis
        st.caption(f"Exibindo {total} de {len(itens)} desvio(s) desfavorável(is) em {MESES[mes]}/{ano}")

    # histórico só das contas desta página (evita carregar as ~50 mil notas do mês inteiro)
    contas_pg = sorted({int(v["conta_cod"]) for v, _, _, _ in page_vis})
    crs_pg = sorted({int(v["cr_cod"]) for v, _, _, _ in page_vis})
    try:
        _op = (c.table("operacional_detalhe").select("uni_cod,cr_cod,conta_cod,num_doc,valor,historico")
               .eq("ano", ano).eq("mes", mes).in_("cr_cod", crs_pg).in_("conta_cod", contas_pg).execute().data or [])
    except Exception:
        _op = []
    oper_all = pd.DataFrame(_op)
    janela_aberta = get_janela(c)  # uma consulta só (antes era 1 por item da página)
    for v, raw, pct, j in page_vis:
        status = j.get("status", "PENDENTE")
        isenta = int(v["conta_cod"]) in isentas
        status_disp = "ISENTA" if isenta else status
        titulo = f"{v['conta_cod']} · {v.get('conta_desc','')} — {v.get('cr_nome','')} ({v.get('unidade','')}) · {brl(raw)} · [{STATUS_LABEL.get(status_disp, status_disp)}]"
        with st.expander(titulo):
            a, b, d = st.columns(3)
            a.metric("Orçado", brl(v["valor_planejado"])); b.metric("Realizado", brl(v["valor_realizado"])); d.metric("Variação", brl(raw), pct_txt(pct))
            st.markdown(f"**Histórico das notas — {v.get('unidade', '')}**")
            det = ([] if oper_all.empty else oper_all[(oper_all["uni_cod"] == v["uni_cod"]) & (oper_all["cr_cod"] == v["cr_cod"]) & (oper_all["conta_cod"] == v["conta_cod"])]
                   [["uni_cod", "num_doc", "valor", "historico"]].to_dict("records"))
            if det:
                dfd = pd.DataFrame(det)
                dfd["uni_cod"] = dfd["uni_cod"].map(lambda u: f"{u} · {'PISA' if u == 1 else 'KING' if u == 2 else '—'}")
                dfd["valor"] = dfd["valor"].map(brl)
                dfd = dfd[["uni_cod", "num_doc", "valor", "historico"]]
                dfd.columns = ["Empresa", "NF/Doc", "Valor", "Histórico"]
                st.dataframe(dfd, use_container_width=True, hide_index=True)
            else:
                st.caption("Sem detalhe operacional para esta conta neste mês.")
            key = dict(ano=ano, mes=mes, uni_cod=int(v["uni_cod"]), cr_cod=int(v["cr_cod"]), conta_cod=int(v["conta_cod"]))
            kb = f"{mes}_{v['uni_cod']}_{v['cr_cod']}_{v['conta_cod']}"
            if status == "DEVOLVIDO" and j.get("comentario_controladoria"):
                st.warning(f"Controladoria: {j['comentario_controladoria']}")
            if isenta:
                st.markdown(f"{chip('Isenta de justificativa', '#8A6D1F')}", unsafe_allow_html=True)
                st.caption("Conta isenta pela regra da LLE (ex.: depreciação, contingências, impostos) — não há cobrança de justificativa.")
            else:
                pode_editar = (not is_ctrl) and status != "APROVADO"
                if pode_editar:
                    if not janela_aberta:
                        st.info(f"Justificativa: {j.get('texto') or '—'}")
                        st.caption("🔒 Janela de justificativas fechada pela controladoria — não é possível enviar ou editar agora.")
                    else:
                        txt = st.text_area("Justificativa", value=j.get("texto", "") or "", key=f"txt_{kb}")
                        c1, c2 = st.columns(2)
                        if c1.button("Salvar rascunho", key=f"sv_{kb}"):
                            ok = False
                            try:
                                c.table("justificativa").upsert({**key, "texto": txt, "status": "PENDENTE", "atualizado_por": prof["nome"]}, on_conflict="ano,mes,uni_cod,cr_cod,conta_cod").execute(); ok = True
                            except Exception:
                                st.error("Não foi possível salvar. Se esta conta já foi enviada/aprovada, é permissão no banco — a controladoria precisa aplicar o ajuste de política de justificativa (RLS por status).")
                            if ok:
                                limpar_cache_justif(); st.rerun()
                        if c2.button("Enviar justificativa", key=f"en_{kb}", type="primary"):
                            if not txt.strip():
                                st.error("Escreva a justificativa antes de enviar.")
                            else:
                                ok = False
                                try:
                                    c.table("justificativa").upsert({**key, "texto": txt, "status": "JUSTIFICADO", "atualizado_por": prof["nome"]}, on_conflict="ano,mes,uni_cod,cr_cod,conta_cod").execute(); ok = True
                                except Exception:
                                    st.error("Não foi possível enviar. Se esta conta já foi enviada/aprovada, é permissão no banco — a controladoria precisa aplicar o ajuste de política de justificativa (RLS por status).")
                                if ok:
                                    limpar_cache_justif(); st.rerun()
                        if status in ("JUSTIFICADO", "EM_REVISAO"):
                            st.caption("Já enviada — você pode editar e reenviar enquanto a janela estiver aberta. Salvar rascunho volta para pendente.")
                elif not is_ctrl:
                    st.info(f"Justificativa: {j.get('texto') or '—'}"); st.caption("Justificativa aprovada — não editável.")
                if is_ctrl:
                    st.info(f"Justificativa do gestor: {j.get('texto') or '— (ainda não enviada)'}")
                    if status in ("JUSTIFICADO", "EM_REVISAO"):
                        coment = st.text_input("Comentário (para devolução)", key=f"cm_{kb}")
                        c1, c2 = st.columns(2)
                        if c1.button("Aprovar", key=f"ap_{kb}", type="primary"):
                            c.table("justificativa").update({"status": "APROVADO"}).match(key).execute(); limpar_cache_justif(); st.rerun()
                        if c2.button("Devolver", key=f"dv_{kb}"):
                            c.table("justificativa").update({"status": "DEVOLVIDO", "comentario_controladoria": coment}).match(key).execute(); limpar_cache_justif(); st.rerun()

# ---------------------------------------------------------------- importar / config
def modelo_hc_treasy_xlsx():
    import io
    cols = HC_COLS_DIM + [col for cat in ("SALARIOS", "ENCARGOS", "BENEFICIOS", "OUTROS") for col in HC_MAP[cat]]
    ex1 = {c: 0 for c in cols}
    ex1.update({"ANO": 2026, "MES": 6, "CODIGO_UNIDADE_NEGOCIO": 1, "DESCRICAO_UNIDADE_NEGOCIO": "PISA",
                "CODIGO_CENTRO_RESULTADO": 102001, "DESCRICAO_CENTRO_RESULTADO": "CONTROLADORIA E AUDITORIA",
                "CODIGO_CARGO_FUNCIONARIO": 427, "DESCRICAO_CARGO_FUNCIONARIO": "ANALISTA DE CONTROLADORIA",
                "QUANTIDADE_FUNCIONARIOS": 1, "SALARIO_TOTAL": 5803, "INSS_SALARIOS": 1584.22, "FGTS_SALARIOS": 464.24})
    df = pd.DataFrame([ex1], columns=cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as x:
        df.to_excel(x, index=False, sheet_name="Planilha2")
    return buf.getvalue()

# ---------------------------------------------------------------- edição de pessoal (fragmento isolado)
@fragment
def _edicao_pessoal_frag(qf, kf, mes, uni_sel, cr_sel, c, prof):
    st.caption("Ajuste headcount e custos por cargo. A edição roda isolada — o app não recarrega a cada célula. Toda alteração fica no log.")
    keys_q, disp_q, oqo, oqr = [], [], [], []
    if not qf.empty:
        qe = qf.sort_values("cargo_cod").reset_index(drop=True)
        for r in qe.itertuples():
            keys_q.append((int(r.ano), int(r.mes), int(r.uni_cod), int(r.cr_cod), str(r.cargo_cod)))
            oqo.append(round(float(getattr(r, "qtd_orcada", 0) or 0), 2))
            oqr.append(round(float(getattr(r, "qtd_realizada", 0) or 0), 2))
            disp_q.append(f"{r.cargo_cod} · {r.cargo_nome}")
        dq = pd.DataFrame({"Cargo": disp_q, "HC orçado": oqo, "HC realizado": oqr})
        st.markdown("**Quadro (quantidade de pessoas)**")
        edq = st.data_editor(dq, key=f"hce_q_{mes}_{uni_sel}_{cr_sel}", hide_index=True, use_container_width=True,
                             num_rows="fixed", disabled=["Cargo"],
                             column_config={"HC orçado": st.column_config.NumberColumn(format="%.0f", step=1),
                                            "HC realizado": st.column_config.NumberColumn(format="%.0f", step=1)})
    else:
        edq = None
    keys_k, disp_cargo, disp_cat, oko, okr = [], [], [], [], []
    if not kf.empty:
        ke = kf.sort_values(["cargo_cod", "categoria"]).reset_index(drop=True)
        for r in ke.itertuples():
            keys_k.append((int(r.ano), int(r.mes), int(r.uni_cod), int(r.cr_cod), str(r.cargo_cod), str(r.categoria)))
            oko.append(round(float(getattr(r, "valor_orcado", 0) or 0), 2))
            okr.append(round(float(getattr(r, "valor_realizado", 0) or 0), 2))
            disp_cargo.append(f"{r.cargo_cod} · {r.cargo_nome}")
            disp_cat.append(CAT_LABEL.get(str(r.categoria), str(r.categoria)))
        dk = pd.DataFrame({"Cargo": disp_cargo, "Categoria": disp_cat, "Orçado": oko, "Realizado": okr})
        st.markdown("**Custo por categoria**")
        edk = st.data_editor(dk, key=f"hce_k_{mes}_{uni_sel}_{cr_sel}", hide_index=True, use_container_width=True,
                             num_rows="fixed", disabled=["Cargo", "Categoria"],
                             column_config={"Orçado": st.column_config.NumberColumn(format="%.0f", step=1),
                                            "Realizado": st.column_config.NumberColumn(format="%.0f", step=1)})
    else:
        edk = None
    if st.button("Salvar alterações de pessoal", key="hce_save", type="primary"):
        mudou = 0
        if edq is not None:
            no, nr = list(edq["HC orçado"]), list(edq["HC realizado"])
            for i, kk in enumerate(keys_q):
                an, me, uni, cr, cgo = kk
                match = dict(ano=an, mes=me, uni_cod=uni, cr_cod=cr, cargo_cod=cgo)
                for coluna, novos, orig in (("qtd_orcada", no, oqo), ("qtd_realizada", nr, oqr)):
                    try: nv = round(float(novos[i]))
                    except (TypeError, ValueError): continue
                    if abs(nv - orig[i]) > 0.005:
                        c.table("hc_quadro").update({coluna: nv}).match(match).execute()
                        c.table("hc_log").insert(dict(**match, categoria=None, campo=coluna,
                            valor_antigo=orig[i], valor_novo=nv, alterado_por=prof.get("nome", ""))).execute()
                        mudou += 1
        if edk is not None:
            no, nr = list(edk["Orçado"]), list(edk["Realizado"])
            for i, kk in enumerate(keys_k):
                an, me, uni, cr, cgo, cat = kk
                match = dict(ano=an, mes=me, uni_cod=uni, cr_cod=cr, cargo_cod=cgo, categoria=cat)
                for coluna, novos, orig in (("valor_orcado", no, oko), ("valor_realizado", nr, okr)):
                    try: nv = round(float(novos[i]))
                    except (TypeError, ValueError): continue
                    if abs(nv - orig[i]) > 0.005:
                        c.table("hc_custo").update({coluna: nv}).match(match).execute()
                        c.table("hc_log").insert(dict(**match, campo=coluna,
                            valor_antigo=orig[i], valor_novo=nv, alterado_por=prof.get("nome", ""))).execute()
                        mudou += 1
        if mudou:
            limpar_cache(); st.success(f"{mudou} alteração(ões) de pessoal salva(s) e registrada(s) no log."); st.rerun()
        else:
            st.info("Nenhuma alteração detectada.")

def tela_headcount(c, prof, ano, mes, somente_leitura=False):
    st.markdown("<div class='modtag'>Gestão de Gastos com Pessoal</div>", unsafe_allow_html=True)
    sub = ("Quadro de pessoal e gastos — orçado x realizado (Unidade × CR × Cargo)" if not somente_leitura
           else "Quadro de pessoal e gastos dos seus centros de resultado — orçado x realizado (somente leitura)")
    st.markdown(f"<div class='modsub'>{sub}</div>", unsafe_allow_html=True)

    q = pd.DataFrame(carregar_hc_quadro(ano))
    k = pd.DataFrame(carregar_hc_custo(ano))
    k_prev = pd.DataFrame(carregar_hc_custo(ano - 1))
    if q.empty and k.empty:
        st.info("Nenhum dado de pessoal ainda. Vá à aba **Importar dados**, seção *Gestão de Gastos com Pessoal*, baixe o modelo e importe a planilha (padrão Treasy).")
        return
    for col in ("qtd_orcada", "qtd_realizada"):
        if not q.empty and col not in q.columns: q[col] = 0
    for col in ("valor_orcado", "valor_realizado"):
        if not k.empty and col not in k.columns: k[col] = 0
    if not k_prev.empty and "valor_realizado" not in k_prev.columns: k_prev["valor_realizado"] = 0
    # a planilha pode não trazer 'categoria' — blinda para não quebrar as grades/agrupamentos
    tem_cat = (not k.empty) and ("categoria" in k.columns) and bool(k["categoria"].notna().any())
    if not k.empty and "categoria" not in k.columns: k["categoria"] = "—"
    if not k_prev.empty and "categoria" not in k_prev.columns: k_prev["categoria"] = "—"

    base = pd.concat([q, k], ignore_index=True)
    def _opts_int(cod, nome):
        d = {}
        for _, r in base.iterrows():
            try: d[int(r[cod])] = str(r.get(nome, "") or int(r[cod]))
            except Exception: pass
        return d
    uni_map, cr_map = _opts_int("uni_cod", "unidade"), _opts_int("cr_cod", "cr_nome")
    cargo_map = {}
    for _, r in base.iterrows():
        cc = str(r.get("cargo_cod", "") or "")
        if cc: cargo_map[cc] = str(r.get("cargo_nome", "") or cc)

    fc = st.columns([1.4, 1.9, 1.6])
    uni_sel = fc[0].selectbox("Unidade", [0] + sorted(uni_map), format_func=lambda x: "Todas" if x == 0 else uni_map[x], key="hc_uni")
    cr_sel = fc[1].selectbox("Centro de resultado", [0] + sorted(cr_map), format_func=lambda x: "Todos" if x == 0 else f"{x} · {cr_map[x]}", key="hc_cr")
    cargo_sel = fc[2].selectbox("Cargo", [""] + sorted(cargo_map), format_func=lambda x: "Todos" if x == "" else f"{x} · {cargo_map[x]}", key="hc_cargo")

    def filtra(df, com_mes=True):
        if df.empty: return df
        d = df[df["mes"] == mes] if com_mes else df
        if uni_sel: d = d[d["uni_cod"] == uni_sel]
        if cr_sel: d = d[d["cr_cod"] == cr_sel]
        if cargo_sel: d = d[d["cargo_cod"].astype(str) == cargo_sel]
        return d
    qf, kf = filtra(q), filtra(k)
    kf_prev = filtra(k_prev)

    hc_o = int(qf["qtd_orcada"].sum()) if not qf.empty else 0
    hc_r = int(qf["qtd_realizada"].sum()) if not qf.empty else 0
    custo_o = float(kf["valor_orcado"].sum()) if not kf.empty else 0.0
    custo_r = float(kf["valor_realizado"].sum()) if not kf.empty else 0.0
    custo_hist = float(kf_prev["valor_realizado"].sum()) if not kf_prev.empty else 0.0
    d_hc = hc_r - hc_o
    d_custo_pct = (custo_r - custo_o) / custo_o * 100 if custo_o else 0.0
    medio = custo_r / hc_r if hc_r else 0.0
    cor_hc = CINZA_TXT if d_hc == 0 else (VERMELHO if d_hc > 0 else VERDE)
    cor_ct = CINZA_TXT if abs(d_custo_pct) < 0.05 else (VERMELHO if d_custo_pct > 0 else VERDE)

    cards = [("HC orçado", str(hc_o), AZUL_PROFUNDO), ("HC realizado", str(hc_r), AZUL_PROFUNDO),
             ("Δ headcount", f"{d_hc:+d}", cor_hc), ("Custo realizado", brl(custo_r), AZUL_PROFUNDO),
             (f"Custo {ano-1} (hist.)", brl(custo_hist), CINZA_TXT),
             ("Δ custo", pct_txt(d_custo_pct), cor_ct), ("Custo médio/func.", brl(medio), AZUL_PROFUNDO)]
    html = "<div class='cards'>"
    for lab, val, cor in cards:
        html += f"<div class='card'><div class='lab'>{lab}</div><div class='val' style='color:{cor}'>{val}</div></div>"
    st.markdown(html + "</div>", unsafe_allow_html=True)

    # quadro por cargo
    st.markdown("#### Quadro por cargo — orçado x realizado")
    hcg = (qf.groupby(["cargo_cod", "cargo_nome"], as_index=False).agg(o=("qtd_orcada", "sum"), r=("qtd_realizada", "sum"))
           if not qf.empty else pd.DataFrame(columns=["cargo_cod", "cargo_nome", "o", "r"]))
    custg = (kf.groupby("cargo_cod", as_index=False).agg(co=("valor_orcado", "sum"), cr=("valor_realizado", "sum"))
             if not kf.empty else pd.DataFrame(columns=["cargo_cod", "co", "cr"]))
    custg_prev = (kf_prev.groupby("cargo_cod", as_index=False).agg(chist=("valor_realizado", "sum"))
                  if not kf_prev.empty else pd.DataFrame(columns=["cargo_cod", "chist"]))
    m = hcg.merge(custg, on="cargo_cod", how="outer") if not (hcg.empty and custg.empty) else pd.DataFrame()
    if not m.empty and not custg_prev.empty:
        m = m.merge(custg_prev, on="cargo_cod", how="outer")
    if not m.empty:
        m = m.fillna(0)
        if "chist" not in m.columns: m["chist"] = 0
        m["cargo_nome"] = m["cargo_nome"].replace(0, "").astype(str)
        m = m.sort_values("cr", ascending=False)
        linhas = ""
        for _, r in m.iterrows():
            dhc = int(r["r"]) - int(r["o"])
            ch = CINZA_TXT if dhc == 0 else (VERMELHO if dhc > 0 else VERDE)
            dcp = (r["cr"] - r["co"]) / r["co"] * 100 if r["co"] else 0.0
            cc2 = CINZA_TXT if abs(dcp) < 0.05 else (VERMELHO if dcp > 0 else VERDE)
            nome = r["cargo_nome"] or f"Cargo {r['cargo_cod']}"
            linhas += (f"<tr><td style='text-align:left'>{r['cargo_cod']} · {nome}</td>"
                       f"<td style='text-align:center'>{int(r['o'])}</td><td style='text-align:center'>{int(r['r'])}</td>"
                       f"<td style='text-align:center; color:{ch}'>{dhc:+d}</td>"
                       f"<td style='color:{CINZA_TXT}'>{brl(r['chist'])}</td>"
                       f"<td>{brl(r['cr'])}</td><td style='text-align:center; color:{cc2}'>{pct_txt(dcp)}</td></tr>")
        st.markdown(f"""<table class="lle"><tr><th style='text-align:left'>Cargo</th>
            <th style='text-align:center'>HC orç.</th><th style='text-align:center'>HC real.</th>
            <th style='text-align:center'>Δ HC</th><th>Hist. {ano-1}</th><th>Custo real.</th><th style='text-align:center'>Δ %</th></tr>{linhas}</table>""", unsafe_allow_html=True)
    else:
        st.caption("Sem lançamentos para os filtros selecionados.")

    # composição por categoria (só quando a planilha traz a coluna 'categoria')
    if tem_cat and not kf.empty:
        st.markdown("#### Composição do custo de pessoal por categoria")
        catg = kf.groupby("categoria", as_index=False).agg(o=("valor_orcado", "sum"), r=("valor_realizado", "sum"))
        cmap = {row["categoria"]: (row["o"], row["r"]) for _, row in catg.iterrows()}
        cols = st.columns(len(CATEGORIAS))
        for i, (cod, lab) in enumerate(CATEGORIAS):
            o, rr = cmap.get(cod, (0.0, 0.0))
            dcp = (rr - o) / o * 100 if o else 0.0
            cor = CINZA_TXT if abs(dcp) < 0.05 else (VERMELHO if dcp > 0 else VERDE)
            cols[i].markdown(
                f"<div style='border:1px solid {LINHA}; border-radius:12px; padding:12px'>"
                f"<div style='font-size:12px; color:{CINZA_TXT}'>{lab}</div>"
                f"<div style='font-size:16px; font-weight:700; color:{AZUL_PROFUNDO}'>{brl(rr)}</div>"
                f"<div style='font-size:11px; color:{cor}'>{pct_txt(dcp)} vs orçado</div></div>", unsafe_allow_html=True)

    # evolução mensal do headcount
    st.markdown("#### Evolução mensal do headcount")
    qy = filtra(q, com_mes=False)
    if not qy.empty:
        ev = qy.groupby("mes", as_index=False).agg(o=("qtd_orcada", "sum"), r=("qtd_realizada", "sum"))
        emap = {int(row["mes"]): (int(row["o"]), int(row["r"])) for _, row in ev.iterrows()}
        cab = "".join(f"<th style='text-align:center'>{MESES[mm][:3]}</th>" for mm in range(1, 13))
        lo = "".join(f"<td style='text-align:center'>{emap.get(mm,(0,0))[0]}</td>" for mm in range(1, 13))
        lr = "".join(f"<td style='text-align:center'>{emap.get(mm,(0,0))[1]}</td>" for mm in range(1, 13))
        st.markdown(f"""<table class="lle"><tr><th style='text-align:left'>&nbsp;</th>{cab}</tr>
            <tr><td style='text-align:left'>Orçado</td>{lo}</tr>
            <tr><td style='text-align:left'>Realizado</td>{lr}</tr></table>""", unsafe_allow_html=True)
    else:
        st.caption("Sem dados de quadro para os filtros selecionados.")

    # ---------- edição manual dos valores de pessoal (com log) — só controladoria ----------
    if not somente_leitura:
        st.divider()
        with st.expander("💹 Aplicar dissídio / reajuste no Orçado", expanded=False):
            if not tem_cat:
                st.caption("A base de pessoal não tem categorias definidas; o reajuste por verba não está disponível.")
            else:
                st.caption("Aplica um percentual de reajuste sobre o **Orçado** (projeção) das verbas escolhidas, "
                           "no período e empresa selecionados. Grava tudo no histórico automaticamente. **Não altera o Realizado.**")
                cA, cB, cC = st.columns([1.1, 1, 1])
                with cA:
                    emp_r = st.radio("Empresa", ["PISA", "KING", "Ambas"], horizontal=True, key="rj_emp")
                with cB:
                    mi = st.selectbox("Mês inicial", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="rj_mi")
                with cC:
                    mf = st.selectbox("Mês final", list(range(1, 13)), index=11, format_func=lambda m: MESES[m], key="rj_mf")
                cD, cE = st.columns([1, 2])
                with cD:
                    pct = st.number_input("Percentual (%)", value=5.0, step=0.5, format="%.2f", key="rj_pct")
                with cE:
                    labels_all = [lab for _, lab in CATEGORIAS]
                    default_labs = [l for l in ("Salário", "Encargos", "Outros vencimentos") if l in labels_all]
                    cats_lab = st.multiselect("Verbas (categorias)", labels_all, default=default_labs, key="rj_cats")
                unis = {1} if emp_r == "PISA" else {2} if emp_r == "KING" else {1, 2}
                cats_sel = [cod for cod, lab in CATEGORIAS if lab in cats_lab]
                erro = "O mês inicial não pode ser maior que o mês final." if mi > mf else ("Selecione ao menos uma verba." if not cats_sel else None)
                if erro:
                    st.warning(erro)
                else:
                    kk = k.copy()
                    kk["uni_cod"] = pd.to_numeric(kk["uni_cod"], errors="coerce")
                    kk["mes"] = pd.to_numeric(kk["mes"], errors="coerce")
                    kk["valor_orcado"] = pd.to_numeric(kk["valor_orcado"], errors="coerce").fillna(0.0)
                    sel = kk[(kk["uni_cod"].isin(unis)) & (kk["mes"].between(mi, mf)) & (kk["categoria"].isin(cats_sel))]
                    sel = sel[sel["valor_orcado"] != 0]
                    tot_atual = float(sel["valor_orcado"].sum())
                    tot_novo = tot_atual * (1 + pct / 100.0)
                    p1, p2, p3 = st.columns(3)
                    p1.metric("Orçado atual (recorte)", brl(tot_atual))
                    p2.metric(f"Após {pct:.2f}%", brl(tot_novo))
                    p3.metric("Δ", brl(tot_novo - tot_atual))
                    st.caption(f"{len(sel)} linha(s) de custo serão reajustadas ({MESES[mi]}→{MESES[mf]}, {emp_r}). "
                               f"O reajuste multiplica o valor atual — aplicar duas vezes acumula o efeito.")
                    if st.button("Aplicar reajuste no Orçado", type="primary", key="rj_go", disabled=(len(sel) == 0 or pct == 0)):
                        updates, logs = [], []
                        fator = 1 + pct / 100.0
                        nome_q = prof.get("nome", "")
                        for _, r in sel.iterrows():
                            antigo = round(float(r["valor_orcado"]), 2)
                            novo = round(antigo * fator)
                            if novo == antigo: continue
                            rlz_raw = pd.to_numeric(r.get("valor_realizado"), errors="coerce")
                            rlz = round(float(rlz_raw)) if pd.notna(rlz_raw) else 0.0
                            an = int(r.get("ano") or ano); me = int(r["mes"]); uni = int(r["uni_cod"])
                            cr = int(r.get("cr_cod") or 0); cgo = str(r.get("cargo_cod")); cat = r["categoria"]
                            updates.append(dict(ano=an, mes=me, uni_cod=uni, cr_cod=cr, cargo_cod=cgo, categoria=cat,
                                                valor_orcado=novo, valor_realizado=rlz,
                                                unidade=str(r.get("unidade", "") or ""), cr_nome=str(r.get("cr_nome", "") or ""),
                                                cargo_nome=str(r.get("cargo_nome", "") or "")))
                            logs.append(dict(ano=an, mes=me, uni_cod=uni, cr_cod=cr, cargo_cod=cgo, categoria=cat,
                                             campo="valor_orcado", valor_antigo=antigo, valor_novo=novo, alterado_por=nome_q))
                        try:
                            for ch in chunks(updates):
                                c.table("hc_custo").upsert(ch, on_conflict="ano,mes,uni_cod,cr_cod,cargo_cod,categoria").execute()
                            for ch in chunks(logs):
                                c.table("hc_log").insert(ch).execute()
                            limpar_cache()
                            st.success(f"Reajuste de {pct:.2f}% aplicado em {len(updates)} linha(s) do Orçado e registrado no histórico.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Não foi possível aplicar o reajuste: {e}")
        st.divider()
        st.markdown("#### Editar valores de pessoal")
        if not cr_sel:
            st.caption("Selecione uma **Unidade** e um **Centro de resultado** acima para editar quantidades e custos deste recorte.")
        elif qf.empty and kf.empty:
            st.caption("Nada para editar neste recorte.")
        else:
            _edicao_pessoal_frag(qf, kf, mes, uni_sel, cr_sel, c, prof)

        # ---------- histórico de alterações de pessoal ----------
        log = carregar_hc_log(300)
        if log:
            st.markdown("###### Histórico de alterações de pessoal")
            CAMPO_LAB = {"qtd_orcada": "HC orçado", "qtd_realizada": "HC realizado",
                         "valor_orcado": "Custo orçado", "valor_realizado": "Custo realizado"}
            _n = hist_n(len(log), "hist_n_pessoal")
            linhas = ""
            for g in log[:_n]:
                quando = str(g.get("alterado_em", "") or "")[:16].replace("T", " ")
                va = float(g.get("valor_antigo") or 0); vn = float(g.get("valor_novo") or 0)
                seta = VERMELHO if vn > va else VERDE
                campo = g.get("campo", "")
                eh_qtd = campo in ("qtd_orcada", "qtd_realizada")
                de = (f"{va:.0f}" if eh_qtd else brl(va)); para = (f"{vn:.0f}" if eh_qtd else brl(vn))
                cat = CAT_LABEL.get(g.get("categoria") or "", "—") if not eh_qtd else "—"
                linhas += (f"<tr><td style='text-align:left'>{quando}</td><td style='text-align:left'>{g.get('alterado_por','') or '—'}</td>"
                           f"<td style='text-align:center'>{int(g.get('ano') or 0)}</td>"
                           f"<td style='text-align:left'>{MESES[int(g.get('mes') or 1)]}</td>"
                           f"<td style='text-align:left'>{g.get('cargo_cod','')}</td>"
                           f"<td style='text-align:center'>{CAMPO_LAB.get(campo, campo)}</td>"
                           f"<td style='text-align:left'>{cat}</td>"
                           f"<td>{de}</td><td style='color:{seta}'>{para}</td></tr>")
            st.markdown(f"""<table class="lle"><tr>
                <th style='text-align:left'>Quando</th><th style='text-align:left'>Quem</th><th style='text-align:center'>Ano</th><th style='text-align:left'>Mês</th>
                <th style='text-align:left'>Cargo</th><th style='text-align:center'>Campo</th>
                <th style='text-align:left'>Categoria</th><th>De</th><th>Para</th></tr>{linhas}</table>""", unsafe_allow_html=True)

def tela_importar(c, ano):
    st.markdown("<div class='modtag'>Configuração e importação de dados</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Regras de classificação, cobrança e carga das bases mensais</div>", unsafe_allow_html=True)
    st.subheader("Configuração")
    atual = get_faixa(c)
    nova = st.number_input("Faixa neutra (±%) — regra de classificação para todos os gestores",
                           value=float(atual), step=0.5, min_value=0.0)
    if st.button("Salvar faixa neutra"):
        set_faixa(c, nova); st.success(f"Faixa neutra atualizada para ±{nova:.1f}%.".replace(".", ","))
    st.markdown("**Cobrança de justificativas**")
    mc = get_cobranca(c)
    novo_mc = st.selectbox("Cobrar justificativas a partir do mês (vale para todos os anos)", list(range(1, 13)), index=mc - 1, format_func=lambda m: MESES[m])
    st.caption("Meses anteriores a este não geram pendência nem cobrança (continuam visíveis nas análises).")
    if st.button("Salvar início da cobrança"):
        set_cobranca(c, novo_mc); st.success(f"Cobrança a partir de {MESES[novo_mc]}.")
    st.divider()
    st.subheader("Importar dados")
    st.caption("A base orçado x realizado atualiza os números; o arquivo operacional traz o histórico das notas.")
    ano = st.number_input("Ano", value=int(ano), step=1)
    def pick(row, cands):
        for cand in cands:
            for k in row.index:
                if norm(k) == cand or cand in norm(k): return row[k]
        return ""
    def num(x):
        if x is None or pd.isna(x): return 0.0
        if isinstance(x, (int, float)): return float(x)
        s = str(x).replace("R$", "").strip()
        if not s: return 0.0
        if "," in s: s = s.replace(".", "").replace(",", ".")
        try: return float(s)
        except ValueError: return 0.0
    def toint(x):
        if x is None or pd.isna(x): return None
        s = str(x).strip()
        try: return int(float(s)) if s else None
        except ValueError: return None
    f1 = st.file_uploader("1) Base orçado x realizado", type=["xlsx"], key="f1")
    if f1 and st.button("Importar orçado x realizado"):
        df = pd.read_excel(f1); recs = []
        for _, r in df.iterrows():
            uni, cr, ct = toint(pick(r, ["codigo unidade"])), toint(pick(r, ["codigo centro"])), toint(pick(r, ["codigo conta"]))
            if None in (uni, cr, ct): continue
            recs.append(dict(ano=toint(pick(r, ["ano"])) or int(ano), mes=toint(pick(r, ["mes"])) or 1, uni_cod=uni,
                unidade=str(pick(r, ["descricao unidade"]) or ""), cr_cod=cr, cr_nome=str(pick(r, ["descricao centro"]) or ""),
                cr_grupo=str(pick(r, ["cr grupo"]) or ""), conta_cod=ct, conta_desc=str(pick(r, ["descricao conta"]) or ""),
                valor_planejado=round(num(pick(r, ["valor planejado"]))), valor_realizado=round(num(pick(r, ["valor realizado"]))),
                tipo_conta=str(pick(r, ["tipo conta"]) or ""), classificacao=str(pick(r, ["classificacao"]) or "")))
        for ch in chunks(recs):
            c.table("orc_realizado").upsert(ch, on_conflict="ano,mes,uni_cod,cr_cod,conta_cod").execute()
        st.success(f"{len(recs)} linhas de orçado x realizado importadas.")
        limpar_cache()
    f2 = st.file_uploader("2) Detalhe operacional (com COMPLHIST)", type=["xlsx"], key="f2")
    mes_op = st.number_input("Mês do arquivo operacional", value=6, min_value=1, max_value=12, step=1)
    if f2 and st.button("Importar histórico das notas"):
        raw = pd.read_excel(f2, header=None, nrows=10); hi = 0
        for i in range(min(10, len(raw))):
            if any("complhist" in norm(x) for x in raw.iloc[i].tolist()): hi = i; break
        df = pd.read_excel(f2, header=hi).dropna(how="all"); recs = []
        for _, r in df.iterrows():
            uni, cr, ct = toint(pick(r, ["codigo unidade", "codigo_unidade"])), toint(pick(r, ["codcencus", "codigo centro"])), toint(pick(r, ["codctactb", "codigo conta"]))
            if None in (uni, cr, ct): continue
            h = pick(r, ["complhist"])
            m = re.search(r"\b(?:NF|CFE)\s*0*([0-9]{3,})", str(h), re.I) if isinstance(h, str) else None
            hist = re.sub(r"\s*-?\d+\.\d+\s*$", "", str(h)).strip() if isinstance(h, str) else str(h or "")
            recs.append(dict(ano=toint(pick(r, ["ano"])) or int(ano), mes=toint(pick(r, ["mes"])) or int(mes_op),
                uni_cod=uni, cr_cod=cr, conta_cod=ct, num_doc=(m.group(1) if m else None), valor=round(num(pick(r, ["valor"]))), historico=hist))
        c.table("operacional_detalhe").delete().eq("ano", int(ano)).eq("mes", int(mes_op)).execute()
        for ch in chunks(recs):
            c.table("operacional_detalhe").insert(ch).execute()
        st.success(f"{len(recs)} lançamentos com histórico importados.")
        limpar_cache()

    st.divider()
    st.subheader("Gestão de Gastos com Pessoal — planilha (padrão Treasy)")
    st.caption("Arquivo único no layout Treasy: uma linha por unidade × CR × cargo, com QUANTIDADE_FUNCIONARIOS e as "
               "rubricas de custo. O sistema soma as rubricas nas 4 categorias automaticamente. Importe o arquivo de "
               "Realizado e, se tiver, o de Orçado — cada tipo preenche a sua coluna, sem apagar o outro.")
    st.download_button("⬇️ Modelo — layout Treasy", data=modelo_hc_treasy_xlsx(),
                       file_name="modelo_headcount_treasy.xlsx", mime=XLSX_MIME, key="mdl_hc")

    def salvar_cargos(cargos):
        recs = [dict(cargo_cod=k, cargo_nome=v, ativo=True) for k, v in cargos.items() if k]
        for ch in chunks(recs):
            try: c.table("hc_cargo").upsert(ch, on_conflict="cargo_cod").execute()
            except Exception: pass

    def cod_cargo(x):
        if x is None or pd.isna(x): return None
        if isinstance(x, (int, float)):
            return str(int(x)) if float(x).is_integer() else str(x)
        return str(x).strip()

    tipo = st.radio("Este arquivo é", ["Realizado", "Orçado"], horizontal=True, key="hc_tipo")
    fh = st.file_uploader("Planilha de pessoal (layout Treasy)", type=["xlsx"], key="fh")
    mes_hc = st.number_input("Mês deste arquivo de pessoal", value=6, min_value=1, max_value=12, step=1, key="mes_hc",
                             help="Usado quando a planilha não traz a coluna MES preenchida. Se a planilha tiver ANO/MES por linha, eles prevalecem.")
    if fh and st.button("Importar planilha de pessoal", key="imp_hc"):
        df = pd.read_excel(fh).dropna(how="all")
        cmap = {norm(col): col for col in df.columns}
        def cv(row, name):
            col = cmap.get(norm(name))
            return row[col] if col is not None else None
        realizado = tipo == "Realizado"
        qcol = "qtd_realizada" if realizado else "qtd_orcada"
        vcol = "valor_realizado" if realizado else "valor_orcado"
        # coluna do OUTRO tipo, que deve ser PRESERVADA (nunca zerada) neste import
        outro_v = "valor_orcado" if realizado else "valor_realizado"
        outro_q = "qtd_orcada" if realizado else "qtd_realizada"
        # valores já existentes do outro tipo (leitura DIRETA do banco, sem cache,
        # para preservar a outra coluna sem regravar valores defasados)
        exist_v = {}
        try:
            _cur, _p, _i = [], 1000, 0
            while True:
                _lote = c.table("hc_custo").select("ano,mes,uni_cod,cr_cod,cargo_cod,categoria,valor_orcado,valor_realizado").eq("ano", ano).range(_i, _i + _p - 1).execute().data or []
                _cur.extend(_lote)
                if len(_lote) < _p: break
                _i += _p
        except Exception:
            _cur = []
        for x in _cur:
            exist_v[(int(x.get("ano") or 0), int(x.get("mes") or 0), int(x.get("uni_cod") or 0),
                     int(x.get("cr_cod") or 0), str(x.get("cargo_cod")), x.get("categoria"))] = float(x.get(outro_v) or 0)
        exist_q = {}
        try:
            _curq, _p, _i = [], 1000, 0
            while True:
                _lote = c.table("hc_quadro").select("ano,mes,uni_cod,cr_cod,cargo_cod,qtd_orcada,qtd_realizada").eq("ano", ano).range(_i, _i + _p - 1).execute().data or []
                _curq.extend(_lote)
                if len(_lote) < _p: break
                _i += _p
        except Exception:
            _curq = []
        for x in _curq:
            exist_q[(int(x.get("ano") or 0), int(x.get("mes") or 0), int(x.get("uni_cod") or 0),
                     int(x.get("cr_cod") or 0), str(x.get("cargo_cod")))] = float(x.get(outro_q) or 0)
        quadro, custo, cargos = [], [], {}
        for _, r in df.iterrows():
            uni = toint(cv(r, "CODIGO_UNIDADE_NEGOCIO")); cr = toint(cv(r, "CODIGO_CENTRO_RESULTADO"))
            cgo = cod_cargo(cv(r, "CODIGO_CARGO_FUNCIONARIO"))
            if uni is None or cr is None or not cgo: continue
            an = toint(cv(r, "ANO")) or int(ano); me = toint(cv(r, "MES")) or int(mes_hc)
            cargo_nome = str(cv(r, "DESCRICAO_CARGO_FUNCIONARIO") or "")
            cargos[cgo] = cargo_nome
            dim = dict(ano=an, mes=me, uni_cod=uni, unidade=str(cv(r, "DESCRICAO_UNIDADE_NEGOCIO") or ""),
                       cr_cod=cr, cr_nome=str(cv(r, "DESCRICAO_CENTRO_RESULTADO") or ""),
                       cargo_cod=cgo, cargo_nome=cargo_nome)
            quadro.append({**dim, qcol: num(cv(r, "QUANTIDADE_FUNCIONARIOS")),
                           outro_q: exist_q.get((an, me, uni, cr, cgo), 0)})
            for cat, rubricas in HC_MAP.items():
                total = round(sum(num(cv(r, rub)) for rub in rubricas))
                custo.append({**dim, "categoria": cat, vcol: total,
                              outro_v: exist_v.get((an, me, uni, cr, cgo, cat), 0)})
        # --- Aviso defensivo: colunas numéricas que o sistema NÃO reconhece ---
        # Evita repetir o caso de uma rubrica com nome fora do padrão (ex.: SALARIO_UNITARIO
        # em vez de SALARIO_TOTAL) sumir da carga silenciosamente. Não bloqueia a importação.
        mapeadas = {norm(x) for x in HC_COLS_DIM}
        for _rubs in HC_MAP.values():
            mapeadas |= {norm(x) for x in _rubs}
        ignoradas = []
        for col in df.columns:
            if norm(col) in mapeadas:
                continue
            serie = pd.to_numeric(df[col], errors="coerce")
            soma = float(serie.fillna(0).sum())
            if abs(soma) > 0.005:
                ignoradas.append((str(col), soma))
        if ignoradas:
            ignoradas.sort(key=lambda t: -abs(t[1]))
            det = "; ".join(f"{nome} = {brl(v)}" for nome, v in ignoradas)
            st.warning("⚠️ A planilha tem coluna(s) numérica(s) que o sistema **não reconhece** e que "
                       "**não entraram** na carga (não constam no mapa de rubricas HC_MAP): " + det +
                       ". Se alguma for custo de pessoal, renomeie na planilha para a rubrica correta "
                       "(ex.: SALARIO_TOTAL) e reimporte — ou me avise para incluí-la no mapa.")
        salvar_cargos(cargos)
        for ch in chunks(quadro):
            c.table("hc_quadro").upsert(ch, on_conflict="ano,mes,uni_cod,cr_cod,cargo_cod").execute()
        for ch in chunks(custo):
            c.table("hc_custo").upsert(ch, on_conflict="ano,mes,uni_cod,cr_cod,cargo_cod,categoria").execute()
        periodos = sorted({(x["ano"], x["mes"]) for x in quadro})
        per = ", ".join(f"{m:02d}/{a}" for a, m in periodos)
        st.success(f"{len(quadro)} cargo(s) importado(s) como {tipo}" + (f" — período(s): {per}." if per else "."))
        limpar_cache()

# ---------------------------------------------------------------- painel de recebidas
def relatorio_justificativas_xlsx(js_rows, df_orc, cg):
    """Monta um .xlsx das justificativas inseridas, enriquecido com orçado/realizado/variação."""
    import io
    fin = {}
    if df_orc is not None and not df_orc.empty:
        for _, r in df_orc.iterrows():
            fin[(int(r["mes"]), int(r["uni_cod"]), int(r["cr_cod"]), int(r["conta_cod"]))] = (
                float(r["valor_planejado"] or 0), float(r["valor_realizado"] or 0),
                r.get("unidade", "") or "", r.get("cr_nome", "") or "", r.get("conta_desc", "") or "")
    linhas = []
    for j in js_rows:
        mes = int(j["mes"]); uni = int(j["uni_cod"]); cr = int(j["cr_cod"]); ct = int(j["conta_cod"])
        vp, vr, unidade, cr_nome, conta_desc = fin.get((mes, uni, cr, ct), (0.0, 0.0, "", "", ""))
        raw, pct = var_de(vp, vr)
        gestor = cg.get((uni, cr), ("—", ""))[0]
        if not cr_nome:
            cr_nome = cg.get((uni, cr), ("", ""))[1]
        if not unidade:
            unidade = "PISA" if uni == 1 else "KING" if uni == 2 else f"Emp {uni}"
        linhas.append({
            "Ano": int(j.get("ano", 2026)), "Mês": MESES[mes], "Nº Mês": mes,
            "Gestor": gestor, "Cód. Unidade": uni, "Unidade": unidade,
            "Cód. CR": cr, "Centro de Resultado": cr_nome,
            "Cód. Conta": ct, "Descrição da Conta": conta_desc,
            "Orçado": round(vp), "Realizado": round(vr),
            "Variação (R$)": round(raw), "Variação (%)": round(pct, 2),
            "Situação": STATUS_LABEL.get(j.get("status", "PENDENTE"), j.get("status", "")),
            "Justificativa": j.get("texto", "") or "",
            "Comentário Controladoria": j.get("comentario_controladoria", "") or "",
            "Atualizado por": j.get("atualizado_por", "") or "",
            "Atualizado em": str(j.get("atualizado_em", "") or ""),
        })
    df = pd.DataFrame(linhas)
    if not df.empty:
        df = df.sort_values(["Nº Mês", "Gestor", "Centro de Resultado", "Cód. Conta"], kind="stable").drop(columns=["Nº Mês"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Justificativas")
    return buf.getvalue(), len(df)

def tela_painel(c, prof, banda, df_orc, cg, ano, mes):
    st.markdown("<div class='modtag'>Justificativas recebidas</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Pendências por gestor, resumo e detalhe das respostas</div>", unsafe_allow_html=True)

    # ----- Janela de justificativas: abrir/fechar a qualquer momento -----
    janela = get_janela(c)
    jc = st.columns([3, 1.4])
    jc[0].markdown(
        f"<div style='padding-top:8px'>Janela de justificativas: "
        f"<b style='color:{VERDE if janela else VERMELHO}'>{'ABERTA' if janela else 'FECHADA'}</b></div>",
        unsafe_allow_html=True)
    if jc[1].button(("🔒 Fechar janela" if janela else "🔓 Abrir janela"), key="toggle_janela", use_container_width=True):
        set_janela(c, not janela); st.rerun()
    st.caption("Com a janela fechada, os gestores não conseguem enviar nem editar justificativas. Você abre e fecha quando quiser (a aprovação/devolução pela controladoria continua funcionando).")
    st.divider()

    # ----- Exportar relatório (.xlsx) -----
    with st.expander("📄 Exportar relatório de justificativas (.xlsx)"):
        escopo = st.radio("Período", [f"Mês selecionado ({MESES[mes]}/{ano})", f"Ano inteiro ({ano})"],
                          horizontal=True, key="rel_escopo")
        if st.button("Gerar relatório", key="rel_gen"):
            ano_todo = escopo.startswith("Ano")
            js_rel = carregar_justificativas_ano(ano) if ano_todo else carregar_justificativas(ano, mes)
            dados, nlin = relatorio_justificativas_xlsx(js_rel, df_orc, cg)
            st.session_state["rel_bytes"] = dados
            st.session_state["rel_nome"] = f"justificativas_{ano}.xlsx" if ano_todo else f"justificativas_{mes:02d}_{ano}.xlsx"
            st.session_state["rel_n"] = nlin
        if st.session_state.get("rel_bytes") is not None:
            if st.session_state.get("rel_n", 0) == 0:
                st.caption("Nenhuma justificativa inserida no período gerado.")
            else:
                st.download_button(f"⬇️ Baixar {st.session_state['rel_nome']} ({st.session_state['rel_n']} linha(s))",
                                   data=st.session_state["rel_bytes"], file_name=st.session_state["rel_nome"],
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="rel_dl")

    if mes < get_cobranca(c):
        st.info(f"{MESES[mes]}/{ano} não está sujeito à cobrança de justificativa.")
        return

    js = carregar_justificativas(ano, mes)
    isentas = get_isentas(c)

    # ----- Pendências por empresa (mesma régua do gestor: por unidade + CR + conta) -----
    dfo = df_orc[df_orc["mes"] == mes] if not df_orc.empty else df_orc
    enviadas = {(int(j["uni_cod"]), int(j["cr_cod"]), int(j["conta_cod"])) for j in js if j.get("status") in ("JUSTIFICADO", "EM_REVISAO", "APROVADO")}
    devolvidas = {(int(j["uni_cod"]), int(j["cr_cod"]), int(j["conta_cod"])) for j in js if j.get("status") == "DEVOLVIDO"}
    pend = {}
    for _, v in dfo.iterrows():
        if int(v["conta_cod"]) in isentas:
            continue
        raw, pct = var_de(v["valor_planejado"], v["valor_realizado"])
        lab, _ = classifica(raw, pct, v["conta_cod"], banda)
        if lab != "Desfavorável":
            continue
        chave = (int(v["uni_cod"]), int(v["cr_cod"]), int(v["conta_cod"]))
        if chave in enviadas:
            continue
        gestor = cg.get((int(v["uni_cod"]), int(v["cr_cod"])), ("Sem gestor", ""))[0]
        d = pend.setdefault(gestor, {"n": 0, "valor": 0.0, "devolv": 0, "itens": []})
        d["n"] += 1; d["valor"] += raw
        dev = chave in devolvidas
        if dev: d["devolv"] += 1
        d["itens"].append({"cr": v.get("cr_nome", ""), "conta": f"{v['conta_cod']} · {v.get('conta_desc','')} ({v.get('unidade','')})", "raw": raw, "dev": dev})

    st.markdown("###### Pendências por gestor (desvios desfavoráveis ainda não justificados)")
    if not pend:
        st.success("Nenhuma pendência neste mês — todos os desvios desfavoráveis foram justificados.")
    else:
        total = sum(d["valor"] for d in pend.values())
        st.caption(f"Total: {sum(d['n'] for d in pend.values())} conta(s) · {brl(total)} de desvio parado")
        for gestor in sorted(pend, key=lambda g: -pend[g]["valor"]):
            d = pend[gestor]
            titulo = f"{gestor} — {d['n']} conta(s) · {brl(d['valor'])}" + (f" · {d['devolv']} devolvida(s)" if d["devolv"] else "")
            with st.expander(titulo):
                linhas = ""
                for it in sorted(d["itens"], key=lambda x: -x["raw"]):
                    tag = " (devolvida)" if it["dev"] else ""
                    linhas += (f"<tr><td>{it['cr']}</td><td style='text-align:center'>{it['conta']}{tag}</td>"
                               f"<td style='text-align:center; color:{VERMELHO}'>{brl(it['raw'])}</td></tr>")
                st.markdown(f"""<table class="lle"><tr><th style='text-align:left'>Centro de resultado</th>
                    <th style='text-align:center'>Conta</th>
                    <th style='text-align:center'>Desvio (R$)</th></tr>{linhas}</table>""", unsafe_allow_html=True)
    st.divider()

    if not js:
        st.info(f"Ainda não há justificativas enviadas em {MESES[mes]}/{ano}. As pendências acima mostram o que falta.")
        return

    desc_map = dict(zip(df_orc["conta_cod"].astype(int), df_orc["conta_desc"])) if not df_orc.empty else {}
    linhas = []
    for j in js:
        gestor, cr_nome = cg.get((j["uni_cod"], j["cr_cod"]), ("—", str(j["cr_cod"])))
        linhas.append({"gestor": gestor, "cr": cr_nome, "cr_cod": int(j["cr_cod"]), "uni_cod": int(j["uni_cod"]),
                       "conta": j["conta_cod"], "conta_desc": desc_map.get(int(j["conta_cod"]), ""),
                       "status": j.get("status", "PENDENTE"), "texto": j.get("texto", "") or "",
                       "comentario": j.get("comentario_controladoria", "") or "",
                       "_k": (j["uni_cod"], j["cr_cod"], j["conta_cod"])})
    df = pd.DataFrame(linhas)

    st.markdown("###### Resumo por gestor")
    resumo = ""
    for gestor in sorted(df["gestor"].unique()):
        sub = df[df["gestor"] == gestor]
        cnt = {s: int((sub["status"] == s).sum()) for s in STATUS_LABEL}
        resumo += (f"<tr><td style='text-align:center'>{gestor}</td><td style='text-align:center'>{len(sub)}</td>"
                   f"<td style='text-align:center; color:{VERMELHO}'>{cnt['PENDENTE']+cnt['DEVOLVIDO']}</td>"
                   f"<td style='text-align:center; color:{AZUL_CORP}'>{cnt['JUSTIFICADO']+cnt['EM_REVISAO']}</td>"
                   f"<td style='text-align:center; color:{VERDE}'>{cnt['APROVADO']}</td></tr>")
    st.markdown(f"""<table class="lle"><tr><th style='text-align:center'>Gestor</th><th style='text-align:center'>Total</th>
        <th style='text-align:center'>A responder</th><th style='text-align:center'>Aguardando controladoria</th><th style='text-align:center'>Aprovadas</th></tr>{resumo}</table>""", unsafe_allow_html=True)

    st.markdown("###### Detalhe por gestor e centro de resultado")
    st_cor = {"APROVADO": VERDE, "DEVOLVIDO": VERMELHO, "PENDENTE": CINZA_TXT, "JUSTIFICADO": AZUL_CORP, "EM_REVISAO": AZUL_CORP}
    for gestor in sorted(df["gestor"].unique()):
        sub = df[df["gestor"] == gestor]
        with st.expander(f"{gestor} — {len(sub)} justificativa(s)"):
            for cr in sorted(sub["cr"].unique()):
                grp = sub[sub["cr"] == cr]
                crc = int(grp["cr_cod"].iloc[0])
                st.markdown(f"**{cr} · {crc}**")
                linhas2 = ""
                for _, r in grp.iterrows():
                    st_lab = STATUS_LABEL.get(r["status"])
                    cor = st_cor.get(r["status"], AZUL_CORP)
                    txt = (r["texto"][:120] + "…") if len(r["texto"]) > 120 else (r["texto"] or "—")
                    _emp = "PISA" if r["uni_cod"] == 1 else "KING" if r["uni_cod"] == 2 else f"Emp {r['uni_cod']}"
                    conta_disp = (f"{r['conta']} · {r['conta_desc']} ({_emp})" if r["conta_desc"] else f"{r['conta']} ({_emp})")
                    linhas2 += (f"<tr><td style='text-align:center'>{conta_disp}</td>"
                                f"<td style='text-align:center'>{chip(st_lab, cor)}</td>"
                                f"<td style='text-align:left'>{txt}</td></tr>")
                st.markdown(f"""<table class="lle"><tr><th style='text-align:center'>Conta</th>
                    <th style='text-align:center'>Situação</th>
                    <th style='text-align:left'>Justificativa</th></tr>{linhas2}</table>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- edição manual do orçado
# ---------------------------------------------------------------- edição do orçamento (fragmento isolado)
@fragment
def _edicao_orcado_frag(view, ano, mes, uni_sel, cr_sel, conta_sel, c, prof):
    view = view.sort_values(["uni_cod", "cr_cod", "conta_cod"]).reset_index(drop=True)
    keys = [(int(r.ano), int(r.mes), int(r.uni_cod), int(r.cr_cod), int(r.conta_cod)) for r in view.itertuples()]
    orig_o = [round(float(r.valor_planejado or 0), 2) for r in view.itertuples()]
    orig_r = [round(float(r.valor_realizado or 0), 2) for r in view.itertuples()]
    disp = pd.DataFrame({
        "Unidade": [r.unidade for r in view.itertuples()],
        "Centro de resultado": [r.cr_nome for r in view.itertuples()],
        "Conta": [f"{int(r.conta_cod)} · {r.conta_desc}" for r in view.itertuples()],
        "Orçado": orig_o, "Realizado": orig_r,
    })
    st.caption(f"{len(disp)} conta(s) em {MESES[mes]}/{ano}. A edição roda isolada — o app não recarrega a cada célula. Edite **Orçado** e/ou **Realizado** e salve.")
    edited = st.data_editor(
        disp, key=f"edo_grid_{mes}_{uni_sel}_{cr_sel}_{conta_sel}", hide_index=True, use_container_width=True,
        num_rows="fixed", disabled=["Unidade", "Centro de resultado", "Conta"],
        column_config={"Orçado": st.column_config.NumberColumn(format="%.0f", step=1),
                       "Realizado": st.column_config.NumberColumn(format="%.0f", step=1)})
    if st.button("Salvar alterações", key="edo_save", type="primary"):
        novos_o = list(edited["Orçado"]); novos_r = list(edited["Realizado"]); mudou = 0
        for i, kkey in enumerate(keys):
            an, me, uni, cr, ct = kkey
            match = dict(ano=an, mes=me, uni_cod=uni, cr_cod=cr, conta_cod=ct)
            for coluna, novos, orig in (("valor_planejado", novos_o, orig_o), ("valor_realizado", novos_r, orig_r)):
                try: nv = round(float(novos[i]))
                except (TypeError, ValueError): continue
                if abs(nv - orig[i]) > 0.005:
                    c.table("orc_realizado").update({coluna: nv}).match(match).execute()
                    c.table("orc_log").insert(dict(**match, campo=coluna, valor_antigo=orig[i],
                        valor_novo=nv, alterado_por=prof.get("nome", ""))).execute()
                    mudou += 1
        if mudou:
            limpar_cache(); st.success(f"{mudou} alteração(ões) salva(s) e registrada(s) no log."); st.rerun()
        else:
            st.info("Nenhuma alteração detectada.")

def tela_editar_orcado(c, prof, df_orc, ano, mes):
    st.markdown("<div class='modtag'>Manutenção do Orçamento</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Edite orçado e/ou realizado linha a linha — ideal para reclassificações no período, sem reimportar o arquivo. Toda alteração fica registrada no log.</div>", unsafe_allow_html=True)
    if df_orc is None or df_orc.empty:
        st.info("Base de orçado ainda não carregada. Importe na aba **Importar dados**.")
        return

    fc = st.columns([1.2, 1.6, 1.9])
    unis = sorted({(int(r["uni_cod"]), r.get("unidade", "")) for _, r in df_orc.iterrows()})
    uni_sel = fc[0].selectbox("Unidade", [0] + [u[0] for u in unis],
                              format_func=lambda x: "Todas" if x == 0 else next((n for u, n in unis if u == x), str(x)), key="edo_uni")
    dcr = df_orc if not uni_sel else df_orc[df_orc["uni_cod"] == uni_sel]
    crs = sorted({(int(r["cr_cod"]), r.get("cr_nome", "")) for _, r in dcr.iterrows()})
    cr_sel = fc[1].selectbox("Centro de resultado", [0] + [x[0] for x in crs],
                             format_func=lambda x: "Todos" if x == 0 else f"{x} · {next((n for cc, n in crs if cc == x), '')}", key="edo_cr")

    # contas disponíveis para os filtros de unidade/CR no mês (o seletor evita editar a linha errada)
    base_ct = df_orc[df_orc["mes"] == mes]
    if uni_sel: base_ct = base_ct[base_ct["uni_cod"] == uni_sel]
    if cr_sel: base_ct = base_ct[base_ct["cr_cod"] == cr_sel]
    contas_opt = sorted({(int(r["conta_cod"]), r.get("conta_desc", "")) for _, r in base_ct.iterrows()})
    conta_sel = fc[2].selectbox("Conta", [0] + [x[0] for x in contas_opt],
                                format_func=lambda x: "Todas" if x == 0 else f"{x} · {next((n for cc, n in contas_opt if cc == x), '')}", key="edo_conta")

    view = df_orc[df_orc["mes"] == mes]
    if uni_sel: view = view[view["uni_cod"] == uni_sel]
    if cr_sel: view = view[view["cr_cod"] == cr_sel]
    if conta_sel: view = view[view["conta_cod"] == conta_sel]
    if view.empty:
        st.info("Nenhuma conta para os filtros selecionados.")
    else:
        _edicao_orcado_frag(view, ano, mes, uni_sel, cr_sel, conta_sel, c, prof)

    # ---------- histórico de alterações ----------
    st.divider()
    st.markdown("#### Histórico de alterações")
    log = carregar_orc_log(300)
    if not log:
        st.caption("Nenhuma alteração registrada ainda.")
        return
    nome = {}
    for _, r in df_orc.iterrows():
        nome[(int(r["uni_cod"]), int(r["cr_cod"]), int(r["conta_cod"]))] = (r.get("unidade", ""), r.get("cr_nome", ""), r.get("conta_desc", ""))
    _n = hist_n(len(log), "hist_n_orcado")
    linhas = ""
    for g in log[:_n]:
        uni = int(g.get("uni_cod") or 0); cr = int(g.get("cr_cod") or 0); ct = int(g.get("conta_cod") or 0)
        un, crn, cd = nome.get((uni, cr, ct), ("", "", ""))
        quando = str(g.get("alterado_em", "") or "")[:16].replace("T", " ")
        va = float(g.get("valor_antigo") or 0); vn = float(g.get("valor_novo") or 0)
        seta = VERMELHO if vn > va else VERDE
        campo_lab = "Orçado" if g.get("campo") == "valor_planejado" else "Realizado"
        linhas += (f"<tr><td style='text-align:left'>{quando}</td><td style='text-align:left'>{g.get('alterado_por','') or '—'}</td>"
                   f"<td style='text-align:center'>{int(g.get('ano') or 0)}</td>"
                   f"<td style='text-align:left'>{MESES[int(g.get('mes') or 1)]}</td><td style='text-align:left'>{un or uni}</td>"
                   f"<td style='text-align:left'>{cr} · {crn}</td><td style='text-align:left'>{ct} · {cd}</td>"
                   f"<td style='text-align:center'>{campo_lab}</td><td>{brl(va)}</td><td style='color:{seta}'>{brl(vn)}</td></tr>")
    st.markdown(f"""<table class="lle"><tr>
        <th style='text-align:left'>Quando</th><th style='text-align:left'>Quem</th><th style='text-align:center'>Ano</th><th style='text-align:left'>Mês</th>
        <th style='text-align:left'>Unidade</th><th style='text-align:left'>Centro de resultado</th>
        <th style='text-align:left'>Conta</th><th style='text-align:center'>Campo</th><th>De</th><th>Para</th></tr>{linhas}</table>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- registro mensal (fragmento isolado)
@fragment
def _registro_mensal_frag(tabela, tabela_log, log_loader, dfr, emp, ano, c, prof, rotulo, custo=False):
    """Grade editável Planejado/Realizado por empresa e mês, com log. Roda isolada (fragment)."""
    EMP = {1: "PISA", 2: "KING"}
    st.divider()
    st.markdown(f"#### Registrar / editar {rotulo}")
    st.caption("Preencha Planejado e Realizado por empresa e mês. A edição roda isolada — o app não recarrega a cada célula. Toda alteração fica no log.")
    emps = [1, 2] if not emp else [emp]
    idx = {(int(r["mes"]), int(r["uni_cod"])): r for _, r in dfr.iterrows()}
    # histórico: realizado do ano anterior (referência), da própria tabela
    try:
        prev = c.table(tabela).select("mes,uni_cod,valor_realizado").eq("ano", ano - 1).execute().data or []
    except Exception:
        prev = []
    hist_idx = {}
    for r in prev:
        k = (int(r.get("mes", 0) or 0), int(r.get("uni_cod", 0) or 0))
        hist_idx[k] = hist_idx.get(k, 0.0) + float(r.get("valor_realizado") or 0)
    hcol = f"Hist. {ano-1}"
    hist_on = st.checkbox(f"Mostrar {hcol} (mês a mês)", value=False, key=f"{tabela}_hist_{emp}",
                          help="Coluna de referência com o realizado do mesmo mês no ano anterior.")
    keys, disp_mes, disp_emp, ohist, oplan, oreal = [], [], [], [], [], []
    for u in emps:
        for m in range(1, 13):
            r = idx.get((m, u))
            keys.append((m, u)); disp_mes.append(MESES[m]); disp_emp.append(EMP[u])
            ohist.append(round(hist_idx.get((m, u), 0.0), 2))
            oplan.append(round(float(r["valor_planejado"]) if r is not None else 0.0, 2))
            oreal.append(round(float(r["valor_realizado"]) if r is not None else 0.0, 2))
    cols = {"Mês": disp_mes, "Empresa": disp_emp}
    if hist_on: cols[hcol] = ohist
    cols["Planejado"] = oplan; cols["Realizado"] = oreal
    dedit = pd.DataFrame(cols)
    ccfg = {"Planejado": st.column_config.NumberColumn(format="%.0f", step=1),
            "Realizado": st.column_config.NumberColumn(format="%.0f", step=1)}
    disc = ["Mês", "Empresa"]
    if hist_on:
        ccfg[hcol] = st.column_config.NumberColumn(f"{hcol} (R$)", format="%.0f", help="Realizado do ano anterior — referência, não editável")
        disc.append(hcol)
    ed = st.data_editor(dedit, key=f"{tabela}_grid_{emp}", hide_index=True, use_container_width=True,
                        num_rows="fixed", disabled=disc, column_config=ccfg)
    if st.button(f"Salvar {rotulo}", key=f"{tabela}_save", type="primary"):
        np_, nr_ = list(ed["Planejado"]), list(ed["Realizado"]); mudou = 0
        for i, (m, u) in enumerate(keys):
            try: p_novo = round(float(np_[i])); r_novo = round(float(nr_[i]))
            except (TypeError, ValueError): continue
            mud_p = abs(p_novo - oplan[i]) > 0.005; mud_r = abs(r_novo - oreal[i]) > 0.005
            if not (mud_p or mud_r): continue
            c.table(tabela).upsert(dict(ano=ano, mes=m, uni_cod=u, unidade=EMP[u],
                valor_planejado=p_novo, valor_realizado=r_novo, atualizado_por=prof.get("nome", "")),
                on_conflict="ano,mes,uni_cod").execute()
            if mud_p:
                c.table(tabela_log).insert(dict(ano=ano, mes=m, uni_cod=u, campo="valor_planejado",
                    valor_antigo=oplan[i], valor_novo=p_novo, alterado_por=prof.get("nome", ""))).execute(); mudou += 1
            if mud_r:
                c.table(tabela_log).insert(dict(ano=ano, mes=m, uni_cod=u, campo="valor_realizado",
                    valor_antigo=oreal[i], valor_novo=r_novo, alterado_por=prof.get("nome", ""))).execute(); mudou += 1
        if mudou:
            limpar_cache(); st.success(f"{mudou} alteração(ões) salva(s) e registrada(s) no log."); st.rerun()
        else:
            st.info("Nenhuma alteração detectada.")
    log = log_loader(300)
    if log:
        st.markdown(f"###### Histórico de alterações de {rotulo}")
        CL = {"valor_planejado": "Planejado", "valor_realizado": "Realizado"}
        _n = hist_n(len(log), f"hist_n_{tabela}")
        ln = ""
        for glog in log[:_n]:
            quando = str(glog.get("alterado_em", "") or "")[:16].replace("T", " ")
            va = float(glog.get("valor_antigo") or 0); vn = float(glog.get("valor_novo") or 0)
            seta = (VERMELHO if vn > va else VERDE) if custo else (VERDE if vn > va else VERMELHO)
            ln += (f"<tr><td style='text-align:left'>{quando}</td><td style='text-align:left'>{glog.get('alterado_por','') or '\u2014'}</td>"
                   f"<td style='text-align:center'>{int(glog.get('ano') or 0)}</td>"
                   f"<td style='text-align:left'>{MESES[int(glog.get('mes') or 1)]}</td>"
                   f"<td style='text-align:center'>{EMP.get(int(glog.get('uni_cod') or 0), glog.get('uni_cod'))}</td>"
                   f"<td style='text-align:center'>{CL.get(glog.get('campo'), glog.get('campo'))}</td>"
                   f"<td>{brl(va)}</td><td style='color:{seta}'>{brl(vn)}</td></tr>")
        st.markdown(f"""<table class="lle"><tr>
            <th style='text-align:left'>Quando</th><th style='text-align:left'>Quem</th><th style='text-align:center'>Ano</th><th style='text-align:left'>Mês</th>
            <th style='text-align:center'>Empresa</th><th style='text-align:center'>Campo</th><th>De</th><th>Para</th></tr>{ln}</table>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- receita bruta de vendas
def tela_receita(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    banda = get_faixa(c)
    st.markdown("<div class='modtag'>Receita Bruta de Vendas</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Planejado x realizado por empresa e mês. Convenção de receita: realizado acima do planejado é favorável.</div>", unsafe_allow_html=True)

    rows = carregar_receita(ano)
    dfr = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ano", "mes", "uni_cod", "unidade", "valor_planejado", "valor_realizado"])
    for col in ("valor_planejado", "valor_realizado"):
        if col not in dfr.columns: dfr[col] = 0.0

    emp = st.selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="rec_emp")
    scope = dfr if not emp else dfr[dfr["uni_cod"] == emp]

    if scope.empty:
        g = pd.DataFrame(0.0, index=range(1, 13), columns=["valor_planejado", "valor_realizado"])
    else:
        g = scope.groupby("mes")[["valor_planejado", "valor_realizado"]].sum().reindex(range(1, 13), fill_value=0.0)
    tp = float(g["valor_planejado"].sum()); tr = float(g["valor_realizado"].sum())
    mask_real = g["valor_realizado"] != 0
    tp_real = float(g.loc[mask_real, "valor_planejado"].sum())
    var = tr - tp_real; pct = (var / tp_real * 100) if tp_real else 0.0  # variação só sobre meses com realizado

    def cor_receita(v, pl):
        p = (v / pl * 100) if pl else 0.0
        if pl and abs(p) <= banda: return CINZA_TXT
        return VERDE if v >= 0 else VERMELHO

    # KPIs
    k = st.columns(4)
    kpi = [("Planejado (ano)", brl(tp), CINZA_TXT), ("Realizado (ano)", brl(tr), CINZA_TXT),
           ("Variação (R$)", brl(var), cor_receita(var, tp_real)), ("Variação (%)", pct_txt(pct), cor_receita(var, tp_real))]
    for col, (t, v, cr) in zip(k, kpi):
        col.markdown(f"<div class='card' style='text-align:center'><div style='font-size:.8rem;color:{CINZA_TXT}'>{t}</div>"
                     f"<div style='font-size:1.4rem;font-weight:700;color:{cr}'>{v}</div></div>", unsafe_allow_html=True)

    # evolução mensal
    st.caption(f"Variação calculada apenas sobre os meses com realizado (planejado desses meses: {brl(tp_real)}). Orçado cheio do ano: {brl(tp)}.")
    st.markdown("#### Evolução mensal")
    linhas = ""
    for m in range(1, 13):
        vp = float(g.loc[m, "valor_planejado"]); vr = float(g.loc[m, "valor_realizado"])
        v = vr - vp; p = (v / vp * 100) if vp else 0.0
        if vr == 0 and vp == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem dados", CINZA_TXT); cr = CINZA_TXT
        elif vr == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem realizado", CINZA_TXT); cr = CINZA_TXT
        else:
            cr = cor_receita(v, vp)
            lab = "Neutro" if cr == CINZA_TXT else ("Favorável" if v >= 0 else "Desfavorável")
            vr_txt = brl(vr); var_txt = brl(v); pctv = pct_txt(p); status = chip(lab, cr)
        linhas += (f"<tr><td style='text-align:left'>{MESES[m]}</td><td>{brl(vp)}</td><td>{vr_txt}</td>"
                   f"<td style='color:{cr}'>{var_txt}</td><td style='color:{cr}'>{pctv}</td><td>{status}</td></tr>")
    tcor = cor_receita(var, tp_real)
    total = (f"<tr class='mark'><td style='text-align:left'><b>Total</b></td><td><b>{brl(tp)}</b></td>"
             f"<td><b>{brl(tr) if tr else '\u2014'}</b></td><td style='color:{tcor}'><b>{brl(var) if tr else '\u2014'}</b></td>"
             f"<td style='color:{tcor}'><b>{pct_txt(pct) if tr else '\u2014'}</b></td><td></td></tr>")
    st.markdown(f"""<div class='scroll'><table class="lle"><tr>
        <th style='text-align:left'>Mês</th><th>Planejado</th><th>Realizado</th>
        <th>Var. (R$)</th><th>Var. (%)</th><th>Status</th></tr>{linhas}{total}</table></div>""", unsafe_allow_html=True)

    _registro_mensal_frag("receita_venda", "receita_log", carregar_receita_log, dfr, emp, ano, c, prof, "receita")

# ---------------------------------------------------------------- CMV (custo da mercadoria vendida)
def tela_cmv(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    banda = get_faixa(c)
    st.markdown("<div class='modtag'>CMV — Custo da Mercadoria Vendida</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Planejado x realizado por empresa e mês. Convenção de custo: realizado abaixo do planejado é favorável (gastou menos).</div>", unsafe_allow_html=True)

    rows = carregar_cmv(ano)
    dfr = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ano", "mes", "uni_cod", "unidade", "valor_planejado", "valor_realizado"])
    for col in ("valor_planejado", "valor_realizado"):
        if col not in dfr.columns: dfr[col] = 0.0

    emp = st.selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="cmv_emp")
    scope = dfr if not emp else dfr[dfr["uni_cod"] == emp]

    if scope.empty:
        g = pd.DataFrame(0.0, index=range(1, 13), columns=["valor_planejado", "valor_realizado"])
    else:
        g = scope.groupby("mes")[["valor_planejado", "valor_realizado"]].sum().reindex(range(1, 13), fill_value=0.0)
    tp = float(g["valor_planejado"].sum()); tr = float(g["valor_realizado"].sum())
    mask_real = g["valor_realizado"] != 0
    tp_real = float(g.loc[mask_real, "valor_planejado"].sum())
    var = tr - tp_real; pct = (var / tp_real * 100) if tp_real else 0.0  # variação só sobre meses com realizado

    def cor_cmv(v, pl):  # custo: gastar menos que o previsto (v<=0) é favorável
        p = (v / pl * 100) if pl else 0.0
        if pl and abs(p) <= banda: return CINZA_TXT
        return VERDE if v <= 0 else VERMELHO

    # KPIs
    k = st.columns(4)
    kpi = [("Planejado (ano)", brl(tp), CINZA_TXT), ("Realizado (ano)", brl(tr), CINZA_TXT),
           ("Variação (R$)", brl(var), cor_cmv(var, tp_real)), ("Variação (%)", pct_txt(pct), cor_cmv(var, tp_real))]
    for col, (t, v, cr) in zip(k, kpi):
        col.markdown(f"<div class='card' style='text-align:center'><div style='font-size:.8rem;color:{CINZA_TXT}'>{t}</div>"
                     f"<div style='font-size:1.4rem;font-weight:700;color:{cr}'>{v}</div></div>", unsafe_allow_html=True)

    # evolução mensal
    st.caption(f"Variação calculada apenas sobre os meses com realizado (planejado desses meses: {brl(tp_real)}). Orçado cheio do ano: {brl(tp)}.")
    st.markdown("#### Evolução mensal")
    linhas = ""
    for m in range(1, 13):
        vp = float(g.loc[m, "valor_planejado"]); vr = float(g.loc[m, "valor_realizado"])
        v = vr - vp; p = (v / vp * 100) if vp else 0.0
        if vr == 0 and vp == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem dados", CINZA_TXT); cr = CINZA_TXT
        elif vr == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem realizado", CINZA_TXT); cr = CINZA_TXT
        else:
            cr = cor_cmv(v, vp)
            lab = "Neutro" if cr == CINZA_TXT else ("Favorável" if v <= 0 else "Desfavorável")
            vr_txt = brl(vr); var_txt = brl(v); pctv = pct_txt(p); status = chip(lab, cr)
        linhas += (f"<tr><td style='text-align:left'>{MESES[m]}</td><td>{brl(vp)}</td><td>{vr_txt}</td>"
                   f"<td style='color:{cr}'>{var_txt}</td><td style='color:{cr}'>{pctv}</td><td>{status}</td></tr>")
    tcor = cor_cmv(var, tp_real)
    total = (f"<tr class='mark'><td style='text-align:left'><b>Total</b></td><td><b>{brl(tp)}</b></td>"
             f"<td><b>{brl(tr) if tr else '\u2014'}</b></td><td style='color:{tcor}'><b>{brl(var) if tr else '\u2014'}</b></td>"
             f"<td style='color:{tcor}'><b>{pct_txt(pct) if tr else '\u2014'}</b></td><td></td></tr>")
    st.markdown(f"""<div class='scroll'><table class="lle"><tr>
        <th style='text-align:left'>Mês</th><th>Planejado</th><th>Realizado</th>
        <th>Var. (R$)</th><th>Var. (%)</th><th>Status</th></tr>{linhas}{total}</table></div>""", unsafe_allow_html=True)

    _registro_mensal_frag("cmv_valor", "cmv_log", carregar_cmv_log, dfr, emp, ano, c, prof, "CMV", custo=True)

# ---------------------------------------------------------------- deduções de vendas
@fragment
def _registro_deducao_frag(dfr, emp, conta, ano, c, prof):
    """Grade editável Planejado/Realizado da dedução escolhida, por empresa e mês. Roda isolada."""
    EMP = {1: "PISA", 2: "KING"}
    st.divider()
    st.markdown(f"#### Registrar / editar — {conta}")
    st.caption("Preencha Planejado e Realizado por empresa e mês desta dedução. A edição roda isolada — o app não recarrega a cada célula. Toda alteração fica no log.")
    hcol = f"Hist. {ano-1}"
    hist_on = st.checkbox(f"Mostrar {hcol} (mês a mês)", value=False, key=f"ded_hist_{conta}_{emp}",
                          help="Realizado do mesmo mês no ano anterior, para esta dedução — referência.")
    hist_idx = {}
    if hist_on:
        try:
            prev = c.table("deducao_valor").select("mes,uni_cod,valor_realizado").eq("ano", ano - 1).eq("conta", conta).execute().data or []
        except Exception:
            prev = []
        for r in prev:
            k = (int(r.get("mes", 0) or 0), int(r.get("uni_cod", 0) or 0))
            hist_idx[k] = hist_idx.get(k, 0.0) + float(r.get("valor_realizado") or 0)
    emps = [1, 2] if not emp else [emp]
    sub = dfr[dfr["conta"] == conta] if "conta" in dfr.columns else dfr.iloc[0:0]
    idx = {(int(r["mes"]), int(r["uni_cod"])): r for _, r in sub.iterrows()}
    keys, disp_mes, disp_emp, ohist, oplan, oreal = [], [], [], [], [], []
    for u in emps:
        for m in range(1, 13):
            r = idx.get((m, u))
            keys.append((m, u)); disp_mes.append(MESES[m]); disp_emp.append(EMP[u])
            ohist.append(round(hist_idx.get((m, u), 0.0), 2))
            oplan.append(round(float(r["valor_planejado"]) if r is not None else 0.0, 2))
            oreal.append(round(float(r["valor_realizado"]) if r is not None else 0.0, 2))
    cols = {"Mês": disp_mes, "Empresa": disp_emp}
    if hist_on: cols[hcol] = ohist
    cols["Planejado"] = oplan; cols["Realizado"] = oreal
    dedit = pd.DataFrame(cols)
    ccfg = {"Planejado": st.column_config.NumberColumn(format="%.0f", step=1),
            "Realizado": st.column_config.NumberColumn(format="%.0f", step=1)}
    disc = ["Mês", "Empresa"]
    if hist_on:
        ccfg[hcol] = st.column_config.NumberColumn(f"{hcol} (R$)", format="%.0f", help="Realizado do ano anterior — referência, não editável")
        disc.append(hcol)
    ed = st.data_editor(dedit, key=f"ded_grid_{conta}_{emp}", hide_index=True, use_container_width=True,
                        num_rows="fixed", disabled=disc, column_config=ccfg)
    if st.button("Salvar deduções", key="ded_save", type="primary"):
        np_, nr_ = list(ed["Planejado"]), list(ed["Realizado"]); mudou = 0
        for i, (m, u) in enumerate(keys):
            try: p_novo = round(float(np_[i])); r_novo = round(float(nr_[i]))
            except (TypeError, ValueError): continue
            mud_p = abs(p_novo - oplan[i]) > 0.005; mud_r = abs(r_novo - oreal[i]) > 0.005
            if not (mud_p or mud_r): continue
            c.table("deducao_valor").upsert(dict(ano=ano, mes=m, uni_cod=u, unidade=EMP[u], conta=conta,
                valor_planejado=p_novo, valor_realizado=r_novo, atualizado_por=prof.get("nome", "")),
                on_conflict="ano,mes,uni_cod,conta").execute()
            if mud_p:
                c.table("deducao_log").insert(dict(ano=ano, mes=m, uni_cod=u, conta=conta, campo="valor_planejado",
                    valor_antigo=oplan[i], valor_novo=p_novo, alterado_por=prof.get("nome", ""))).execute(); mudou += 1
            if mud_r:
                c.table("deducao_log").insert(dict(ano=ano, mes=m, uni_cod=u, conta=conta, campo="valor_realizado",
                    valor_antigo=oreal[i], valor_novo=r_novo, alterado_por=prof.get("nome", ""))).execute(); mudou += 1
        if mudou:
            limpar_cache(); st.success(f"{mudou} alteração(ões) salva(s) e registrada(s) no log."); st.rerun()
        else:
            st.info("Nenhuma alteração detectada.")

def tela_deducao(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    banda = get_faixa(c)
    st.markdown("<div class='modtag'>Deduções de Vendas</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Planejado x realizado por empresa, mês e dedução (devoluções e impostos sobre vendas). Convenção: realizado abaixo do planejado é favorável (deduziu menos).</div>", unsafe_allow_html=True)

    rows = carregar_deducao(ano)
    dfr = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ano", "mes", "uni_cod", "unidade", "conta", "valor_planejado", "valor_realizado"])
    for col in ("valor_planejado", "valor_realizado"):
        if col not in dfr.columns: dfr[col] = 0.0
    if "conta" not in dfr.columns: dfr["conta"] = ""

    fcol = st.columns([1.3, 2.2])
    emp = fcol[0].selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="ded_emp")
    conta = fcol[1].selectbox("Dedução", ["Todas"] + DEDUCOES, key="ded_conta")

    scope = dfr if not emp else dfr[dfr["uni_cod"] == emp]
    if conta != "Todas": scope = scope[scope["conta"] == conta]

    if scope.empty:
        g = pd.DataFrame(0.0, index=range(1, 13), columns=["valor_planejado", "valor_realizado"])
    else:
        g = scope.groupby("mes")[["valor_planejado", "valor_realizado"]].sum().reindex(range(1, 13), fill_value=0.0)
    tp = float(g["valor_planejado"].sum()); tr = float(g["valor_realizado"].sum())
    mask_real = g["valor_realizado"] != 0
    tp_real = float(g.loc[mask_real, "valor_planejado"].sum())
    var = tr - tp_real; pct = (var / tp_real * 100) if tp_real else 0.0  # variação só sobre meses com realizado

    def cor_ded(v, pl):
        p = (v / pl * 100) if pl else 0.0
        if pl and abs(p) <= banda: return CINZA_TXT
        return VERDE if v <= 0 else VERMELHO

    escopo_lbl = "todas as deduções" if conta == "Todas" else conta
    k = st.columns(4)
    kpi = [(f"Planejado (ano)", brl(tp), CINZA_TXT), ("Realizado (ano)", brl(tr), CINZA_TXT),
           ("Variação (R$)", brl(var), cor_ded(var, tp_real)), ("Variação (%)", pct_txt(pct), cor_ded(var, tp_real))]
    for col, (t, v, cr) in zip(k, kpi):
        col.markdown(f"<div class='card' style='text-align:center'><div style='font-size:.8rem;color:{CINZA_TXT}'>{t}</div>"
                     f"<div style='font-size:1.4rem;font-weight:700;color:{cr}'>{v}</div></div>", unsafe_allow_html=True)
    st.caption(f"Exibindo: {escopo_lbl}.")

    st.caption(f"Variação calculada apenas sobre os meses com realizado (planejado desses meses: {brl(tp_real)}). Orçado cheio do ano: {brl(tp)}.")
    st.markdown("#### Evolução mensal")
    linhas = ""
    for m in range(1, 13):
        vp = float(g.loc[m, "valor_planejado"]); vr = float(g.loc[m, "valor_realizado"])
        v = vr - vp; p = (v / vp * 100) if vp else 0.0
        if vr == 0 and vp == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem dados", CINZA_TXT); cr = CINZA_TXT
        elif vr == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem realizado", CINZA_TXT); cr = CINZA_TXT
        else:
            cr = cor_ded(v, vp)
            lab = "Neutro" if cr == CINZA_TXT else ("Favorável" if v <= 0 else "Desfavorável")
            vr_txt = brl(vr); var_txt = brl(v); pctv = pct_txt(p); status = chip(lab, cr)
        linhas += (f"<tr><td style='text-align:left'>{MESES[m]}</td><td>{brl(vp)}</td><td>{vr_txt}</td>"
                   f"<td style='color:{cr}'>{var_txt}</td><td style='color:{cr}'>{pctv}</td><td>{status}</td></tr>")
    tcor = cor_ded(var, tp_real)
    total = (f"<tr class='mark'><td style='text-align:left'><b>Total</b></td><td><b>{brl(tp)}</b></td>"
             f"<td><b>{brl(tr) if tr else '\u2014'}</b></td><td style='color:{tcor}'><b>{brl(var) if tr else '\u2014'}</b></td>"
             f"<td style='color:{tcor}'><b>{pct_txt(pct) if tr else '\u2014'}</b></td><td></td></tr>")
    st.markdown(f"""<div class='scroll'><table class="lle"><tr>
        <th style='text-align:left'>Mês</th><th>Planejado</th><th>Realizado</th>
        <th>Var. (R$)</th><th>Var. (%)</th><th>Status</th></tr>{linhas}{total}</table></div>""", unsafe_allow_html=True)

    # registro por dedução escolhida
    if conta == "Todas":
        st.divider()
        st.info("Selecione uma **dedução específica** no seletor acima para registrar/editar seus valores por mês.")
    else:
        _registro_deducao_frag(dfr, emp, conta, ano, c, prof)

    # histórico (todas as deduções)
    log = carregar_deducao_log(300)
    if log:
        st.markdown("###### Histórico de alterações de deduções")
        CL = {"valor_planejado": "Planejado", "valor_realizado": "Realizado"}
        _n = hist_n(len(log), "hist_n_deducao")
        ln = ""
        for glog in log[:_n]:
            quando = str(glog.get("alterado_em", "") or "")[:16].replace("T", " ")
            va = float(glog.get("valor_antigo") or 0); vn = float(glog.get("valor_novo") or 0)
            seta = VERMELHO if vn > va else VERDE
            ln += (f"<tr><td style='text-align:left'>{quando}</td><td style='text-align:left'>{glog.get('alterado_por','') or '\u2014'}</td>"
                   f"<td style='text-align:center'>{int(glog.get('ano') or 0)}</td>"
                   f"<td style='text-align:left'>{MESES[int(glog.get('mes') or 1)]}</td>"
                   f"<td style='text-align:center'>{EMP.get(int(glog.get('uni_cod') or 0), glog.get('uni_cod'))}</td>"
                   f"<td style='text-align:left'>{glog.get('conta','') or '\u2014'}</td>"
                   f"<td style='text-align:center'>{CL.get(glog.get('campo'), glog.get('campo'))}</td>"
                   f"<td>{brl(va)}</td><td style='color:{seta}'>{brl(vn)}</td></tr>")
        st.markdown(f"""<table class="lle"><tr>
            <th style='text-align:left'>Quando</th><th style='text-align:left'>Quem</th><th style='text-align:center'>Ano</th><th style='text-align:left'>Mês</th>
            <th style='text-align:center'>Empresa</th><th style='text-align:left'>Dedução</th>
            <th style='text-align:center'>Campo</th><th>De</th><th>Para</th></tr>{ln}</table>""", unsafe_allow_html=True)

# ---------------------------------------------------------------- DRE consolidada
def dre_xlsx(linhas, base_r, ano, faixa, empresa):
    import io
    dados = []
    for nome, d, tipo, forte in linhas:
        p, r, a = d["p"], d["r"], d["a"]; vr = r - p; vp = (vr / p * 100) if p else 0.0
        av = (r / base_r * 100) if base_r else 0.0; ah = ((r - a) / a * 100) if a else 0.0
        dados.append({"Linha": nome, "Planejado": round(p, 2), "Realizado": round(r, 2),
                      f"Ano anterior ({ano-1})": round(a, 2), "Var (R$)": round(vr, 2),
                      "Var (%)": round(vp, 2), "AV %": round(av, 2), "AH %": round(ah, 2)})
    df = pd.DataFrame(dados)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as x:
        df.to_excel(x, index=False, sheet_name="DRE", startrow=2)
        ws = x.sheets["DRE"]
        ws["A1"] = f"DRE — {empresa} — {faixa}"
        try:
            from openpyxl.styles import Font
            ws["A1"].font = Font(bold=True, size=13)
            for cell in ws[3]:  # cabeçalho (linha 3, pois startrow=2)
                cell.font = Font(bold=True)
            ws.column_dimensions["A"].width = 42
            for col in "BCDEFGHI":
                ws.column_dimensions[col].width = 16
        except Exception:
            pass
    return buf.getvalue()

@fragment
def _mapa_dre_frag(contas, mapa, c, prof):
    """Grade para mapear contas do orçamento -> linha da DRE. Roda isolada (fragment)."""
    OPCOES = ["Fora da DRE"] + DRE_LINHAS_OPC
    st.caption("Marque quais contas do orçamento são despesa operacional e em qual grupo entram. "
               "Deixe **Fora da DRE** o que já vem dos módulos próprios (receita, deduções, CMV, pessoal) ou não é despesa. A edição roda isolada.")
    dm = pd.DataFrame({
        "Conta": [f"{cod} · {desc}" for cod, desc in contas],
        "Linha da DRE": [mapa.get(cod, "Fora da DRE") for cod, desc in contas],
    })
    ed = st.data_editor(dm, key="dre_mapa_grid", hide_index=True, use_container_width=True, num_rows="fixed",
                        disabled=["Conta"],
                        column_config={"Linha da DRE": st.column_config.SelectboxColumn(options=OPCOES)})
    if st.button("Salvar mapeamento", key="dre_mapa_save", type="primary"):
        novos = list(ed["Linha da DRE"]); mudou = 0
        for i, (cod, desc) in enumerate(contas):
            nova = str(novos[i]) if novos[i] in OPCOES else "Fora da DRE"
            atual = mapa.get(cod, "Fora da DRE")
            if nova != atual:
                c.table("dre_mapa").upsert(dict(conta_cod=int(cod), conta_desc=desc, linha=nova,
                    atualizado_por=prof.get("nome", "")), on_conflict="conta_cod").execute()
                mudou += 1
        if mudou:
            limpar_cache(); st.success(f"{mudou} conta(s) remapeada(s)."); st.rerun()
        else:
            st.info("Nenhuma alteração no mapeamento.")

@fragment
def tela_dre(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    banda = get_faixa(c)
    st.markdown("<div class='modtag'>DRE — Demonstrativo de Resultados</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Receita, Deduções, CMV e Pessoal vêm dos módulos dedicados; as demais despesas vêm do orçamento conforme o mapeamento de contas. Planejado x Realizado, com ano anterior, AV e AH.</div>", unsafe_allow_html=True)

    f = st.columns([1.1, 1.0, 1.0, 1.2, 1.2])
    emp = f[0].selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="dre_emp")
    de = f[1].selectbox("Mês inicial", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="dre_de")
    ate = f[2].selectbox("Mês final", list(range(1, 13)), index=11, format_func=lambda m: MESES[m], key="dre_ate")
    visao = f[3].radio("Visão", ["Mensal", "Acumulada"], horizontal=True, key="dre_visao")
    formato = f[4].radio("Formato", ["Consolidado", "Por unidade", "Por mês", "Comparar períodos"], horizontal=True, key="dre_formato")
    if ate < de: ate = de
    meses_range = list(range(de, ate + 1))
    faixa = f"{MESES[de]}–{MESES[ate]}/{ano}" + (" (acumulado)" if visao == "Acumulada" else "")

    # ----- pré-cálculo por mês (uma passada por tabela; barato e cacheado) -----
    def por_mes(loader, kp, kr):
        cur = loader(ano) or []; prev = loader(ano - 1) or []
        d = {m: {"p": 0.0, "r": 0.0, "a": 0.0} for m in range(1, 13)}
        for x in cur:
            m = int(x.get("mes", 0) or 0)
            if 1 <= m <= 12 and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                d[m]["p"] += float(x.get(kp) or 0); d[m]["r"] += float(x.get(kr) or 0)
        for x in prev:
            m = int(x.get("mes", 0) or 0)
            if 1 <= m <= 12 and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                d[m]["a"] += float(x.get(kr) or 0)
        return d
    rec_m = por_mes(carregar_receita, "valor_planejado", "valor_realizado")
    ded_m = por_mes(carregar_deducao, "valor_planejado", "valor_realizado")
    cmv_m = por_mes(carregar_cmv, "valor_planejado", "valor_realizado")
    pes_m = por_mes(carregar_hc_custo, "valor_orcado", "valor_realizado")

    mapa = {int(x["conta_cod"]): x.get("linha") for x in carregar_dre_mapa() if x.get("conta_cod") is not None}
    orc_cur = carregar_orc(ano) or []; orc_prev = carregar_orc(ano - 1) or []
    grp_m = {m: {g: {"p": 0.0, "r": 0.0, "a": 0.0} for g, _ in DRE_GRUPOS} for m in range(1, 13)}
    for x in orc_cur:
        m = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
        if 1 <= m <= 12 and g in grp_m[m] and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
            sg = -1.0 if g in DRE_REV else 1.0  # receita vem como crédito (negativo) -> inverte
            grp_m[m][g]["p"] += sg * float(x.get("valor_planejado") or 0); grp_m[m][g]["r"] += sg * float(x.get("valor_realizado") or 0)
    for x in orc_prev:
        m = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
        if 1 <= m <= 12 and g in grp_m[m] and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
            sg = -1.0 if g in DRE_REV else 1.0
            grp_m[m][g]["a"] += sg * float(x.get("valor_realizado") or 0)

    dif = lambda a, b: {"p": a["p"] - b["p"], "r": a["r"] - b["r"], "a": a["a"] - b["a"]}
    somar = lambda md, meses: {k: sum(md[m][k] for m in meses) for k in ("p", "r", "a")}

    # ----- quais linhas opcionais aparecem: decidido sobre o SPAN inteiro (colunas alinhadas) -----
    span = meses_range if visao == "Mensal" else list(range(1, ate + 1))
    gsum_span = {g: somar({m: grp_m[m][g] for m in range(1, 13)}, span) for g, _ in DRE_GRUPOS}
    nz = lambda d: any(abs(d[k]) > 0.005 for k in ("p", "r", "a"))
    op_incl = [g for g in DRE_OP_COST if nz(gsum_span[g])]
    inc_var = nz(gsum_span["Despesas Variáveis"])
    inc_rf = nz(gsum_span["Receitas Financeiras"]); inc_onop = nz(gsum_span["Outras Receitas Não Operacionais"])
    inc_df = nz(gsum_span["Despesas Financeiras"]); tem_pre = inc_rf or inc_onop or inc_df
    inc_imp = nz(gsum_span[DRE_IMPOSTO])

    def montar(meses, src=None, incl=None):
        rm, dm, cm, pm_, gm = src if src else (rec_m, ded_m, cmv_m, pes_m, grp_m)
        if incl is None:
            _var, _op, _rf, _onop, _df, _pre, _imp = inc_var, op_incl, inc_rf, inc_onop, inc_df, tem_pre, inc_imp
        else:
            _var = incl["var"]; _op = incl["op"]; _rf = incl["rf"]; _onop = incl["onop"]
            _df = incl["df"]; _pre = incl["pre"]; _imp = incl["imp"]
        rec = somar(rm, meses); ded = somar(dm, meses); cmv = somar(cm, meses); pes = somar(pm_, meses)
        rl = dif(rec, ded); lb = dif(rl, cmv)
        gsum = {g: somar({m: gm[m][g] for m in range(1, 13)}, meses) for g, _ in DRE_GRUPOS}
        var_sub = {k: sum(gsum[g][k] for g in DRE_VAR_COST) for k in ("p", "r", "a")}
        mc = {k: lb[k] - var_sub[k] for k in ("p", "r", "a")}  # margem de contribuição
        op_sub = {k: sum(gsum[g][k] for g in DRE_OP_COST) for k in ("p", "r", "a")}
        resop = {k: mc[k] - pes[k] - op_sub[k] for k in ("p", "r", "a")}
        pre_add = {k: sum(gsum[g][k] for g in DRE_PRE_ADD) for k in ("p", "r", "a")}
        pre_sub = {k: sum(gsum[g][k] for g in DRE_PRE_SUB) for k in ("p", "r", "a")}
        res_ai = {k: resop[k] + pre_add[k] - pre_sub[k] for k in ("p", "r", "a")}
        imp = gsum[DRE_IMPOSTO]; res_liq = {k: res_ai[k] - imp[k] for k in ("p", "r", "a")}
        L = [("Receita Bruta de Vendas", rec, "rev", False),
             ("(−) Deduções de Vendas", ded, "cost", False),
             ("(=) Receita Líquida", rl, "rev", True),
             ("(−) CMV", cmv, "cost", False),
             ("(=) Lucro Bruto", lb, "rev", True),
             ("(−) Despesas com Pessoal", pes, "cost", False)]
        if _var:
            # Lucro Bruto = Margem de Contribuição Bruta; após as variáveis vem a Margem de Contribuição
            L.insert(5, ("(−) Despesas Variáveis", gsum["Despesas Variáveis"], "cost", False))
            L.insert(6, ("(=) Margem de Contribuição", mc, "rev", True))
        for g in _op:
            L.append((f"(−) {g}", gsum[g], "cost", False))
        L.append(("(=) Resultado Operacional", resop, "rev", True))
        if _pre:
            if _rf: L.append(("(+) Receitas Financeiras", gsum["Receitas Financeiras"], "rev", False))
            if _df: L.append(("(−) Despesas Financeiras", gsum["Despesas Financeiras"], "cost", False))
            if _rf or _df:
                resfin = {k: gsum["Receitas Financeiras"][k] - gsum["Despesas Financeiras"][k] for k in ("p", "r", "a")}
                L.append(("(=) Resultado Financeiro", resfin, "rev", True))
            if _onop: L.append(("(+) Outras Receitas Não Operacionais", gsum["Outras Receitas Não Operacionais"], "rev", False))
            L.append(("(=) Resultado antes de Impostos", res_ai, "rev", True))
        if _imp:
            L.append(("(−) Impostos (IRPJ/CSLL)", imp, "cost", False))
            L.append(("(=) Resultado Líquido", res_liq, "rev", True))
        return L

    st.caption(f"Período: {faixa} · Empresa: {'Todas' if not emp else EMP[emp]}. "
               "Mensal = cada período isolado; Acumulada = somado desde Janeiro (com Mês inicial em Janeiro, os dois coincidem).")

    # ================= CONSOLIDADO (uma coluna para o período) =================
    if formato == "Consolidado":
        st.caption("Exibir colunas:")
        cc = st.columns(7)
        col_plan = cc[0].checkbox("Planejado", True, key="dc_p")
        col_real = cc[1].checkbox("Realizado", True, key="dc_r")
        col_ant = cc[2].checkbox(f"Ano ant. ({ano-1})", False, key="dc_a")
        col_vr = cc[3].checkbox("Var (R$)", True, key="dc_vr")
        col_vp = cc[4].checkbox("Var (%)", True, key="dc_vp")
        col_av = cc[5].checkbox("AV %", False, key="dc_av")
        col_ah = cc[6].checkbox("AH %", False, key="dc_ah")

        meses = meses_range if visao == "Mensal" else list(range(1, ate + 1))
        linhas_dre = montar(meses)
        base_r = next((d["r"] for n, d, t, fo in linhas_dre if n == "(=) Receita Líquida"), 0.0)

        heads = ["Linha"]
        if col_plan: heads.append("Planejado")
        if col_real: heads.append("Realizado")
        if col_ant: heads.append(f"Ano ant. ({ano-1})")
        if col_vr: heads.append("Var. (R$)")
        if col_vp: heads.append("Var. (%)")
        if col_av: heads.append("AV %")
        if col_ah: heads.append("AH %")
        th = f"<th style='text-align:left'>{heads[0]}</th>" + "".join(f"<th>{h}</th>" for h in heads[1:])

        def linha(nome, d, tipo, forte=False):
            p, r, a = d["p"], d["r"], d["a"]
            vr = r - p; vp = (vr / p * 100) if p else 0.0
            if tipo == "cost":
                cor = CINZA_TXT if (p and abs(vp) <= banda) else (VERMELHO if vr > 0 else VERDE)
            else:
                cor = CINZA_TXT if (p and abs(vp) <= banda) else (VERDE if vr >= 0 else VERMELHO)
            av = (r / base_r * 100) if base_r else 0.0
            ah = ((r - a) / a * 100) if a else 0.0
            b0, b1 = ("<b>", "</b>") if forte else ("", "")
            tds = [f"<td style='text-align:left'>{b0}{nome}{b1}</td>"]
            if col_plan: tds.append(f"<td>{b0}{brl(p)}{b1}</td>")
            if col_real: tds.append(f"<td>{b0}{brl(r)}{b1}</td>")
            if col_ant: tds.append(f"<td>{b0}{brl(a) if a else '—'}{b1}</td>")
            if col_vr: tds.append(f"<td style='color:{cor}'>{b0}{brl(vr)}{b1}</td>")
            if col_vp: tds.append(f"<td style='color:{cor}'>{b0}{pct_txt(vp)}{b1}</td>")
            if col_av: tds.append(f"<td>{b0}{pct_txt(av)}{b1}</td>")
            if col_ah: tds.append(f"<td>{b0}{(pct_txt(ah) if a else '—')}{b1}</td>")
            cls = " class='mark'" if forte else ""
            return f"<tr{cls}>{''.join(tds)}</tr>"

        corpo = "".join(linha(n, d, t, forte=fo) for n, d, t, fo in linhas_dre)
        st.markdown(f"""<div class='scroll'><table class="lle"><tr>{th}</tr>{corpo}</table></div>""", unsafe_allow_html=True)
        try:
            st.download_button("📥 Baixar DRE (Excel)",
                               data=dre_xlsx(linhas_dre, base_r, ano, faixa, ("Todas" if not emp else EMP[emp])),
                               file_name=f"DRE_{('todas' if not emp else EMP[emp])}_{ano}.xlsx", mime=XLSX_MIME, key="dre_dl")
        except Exception:
            pass

    # ================= POR UNIDADE (PISA | KING | Consolidado) =================
    elif formato == "Por unidade":
        meses_u = meses_range if visao == "Mensal" else list(range(1, ate + 1))
        corp_set = get_cr_corporativos(c)   # CRs marcados como corporativos

        def dicts_unidade(kind):
            # kind: 'all' | 'pisa' | 'king' | 'corp'
            # Receita/Deduções/CMV são por EMPRESA (sem CR) → ficam em PISA/KING; 'corp' = 0.
            # Despesas (orçamento) e Pessoal são por CR → CRs corporativos migram para 'corp'.
            def emp_ok(uni):
                if kind == "pisa": return uni == 1
                if kind == "king": return uni == 2
                return True
            def cr_ok(uni, cr):
                if kind == "all": return True
                if kind == "corp": return cr in corp_set
                if kind == "pisa": return uni == 1 and cr not in corp_set
                if kind == "king": return uni == 2 and cr not in corp_set
                return False
            def pm_emp(loader, kp, kr):
                dd = {mm: {"p": 0.0, "r": 0.0, "a": 0.0} for mm in range(1, 13)}
                if kind == "corp":
                    return dd
                cur = loader(ano) or []; prev = loader(ano - 1) or []
                for x in cur:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and emp_ok(int(x.get("uni_cod", 0) or 0)):
                        dd[mm]["p"] += float(x.get(kp) or 0); dd[mm]["r"] += float(x.get(kr) or 0)
                for x in prev:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and emp_ok(int(x.get("uni_cod", 0) or 0)):
                        dd[mm]["a"] += float(x.get(kr) or 0)
                return dd
            def pm_cr(loader, kp, kr):
                dd = {mm: {"p": 0.0, "r": 0.0, "a": 0.0} for mm in range(1, 13)}
                cur = loader(ano) or []; prev = loader(ano - 1) or []
                for x in cur:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and cr_ok(int(x.get("uni_cod", 0) or 0), int(x.get("cr_cod", 0) or 0)):
                        dd[mm]["p"] += float(x.get(kp) or 0); dd[mm]["r"] += float(x.get(kr) or 0)
                for x in prev:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and cr_ok(int(x.get("uni_cod", 0) or 0), int(x.get("cr_cod", 0) or 0)):
                        dd[mm]["a"] += float(x.get(kr) or 0)
                return dd
            rec = pm_emp(carregar_receita, "valor_planejado", "valor_realizado")
            ded = pm_emp(carregar_deducao, "valor_planejado", "valor_realizado")
            cmv = pm_emp(carregar_cmv, "valor_planejado", "valor_realizado")
            pes = pm_cr(carregar_hc_custo, "valor_orcado", "valor_realizado")
            gm = {mm: {g: {"p": 0.0, "r": 0.0, "a": 0.0} for g, _ in DRE_GRUPOS} for mm in range(1, 13)}
            for x in orc_cur:
                mm = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
                if 1 <= mm <= 12 and g in gm[mm] and cr_ok(int(x.get("uni_cod", 0) or 0), int(x.get("cr_cod", 0) or 0)):
                    sg = -1.0 if g in DRE_REV else 1.0
                    gm[mm][g]["p"] += sg * float(x.get("valor_planejado") or 0); gm[mm][g]["r"] += sg * float(x.get("valor_realizado") or 0)
            for x in orc_prev:
                mm = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
                if 1 <= mm <= 12 and g in gm[mm] and cr_ok(int(x.get("uni_cod", 0) or 0), int(x.get("cr_cod", 0) or 0)):
                    sg = -1.0 if g in DRE_REV else 1.0
                    gm[mm][g]["a"] += sg * float(x.get("valor_realizado") or 0)
            return rec, ded, cmv, pes, gm

        def montar_u(dd, meses, op_inc, tp, i_rf, i_onop, i_df, i_imp, i_var):
            rm, dm, cm, pm2, gmv = dd
            rec = somar(rm, meses); ded = somar(dm, meses); cmv = somar(cm, meses); pes = somar(pm2, meses)
            rl = dif(rec, ded); lb = dif(rl, cmv)
            gsum = {g: somar({mm: gmv[mm][g] for mm in range(1, 13)}, meses) for g, _ in DRE_GRUPOS}
            var_sub = {k: sum(gsum[g][k] for g in DRE_VAR_COST) for k in ("p", "r", "a")}
            mc = {k: lb[k] - var_sub[k] for k in ("p", "r", "a")}
            op_sub = {k: sum(gsum[g][k] for g in DRE_OP_COST) for k in ("p", "r", "a")}
            resop = {k: mc[k] - pes[k] - op_sub[k] for k in ("p", "r", "a")}
            pre_add = {k: sum(gsum[g][k] for g in DRE_PRE_ADD) for k in ("p", "r", "a")}
            pre_sub = {k: sum(gsum[g][k] for g in DRE_PRE_SUB) for k in ("p", "r", "a")}
            res_ai = {k: resop[k] + pre_add[k] - pre_sub[k] for k in ("p", "r", "a")}
            imp = gsum[DRE_IMPOSTO]; res_liq = {k: res_ai[k] - imp[k] for k in ("p", "r", "a")}
            L = [("Receita Bruta de Vendas", rec), ("(−) Deduções de Vendas", ded),
                 ("(=) Receita Líquida", rl), ("(−) CMV", cmv), ("(=) Lucro Bruto", lb)]
            if i_var:
                L.append(("(−) Despesas Variáveis", gsum["Despesas Variáveis"]))
                L.append(("(=) Margem de Contribuição", mc))
            L.append(("(−) Despesas com Pessoal", pes))
            for g in op_inc:
                L.append((f"(−) {g}", gsum[g]))
            L.append(("(=) Resultado Operacional", resop))
            if tp:
                if i_rf: L.append(("(+) Receitas Financeiras", gsum["Receitas Financeiras"]))
                if i_df: L.append(("(−) Despesas Financeiras", gsum["Despesas Financeiras"]))
                if i_rf or i_df:
                    L.append(("(=) Resultado Financeiro", {k: gsum["Receitas Financeiras"][k] - gsum["Despesas Financeiras"][k] for k in ("p", "r", "a")}))
                if i_onop: L.append(("(+) Outras Receitas Não Operacionais", gsum["Outras Receitas Não Operacionais"]))
                L.append(("(=) Resultado antes de Impostos", res_ai))
            if i_imp:
                L.append(("(−) Impostos (IRPJ/CSLL)", imp))
                L.append(("(=) Resultado Líquido", res_liq))
            return L

        dd_all = dicts_unidade("all"); dd_pisa = dicts_unidade("pisa")
        dd_king = dicts_unidade("king"); dd_corp = dicts_unidade("corp")
        # inclusão de linhas decidida sobre o CONSOLIDADO (para alinhar as colunas)
        gall = dd_all[4]
        gsp = {g: somar({mm: gall[mm][g] for mm in range(1, 13)}, meses_u) for g, _ in DRE_GRUPOS}
        op_inc = [g for g in DRE_OP_COST if nz(gsp[g])]
        i_var = nz(gsp["Despesas Variáveis"])
        i_rf = nz(gsp["Receitas Financeiras"]); i_onop = nz(gsp["Outras Receitas Não Operacionais"])
        i_df = nz(gsp["Despesas Financeiras"]); tp = i_rf or i_onop or i_df; i_imp = nz(gsp[DRE_IMPOSTO])
        L_all = montar_u(dd_all, meses_u, op_inc, tp, i_rf, i_onop, i_df, i_imp, i_var)
        L_pisa = montar_u(dd_pisa, meses_u, op_inc, tp, i_rf, i_onop, i_df, i_imp, i_var)
        L_king = montar_u(dd_king, meses_u, op_inc, tp, i_rf, i_onop, i_df, i_imp, i_var)
        L_corp = montar_u(dd_corp, meses_u, op_inc, tp, i_rf, i_onop, i_df, i_imp, i_var)

        med = st.radio("Medida", ["Realizado", "Planejado", f"Ano ant. ({ano-1})"], horizontal=True, key="dre_uni_medida")
        key_med = {"Realizado": "r", "Planejado": "p"}.get(med, "a")
        tem_corp = len(corp_set) > 0
        st.caption("Comparativo por unidade de negócio — PISA, KING e Corporativo (CRs que atendem ao corporativo, "
                   "movidos apenas nesta visão) e o Consolidado (soma das unidades). Receita/Deduções/CMV ficam em "
                   "PISA/KING; o Corporativo concentra despesas e pessoal dos CRs marcados. Configure os CRs em "
                   "Administração › Unidades de negócio / Config DRE.")

        if tem_corp:
            th = "<th style='text-align:left'>Linha</th><th>PISA</th><th>KING</th><th>Corporativo</th><th>Consolidado</th>"
        else:
            th = "<th style='text-align:left'>Linha</th><th>PISA</th><th>KING</th><th>Consolidado</th>"
        corpo = ""
        for i in range(len(L_all)):
            n, da = L_all[i]; dp = L_pisa[i][1]; dk = L_king[i][1]; dc = L_corp[i][1]
            forte = n.startswith("(=")
            b0, b1 = ("<b>", "</b>") if forte else ("", "")
            cls = " class='mark'" if forte else ""
            cells = f"<td>{b0}{brl(dp[key_med])}{b1}</td><td>{b0}{brl(dk[key_med])}{b1}</td>"
            if tem_corp:
                cells += f"<td>{b0}{brl(dc[key_med])}{b1}</td>"
            cells += f"<td>{b0}{brl(da[key_med])}{b1}</td>"
            corpo += f"<tr{cls}><td style='text-align:left'>{b0}{n}{b1}</td>{cells}</tr>"
        st.markdown(f"<div class='scroll'><table class='lle'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)

        # ----- lucratividade por unidade de negócio -----
        def _line(L, nome):
            return next((d for nn, d in L if nn == nome), None)
        def _bottom(L):
            res = None
            for nn, d in L:
                if nn.startswith("(="):
                    res = d
            return res
        st.markdown("#### Lucratividade por unidade de negócio")
        unidades = [("PISA", L_pisa), ("KING", L_king)]
        if tem_corp:
            unidades.append(("Corporativo", L_corp))
        unidades.append(("Consolidado", L_all))
        linhas_luc = ""
        for nome_u, L in unidades:
            rl = _line(L, "(=) Receita Líquida"); res = _bottom(L)
            rlv = rl[key_med] if rl else 0.0
            resv = res[key_med] if res else 0.0
            marg = (resv / rlv * 100) if rlv else None
            cor = CINZA_TXT if marg is None else (VERDE if resv >= 0 else VERMELHO)
            marg_txt = pct_txt(marg) if marg is not None else "—"
            linhas_luc += (f"<tr><td style='text-align:left'>{nome_u}</td>"
                           f"<td>{brl(rlv)}</td><td style='color:{cor}'>{brl(resv)}</td>"
                           f"<td style='color:{cor}'>{marg_txt}</td></tr>")
        st.markdown("<div class='scroll'><table class='lle'><tr>"
                    "<th style='text-align:left'>Unidade de negócio</th><th>Receita Líquida</th>"
                    f"<th>Resultado Líquido</th><th>Margem líquida</th></tr>{linhas_luc}</table></div>", unsafe_allow_html=True)
        st.caption("Margem líquida = Resultado Líquido ÷ Receita Líquida da unidade. O Corporativo não tem receita "
                   "própria (concentra despesas), por isso a margem aparece como “—”.")

        try:
            import io
            rows = []
            for i in range(len(L_all)):
                n, da = L_all[i]; dp = L_pisa[i][1]; dk = L_king[i][1]; dc = L_corp[i][1]
                r = {"Linha": n, "PISA": round(dp[key_med]), "KING": round(dk[key_med])}
                if tem_corp:
                    r["Corporativo"] = round(dc[key_med])
                r["Consolidado"] = round(da[key_med])
                rows.append(r)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as x:
                pd.DataFrame(rows).to_excel(x, index=False, sheet_name="DRE por unidade")
            st.download_button("📥 Baixar DRE por unidade (Excel)", data=buf.getvalue(),
                               file_name=f"DRE_por_unidade_{ano}.xlsx", mime=XLSX_MIME, key="dre_uni_dl")
        except Exception:
            pass

    # ================= POR MÊS (matriz, como no Treasy) =================
    elif formato == "Por mês":
        medida = st.radio("Valores por mês", ["Realizado", "Planejado", "Ambos"], horizontal=True, key="dre_medida")
        avc = st.columns([1, 1, 6])
        col_av = avc[0].checkbox("AV %", False, key="dm_av", help="Análise Vertical: linha ÷ Receita Líquida realizada da mesma coluna.")
        col_ah = avc[1].checkbox("AH %", False, key="dm_ah", help="Análise Horizontal: realizado do período vs mesmo período do ano anterior.")
        kk = "p" if medida == "Planejado" else "r"
        skel = montar(meses_range)  # nomes/estrutura de referência
        # valores de cada coluna-mês
        col_vals = {}
        for m in meses_range:
            col_vals[m] = montar([m]) if visao == "Mensal" else montar(list(range(1, m + 1)))
        total_meses = meses_range if visao == "Mensal" else list(range(1, ate + 1))
        total_vals = montar(total_meses)
        # AV/AH são calculados sempre sobre o REALIZADO (padrão DRE)
        def _rl_r(vals):
            return next((d["r"] for n, d, t, f in vals if n == "(=) Receita Líquida"), 0.0)
        base_col = {m: _rl_r(col_vals[m]) for m in meses_range}
        base_tot = _rl_r(total_vals)

        def head_mes(lbl):
            h = (f"<th>{lbl} Plan</th><th>{lbl} Real</th>" if medida == "Ambos" else f"<th>{lbl}</th>")
            if col_av: h += f"<th>{lbl} AV%</th>"
            if col_ah: h += f"<th>{lbl} AH%</th>"
            return h
        sub = "".join(head_mes(MABREV[m]) for m in meses_range) + head_mes("Total")
        th = f"<th style='text-align:left'>Linha</th>{sub}"

        def cells(vi, base, b0, b1):
            r = vi["r"]; a = vi["a"]
            if medida == "Ambos":
                cc = f"<td>{b0}{brl(vi['p'])}{b1}</td><td>{b0}{brl(vi['r'])}{b1}</td>"
            else:
                cc = f"<td>{b0}{brl(vi[kk])}{b1}</td>"
            if col_av:
                cc += f"<td>{b0}{pct_txt((r / base * 100) if base else 0.0)}{b1}</td>"
            if col_ah:
                cc += f"<td>{b0}{(pct_txt((r - a) / a * 100) if a else '—')}{b1}</td>"
            return cc

        corpo = ""
        for i, (nome, d0, tipo, forte) in enumerate(skel):
            b0, b1 = ("<b>", "</b>") if forte else ("", "")
            tds = f"<td style='text-align:left'>{b0}{nome}{b1}</td>"
            for m in meses_range:
                tds += cells(col_vals[m][i][1], base_col[m], b0, b1)
            tds += cells(total_vals[i][1], base_tot, b0, b1)
            cls = " class='mark'" if forte else ""
            corpo += f"<tr{cls}>{tds}</tr>"
        st.markdown(f"""<div class='scroll'><table class="lle matrix"><tr>{th}</tr>{corpo}</table></div>""", unsafe_allow_html=True)

    # ================= COMPARAR PERÍODOS (realizado de dois períodos quaisquer) =================
    else:
        st.caption("Compare o **realizado** de dois períodos quaisquer (inclusive anos diferentes). Cada lado pode ser um "
                   "único mês (mês inicial = final) ou um intervalo. AV% = participação sobre a Receita Líquida do próprio "
                   "período; AH% = variação vs o mesmo período do ano anterior. Var = período B − período A.")
        anos_cmp = list(range(2020, max(ano, 2026) + 1))
        idx_ant = anos_cmp.index(ano - 1) if (ano - 1) in anos_cmp else 0
        idx_atual = anos_cmp.index(ano) if ano in anos_cmp else len(anos_cmp) - 1
        st.markdown(f"<div style='font-weight:600;color:{AZUL_PROFUNDO}'>Período A</div>", unsafe_allow_html=True)
        ca = st.columns([1, 1, 1])
        ano_a = int(ca[0].selectbox("Ano (A)", anos_cmp, index=idx_ant, key="cmp_ano_a"))
        de_a = ca[1].selectbox("Mês inicial (A)", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="cmp_de_a")
        ate_a = ca[2].selectbox("Mês final (A)", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="cmp_ate_a")
        st.markdown(f"<div style='font-weight:600;color:{AZUL_PROFUNDO};margin-top:4px'>Período B</div>", unsafe_allow_html=True)
        cb = st.columns([1, 1, 1])
        ano_b = int(cb[0].selectbox("Ano (B)", anos_cmp, index=idx_atual, key="cmp_ano_b"))
        de_b = cb[1].selectbox("Mês inicial (B)", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="cmp_de_b")
        ate_b = cb[2].selectbox("Mês final (B)", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="cmp_ate_b")
        if ate_a < de_a: ate_a = de_a
        if ate_b < de_b: ate_b = de_b
        meses_a = list(range(de_a, ate_a + 1)); meses_b = list(range(de_b, ate_b + 1))

        # agregador por ano arbitrário: r = realizado do ano, a = realizado do ano-1, p = planejado do ano
        def _agrega_ano(yr):
            def pm(loader, kp, kr):
                cur = loader(yr) or []; prev = loader(yr - 1) or []
                dd = {mm: {"p": 0.0, "r": 0.0, "a": 0.0} for mm in range(1, 13)}
                for x in cur:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                        dd[mm]["p"] += float(x.get(kp) or 0); dd[mm]["r"] += float(x.get(kr) or 0)
                for x in prev:
                    mm = int(x.get("mes", 0) or 0)
                    if 1 <= mm <= 12 and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                        dd[mm]["a"] += float(x.get(kr) or 0)
                return dd
            rec = pm(carregar_receita, "valor_planejado", "valor_realizado")
            ded = pm(carregar_deducao, "valor_planejado", "valor_realizado")
            cmv = pm(carregar_cmv, "valor_planejado", "valor_realizado")
            pes = pm(carregar_hc_custo, "valor_orcado", "valor_realizado")
            oc = carregar_orc(yr) or []; op = carregar_orc(yr - 1) or []
            gm = {mm: {g: {"p": 0.0, "r": 0.0, "a": 0.0} for g, _ in DRE_GRUPOS} for mm in range(1, 13)}
            for x in oc:
                mm = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
                if 1 <= mm <= 12 and g in gm[mm] and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                    sg = -1.0 if g in DRE_REV else 1.0
                    gm[mm][g]["p"] += sg * float(x.get("valor_planejado") or 0); gm[mm][g]["r"] += sg * float(x.get("valor_realizado") or 0)
            for x in op:
                mm = int(x.get("mes", 0) or 0); g = mapa.get(int(x.get("conta_cod", 0) or 0))
                if 1 <= mm <= 12 and g in gm[mm] and (not emp or int(x.get("uni_cod", 0) or 0) == emp):
                    sg = -1.0 if g in DRE_REV else 1.0
                    gm[mm][g]["a"] += sg * float(x.get("valor_realizado") or 0)
            return (rec, ded, cmv, pes, gm)

        srcA = _agrega_ano(ano_a); srcB = _agrega_ano(ano_b)
        # inclusão de linhas opcionais: união dos dois períodos (colunas alinhadas)
        acc = {g: {"p": 0.0, "r": 0.0, "a": 0.0} for g, _ in DRE_GRUPOS}
        for src, mss in ((srcA, meses_a), (srcB, meses_b)):
            gm = src[4]
            for g, _ in DRE_GRUPOS:
                s = somar({m: gm[m][g] for m in range(1, 13)}, mss)
                for k in ("p", "r", "a"): acc[g][k] += s[k]
        incl_cmp = {"op": [g for g in DRE_OP_COST if nz(acc[g])],
                    "var": nz(acc["Despesas Variáveis"]),
                    "rf": nz(acc["Receitas Financeiras"]), "onop": nz(acc["Outras Receitas Não Operacionais"]),
                    "df": nz(acc["Despesas Financeiras"]),
                    "pre": nz(acc["Receitas Financeiras"]) or nz(acc["Outras Receitas Não Operacionais"]) or nz(acc["Despesas Financeiras"]),
                    "imp": nz(acc[DRE_IMPOSTO])}
        LA = montar(meses_a, src=srcA, incl=incl_cmp)
        LB = montar(meses_b, src=srcB, incl=incl_cmp)
        def _rl(L):
            return next((d["r"] for n, d, t, f in L if n == "(=) Receita Líquida"), 0.0)
        baseA = _rl(LA); baseB = _rl(LB)
        lbl_a = f"{MABREV[de_a]}–{MABREV[ate_a]}/{ano_a}" if de_a != ate_a else f"{MABREV[de_a]}/{ano_a}"
        lbl_b = f"{MABREV[de_b]}–{MABREV[ate_b]}/{ano_b}" if de_b != ate_b else f"{MABREV[de_b]}/{ano_b}"
        th = (f"<th style='text-align:left'>Linha</th>"
              f"<th>{lbl_a}</th><th>{lbl_a} AV%</th><th>{lbl_a} AH%</th>"
              f"<th>{lbl_b}</th><th>{lbl_b} AV%</th><th>{lbl_b} AH%</th>"
              f"<th>Var (R$)</th><th>Var (%)</th>")
        corpo = ""
        for (n, da, ta, fa), (_, db, tb, fb) in zip(LA, LB):
            forte = n.startswith("(=")
            b0, b1 = ("<b>", "</b>") if forte else ("", "")
            ra = da["r"]; rb = db["r"]; aa = da["a"]; ab = db["a"]
            ava = (ra / baseA * 100) if baseA else 0.0
            avb = (rb / baseB * 100) if baseB else 0.0
            aha = ((ra - aa) / aa * 100) if aa else None
            ahb = ((rb - ab) / ab * 100) if ab else None
            var = rb - ra; varp = (var / ra * 100) if ra else 0.0
            if ta == "cost":
                cor = CINZA_TXT if (ra and abs(varp) <= banda) else (VERMELHO if var > 0 else VERDE)
            else:
                cor = CINZA_TXT if (ra and abs(varp) <= banda) else (VERDE if var >= 0 else VERMELHO)
            cls = " class='mark'" if forte else ""
            corpo += (f"<tr{cls}><td style='text-align:left'>{b0}{n}{b1}</td>"
                      f"<td>{b0}{brl(ra)}{b1}</td><td>{b0}{pct_txt(ava)}{b1}</td><td>{b0}{(pct_txt(aha) if aha is not None else '—')}{b1}</td>"
                      f"<td>{b0}{brl(rb)}{b1}</td><td>{b0}{pct_txt(avb)}{b1}</td><td>{b0}{(pct_txt(ahb) if ahb is not None else '—')}{b1}</td>"
                      f"<td style='color:{cor}'>{b0}{brl(var)}{b1}</td><td style='color:{cor}'>{b0}{pct_txt(varp)}{b1}</td></tr>")
        st.markdown(f"<div class='scroll'><table class='lle matrix'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)
        st.caption(f"Empresa: {'Todas' if not emp else EMP[emp]}. Só há dados quando o ano escolhido tem realizado carregado "
                   "(receita, deduções, CMV, pessoal e orçado). AH% usa o ano anterior de cada período.")
        try:
            import io
            rows = []
            for (n, da, ta, fa), (_, db, tb, fb) in zip(LA, LB):
                ra = da["r"]; rb = db["r"]; aa = da["a"]; ab = db["a"]
                rows.append({"Linha": n, lbl_a: round(ra),
                             f"{lbl_a} AV%": round((ra / baseA * 100) if baseA else 0.0, 2),
                             f"{lbl_a} AH%": round(((ra - aa) / aa * 100) if aa else 0.0, 2),
                             lbl_b: round(rb),
                             f"{lbl_b} AV%": round((rb / baseB * 100) if baseB else 0.0, 2),
                             f"{lbl_b} AH%": round(((rb - ab) / ab * 100) if ab else 0.0, 2),
                             "Var (R$)": round(rb - ra),
                             "Var (%)": round(((rb - ra) / ra * 100) if ra else 0.0, 2)})
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as x:
                pd.DataFrame(rows).to_excel(x, index=False, sheet_name="DRE comparação")
            st.download_button("📥 Baixar comparação (Excel)", data=buf.getvalue(),
                               file_name=f"DRE_comparacao_{ano_a}_vs_{ano_b}.xlsx", mime=XLSX_MIME, key="cmp_dl")
        except Exception:
            pass

    if formato != "Comparar períodos" and not any(nz(gsum_span[g]) for g, _ in DRE_GRUPOS):
        st.info("Nenhuma conta do orçamento mapeada ainda. Configure em **Administração › Unidades de negócio / Config DRE**.")

# ---------------------------------------------------------------- investimentos
def tela_investimento(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    banda = get_faixa(c)
    st.markdown("<div class='modtag'>Investimentos</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Planejado x realizado por empresa e mês (investimentos / CAPEX). Convenção: realizado abaixo do planejado é favorável (investiu menos que o previsto).</div>", unsafe_allow_html=True)

    rows = carregar_investimento(ano)
    dfr = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ano", "mes", "uni_cod", "unidade", "valor_planejado", "valor_realizado"])
    for col in ("valor_planejado", "valor_realizado"):
        if col not in dfr.columns: dfr[col] = 0.0

    emp = st.selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="inv_emp")
    scope = dfr if not emp else dfr[dfr["uni_cod"] == emp]

    if scope.empty:
        g = pd.DataFrame(0.0, index=range(1, 13), columns=["valor_planejado", "valor_realizado"])
    else:
        g = scope.groupby("mes")[["valor_planejado", "valor_realizado"]].sum().reindex(range(1, 13), fill_value=0.0)
    tp = float(g["valor_planejado"].sum()); tr = float(g["valor_realizado"].sum())
    mask_real = g["valor_realizado"] != 0
    tp_real = float(g.loc[mask_real, "valor_planejado"].sum())
    var = tr - tp_real; pct = (var / tp_real * 100) if tp_real else 0.0  # variação só sobre meses com realizado

    def cor_inv(v, pl):
        p = (v / pl * 100) if pl else 0.0
        if pl and abs(p) <= banda: return CINZA_TXT
        return VERDE if v <= 0 else VERMELHO

    k = st.columns(4)
    kpi = [("Planejado (ano)", brl(tp), CINZA_TXT), ("Realizado (ano)", brl(tr), CINZA_TXT),
           ("Variação (R$)", brl(var), cor_inv(var, tp_real)), ("Variação (%)", pct_txt(pct), cor_inv(var, tp_real))]
    for col, (t, v, cr) in zip(k, kpi):
        col.markdown(f"<div class='card' style='text-align:center'><div style='font-size:.8rem;color:{CINZA_TXT}'>{t}</div>"
                     f"<div style='font-size:1.4rem;font-weight:700;color:{cr}'>{v}</div></div>", unsafe_allow_html=True)

    st.caption(f"Variação calculada apenas sobre os meses com realizado (planejado desses meses: {brl(tp_real)}). Orçado cheio do ano: {brl(tp)}.")
    st.markdown("#### Evolução mensal")
    linhas = ""
    for m in range(1, 13):
        vp = float(g.loc[m, "valor_planejado"]); vr = float(g.loc[m, "valor_realizado"])
        v = vr - vp; p = (v / vp * 100) if vp else 0.0
        if vr == 0 and vp == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem dados", CINZA_TXT); cr = CINZA_TXT
        elif vr == 0:
            vr_txt = "\u2014"; var_txt = "\u2014"; pctv = "\u2014"; status = chip("Sem realizado", CINZA_TXT); cr = CINZA_TXT
        else:
            cr = cor_inv(v, vp)
            lab = "Neutro" if cr == CINZA_TXT else ("Favorável" if v <= 0 else "Desfavorável")
            vr_txt = brl(vr); var_txt = brl(v); pctv = pct_txt(p); status = chip(lab, cr)
        linhas += (f"<tr><td style='text-align:left'>{MESES[m]}</td><td>{brl(vp)}</td><td>{vr_txt}</td>"
                   f"<td style='color:{cr}'>{var_txt}</td><td style='color:{cr}'>{pctv}</td><td>{status}</td></tr>")
    tcor = cor_inv(var, tp_real)
    total = (f"<tr class='mark'><td style='text-align:left'><b>Total</b></td><td><b>{brl(tp)}</b></td>"
             f"<td><b>{brl(tr) if tr else '\u2014'}</b></td><td style='color:{tcor}'><b>{brl(var) if tr else '\u2014'}</b></td>"
             f"<td style='color:{tcor}'><b>{pct_txt(pct) if tr else '\u2014'}</b></td><td></td></tr>")
    st.markdown(f"""<div class='scroll'><table class="lle"><tr>
        <th style='text-align:left'>Mês</th><th>Planejado</th><th>Realizado</th>
        <th>Var. (R$)</th><th>Var. (%)</th><th>Status</th></tr>{linhas}{total}</table></div>""", unsafe_allow_html=True)

    _registro_mensal_frag("investimento_valor", "investimento_log", carregar_investimento_log, dfr, emp, ano, c, prof, "investimentos", custo=True)

# ---------------------------------------------------------------- administração
def _dump_tabela(c, tabela, cap=200000):
    """Lê a tabela inteira, paginando de 1000 em 1000 (para backup)."""
    linhas, passo, ini = [], 1000, 0
    while ini < cap:
        try:
            lote = c.table(tabela).select("*").range(ini, ini + passo - 1).execute().data or []
        except Exception:
            return linhas
        linhas.extend(lote)
        if len(lote) < passo:
            break
        ini += passo
    return linhas

def backup_xlsx(c):
    """Backup completo: uma aba por tabela. Não inclui operacional_detalhe (raw, reimportável)."""
    import io
    tabelas = ["orc_realizado", "receita_venda", "deducao_valor", "cmv_valor", "investimento_valor",
               "hc_quadro", "hc_custo", "justificativa", "dre_mapa", "config", "cr_gestor", "gestor_usuario",
               "orc_log", "hc_log", "receita_log", "cmv_log", "deducao_log", "investimento_log"]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as x:
        houve = False
        for t in tabelas:
            try:
                rows = _dump_tabela(c, t)
                df = pd.DataFrame(rows)
                if df.empty:
                    df = pd.DataFrame({"(sem dados)": []})
                else:
                    for col in df.columns:
                        if df[col].dtype == object:
                            df[col] = df[col].apply(lambda v: str(v) if isinstance(v, (dict, list)) else v)
                df.to_excel(x, index=False, sheet_name=t[:31])
                houve = True
            except Exception:
                continue
        if not houve:
            pd.DataFrame({"info": ["sem dados"]}).to_excel(x, index=False, sheet_name="info")
    return buf.getvalue()

def _diagnostico(ano):
    """Verificações leves de integridade dos dados do ano. Retorna [(nome, ok, detalhe)]."""
    def mabrev(m): return MESES[m][:3] if 0 < m < len(MESES) else str(m)
    def meses_com(rows): return {int(r["mes"]) for r in (rows or []) if r.get("mes")}
    def dups(rows, keys):
        seen = {}
        for r in (rows or []):
            k = tuple(r.get(x) for x in keys); seen[k] = seen.get(k, 0) + 1
        return sum(v - 1 for v in seen.values() if v > 1)
    def negativos(rows):
        n = 0
        for r in (rows or []):
            if float(r.get("valor_planejado") or 0) < 0 or float(r.get("valor_realizado") or 0) < 0:
                n += 1
        return n

    orc = carregar_orc(ano) or []
    rec = carregar_receita(ano) or []; ded = carregar_deducao(ano) or []
    cmv = carregar_cmv(ano) or []; inv = carregar_investimento(ano) or []
    mapa_keys = {int(x["conta_cod"]) for x in (carregar_dre_mapa() or []) if x.get("conta_cod") is not None}
    achados = []

    contas = {int(x["conta_cod"]) for x in orc if x.get("conta_cod") is not None}
    nao_map = sorted(contas - mapa_keys)
    achados.append(("Contas do orçamento sem classificação na DRE", not nao_map,
                    "Todas classificadas." if not nao_map
                    else f"{len(nao_map)} conta(s) para revisar (ex.: {', '.join(map(str, nao_map[:6]))}…)"))

    real_so = sum(1 for x in orc if float(x.get("valor_realizado") or 0) > 0 and float(x.get("valor_planejado") or 0) == 0)
    orc_sr = sum(1 for x in orc if float(x.get("valor_planejado") or 0) > 0 and float(x.get("valor_realizado") or 0) == 0)
    achados.append(("Realizado sem orçamento correspondente", real_so == 0,
                    "Nenhum." if real_so == 0 else f"{real_so} linha(s) com realizado e orçado zero."))
    achados.append(("Orçado ainda sem realizado (normal p/ meses futuros)", True,
                    f"{orc_sr} linha(s)."))

    ativos = meses_com(orc)
    for nome, rows in [("Receita", rec), ("Deduções", ded), ("CMV", cmv), ("Investimentos", inv)]:
        faltam = sorted(ativos - meses_com(rows))
        achados.append((f"{nome}: meses sem lançamento (dentre os meses com orçamento)", not faltam,
                        "OK." if not faltam else f"faltam: {', '.join(mabrev(m) for m in faltam)}"))

    neg = negativos(rec) + negativos(ded) + negativos(cmv) + negativos(inv) + negativos(orc)
    achados.append(("Valores negativos nos lançamentos", neg == 0,
                    "Nenhum." if neg == 0 else f"{neg} lançamento(s) com valor negativo."))

    d = (dups(orc, ["ano", "mes", "uni_cod", "cr_cod", "conta_cod"]) + dups(rec, ["ano", "mes", "uni_cod"])
         + dups(ded, ["ano", "mes", "uni_cod", "conta"]) + dups(cmv, ["ano", "mes", "uni_cod"])
         + dups(inv, ["ano", "mes", "uni_cod"]))
    achados.append(("Duplicidades de chave nos registros", d == 0,
                    "Nenhuma." if d == 0 else f"{d} chave(s) repetida(s) — verifique reimportações."))
    return achados

def tela_admin(c, prof, ano):
    st.markdown("<div class='modtag'>Administração de acessos</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Substituição por desligamento, ativação e transferência de centros</div>", unsafe_allow_html=True)
    st.caption("Substitua o e-mail de um gestor desligado, ative/desative acessos e transfira centros de resultado. "
               "O histórico de justificativas é sempre preservado.")

    with st.expander("💾 Backup dos dados"):
        st.caption("Gera um Excel com uma aba por tabela (orçamento, receita, deduções, CMV, investimentos, pessoal, "
                   "justificativas, mapeamento da DRE, cadastros, config e logs de auditoria). Guarde como cópia de segurança. "
                   "Não inclui o detalhe de notas (operacional_detalhe), que é reimportável.")
        bc = st.columns([1, 1, 2])
        if bc[0].button("Gerar backup (.xlsx)", key="bk_gen"):
            with st.spinner("Lendo tabelas…"):
                st.session_state["_bk_bytes"] = backup_xlsx(c)
                st.session_state["_bk_ts"] = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        if st.session_state.get("_bk_bytes"):
            bc[1].download_button("📥 Baixar backup", data=st.session_state["_bk_bytes"],
                                  file_name=f"backup_orcamento_{st.session_state.get('_bk_ts','')}.xlsx",
                                  mime=XLSX_MIME, key="bk_dl")
            bc[2].caption(f"Backup gerado em {st.session_state.get('_bk_ts','')} — clique para baixar.")
    st.divider()

    with st.expander("🩺 Diagnóstico de dados"):
        st.caption("Verifica a integridade dos dados do ano selecionado: contas sem classificação na DRE, lacunas "
                   "orçado/realizado, meses faltando nos módulos, valores negativos e duplicidades de chave. Roda só quando você clica.")
        if st.button("Rodar diagnóstico", key="diag_run"):
            achados = _diagnostico(ano)
            n_alerta = sum(1 for _, ok, _ in achados if not ok)
            if n_alerta == 0:
                st.success("Nenhum ponto de atenção encontrado — dados consistentes.")
            else:
                st.warning(f"{n_alerta} ponto(s) de atenção. Veja abaixo.")
            corpo = ""
            for nome, ok, det in achados:
                ic = "✅" if ok else "⚠️"; cr = VERDE if ok else VERMELHO
                corpo += (f"<tr><td style='text-align:left'>{ic} {nome}</td>"
                          f"<td style='text-align:left;color:{cr}'>{det}</td></tr>")
            st.markdown(f"<table class='lle'><tr><th style='text-align:left'>Verificação</th>"
                        f"<th style='text-align:left'>Resultado</th></tr>{corpo}</table>", unsafe_allow_html=True)
            st.caption(f"Diagnóstico do ano {ano}. A conferência entre o detalhe de notas e o realizado (mais pesada) pode ser adicionada sob demanda.")
    st.divider()

    with st.expander("🚫 Contas isentas de justificativa"):
        st.caption("Contas que NÃO exigem justificativa (regra LLE — ex.: depreciação, contingências, impostos). "
                   "Elas aparecem na tela de justificativas marcadas como 'Isenta' e não entram nas pendências. "
                   "Informe os códigos separados por vírgula, espaço ou quebra de linha.")
        atuais = sorted(get_isentas(c))
        cat = {int(x["conta_cod"]): x.get("conta_desc", "") for x in (carregar_plano_contas() or []) if x.get("conta_cod") is not None}
        if atuais:
            corpo = "".join(f"<tr><td style='text-align:left'>{cod}</td><td style='text-align:left'>{cat.get(cod,'—')}</td></tr>" for cod in atuais)
            st.markdown(f"<table class='lle'><tr><th style='text-align:left'>Conta isenta</th>"
                        f"<th style='text-align:left'>Descrição</th></tr>{corpo}</table>", unsafe_allow_html=True)
        else:
            st.caption("Nenhuma conta isenta cadastrada ainda.")
        txt = st.text_area("Códigos das contas isentas", value=", ".join(str(x) for x in atuais), key="isentas_txt", height=100)
        if st.button("Salvar contas isentas", key="isentas_save"):
            novos = set_isentas(c, txt)
            limpar_cache_justif()
            st.success(f"{len(novos)} conta(s) isenta(s) salva(s)."); st.rerun()
    st.divider()

    gestores = c.table("gestor").select("codigo, nome, papel").order("nome").execute().data or []
    usuarios = c.table("gestor_usuario").select("email, gestor_codigo, papel_acesso, ativo").execute().data or []
    nome_por_cod = {g["codigo"]: g["nome"] for g in gestores}
    op_gestor = {f"{g['nome']} ({g['codigo']})": g["codigo"] for g in gestores}

    aba = st.radio("Ação", ["Substituir e-mail (desligamento)", "Ativar / desativar acesso", "Transferir centro de resultado"], horizontal=True)

    if aba.startswith("Substituir"):
        st.markdown("###### Substituir o e-mail de um gestor")
        st.caption("O e-mail antigo é desativado (mantém o histórico) e o novo passa a acessar os mesmos centros de resultado.")
        g_sel = st.selectbox("Gestor (função)", list(op_gestor.keys()))
        cod = op_gestor[g_sel]
        atuais = [u for u in usuarios if u["gestor_codigo"] == cod]
        if atuais:
            st.write("Acessos atuais desta função:")
            st.markdown("<table class='lle'><tr><th>E-mail</th><th>Tipo</th><th>Situação</th></tr>" +
                "".join(f"<tr><td>{u['email']}</td><td>{u['papel_acesso']}</td><td>{'Ativo' if u['ativo'] else 'Inativo'}</td></tr>" for u in atuais) +
                "</table>", unsafe_allow_html=True)
        email_antigo = st.selectbox("E-mail a desativar (do desligado)", ["(nenhum)"] + [u["email"] for u in atuais if u["ativo"]])
        email_novo = st.text_input("Novo e-mail (do substituto)", placeholder="novo.gestor@grupolle.com.br").strip().lower()
        st.info("Depois de salvar aqui, crie o login do novo e-mail no Supabase (Authentication → Users) com uma senha provisória.")
        if st.button("Aplicar substituição", type="primary"):
            if not email_novo or "@" not in email_novo:
                st.error("Informe um e-mail novo válido.")
            else:
                if email_antigo != "(nenhum)":
                    c.table("gestor_usuario").update({"ativo": False}).eq("email", email_antigo).execute()
                c.table("gestor_usuario").upsert({"email": email_novo, "gestor_codigo": cod, "papel_acesso": "titular", "ativo": True, "senha_provisoria": True}, on_conflict="email").execute(); limpar_cache()
                st.success(f"Feito: {email_novo} agora responde por {nome_por_cod.get(cod, cod)}." + (f" {email_antigo} foi desativado." if email_antigo != '(nenhum)' else ""))
                st.rerun()

    elif aba.startswith("Ativar"):
        st.markdown("###### Ativar ou desativar um acesso")
        if not usuarios:
            st.info("Nenhum usuário cadastrado.")
        else:
            for u in sorted(usuarios, key=lambda x: (nome_por_cod.get(x["gestor_codigo"], ""), x["email"])):
                col1, col2, col3 = st.columns([3, 2, 1.2])
                col1.write(f"**{u['email']}**")
                col2.write(f"{nome_por_cod.get(u['gestor_codigo'], u['gestor_codigo'])} · {'Ativo' if u['ativo'] else 'Inativo'}")
                novo_estado = not u["ativo"]
                rot = "Reativar" if not u["ativo"] else "Desativar"
                if col3.button(rot, key=f"tog_{u['email']}"):
                    c.table("gestor_usuario").update({"ativo": novo_estado}).eq("email", u["email"]).execute(); limpar_cache()
                    st.rerun()

    else:
        st.markdown("###### Transferir um centro de resultado para outro gestor")
        st.caption("Muda o dono do CR. As justificativas já feitas continuam registradas na conta/CR, sem alteração.")
        crs = c.table("cr_gestor").select("uni_cod, cr_cod, cr_nome, gestor_codigo").order("cr_nome").execute().data or []
        op_cr = {f"{r['cr_nome']} (uni {r['uni_cod']} · CR {r['cr_cod']}) — hoje: {nome_por_cod.get(r['gestor_codigo'], r['gestor_codigo'])}": (r["uni_cod"], r["cr_cod"]) for r in crs}
        cr_sel = st.selectbox("Centro de resultado", list(op_cr.keys()))
        destino = st.selectbox("Novo gestor responsável", list(op_gestor.keys()))
        if st.button("Transferir CR", type="primary"):
            uni, crc = op_cr[cr_sel]
            c.table("cr_gestor").update({"gestor_codigo": op_gestor[destino]}).eq("uni_cod", uni).eq("cr_cod", crc).execute(); limpar_cache()
            st.success(f"CR transferido para {destino.split(' (')[0]}.")
            st.rerun()

# ---------------------------------------------------------------- acompanhamento
def secao_valores(c, df_filtrado, ano, mes):
    """Visão Treasy: por mês, Planejado | Realizado | Histórico (realizado do ano anterior)
       + Variação por par escolhido, em Normal (mês) ou Acumulado (Jan até o mês)."""
    keys = {(int(r["uni_cod"]), int(r["cr_cod"]), int(r["conta_cod"])) for _, r in df_filtrado.iterrows()}
    plan = {m: 0.0 for m in range(1, 13)}; real = {m: 0.0 for m in range(1, 13)}; hist = {m: 0.0 for m in range(1, 13)}
    for _, r in df_filtrado.iterrows():
        m = int(r["mes"])
        if 1 <= m <= 12:
            plan[m] += float(r.get("valor_planejado") or 0); real[m] += float(r.get("valor_realizado") or 0)
    for r in (carregar_orc(ano - 1) or []):
        if (int(r.get("uni_cod", 0) or 0), int(r.get("cr_cod", 0) or 0), int(r.get("conta_cod", 0) or 0)) in keys:
            m = int(r.get("mes", 0) or 0)
            if 1 <= m <= 12: hist[m] += float(r.get("valor_realizado") or 0)

    cc = st.columns([1.3, 2.6])
    modo = cc[0].radio("Modo", ["Normal", "Acumulado"], horizontal=True, key="val_modo")
    base = cc[1].radio("Colunas de Variação",
                       ["Planejado × Realizado", "Planejado × Histórico", "Realizado × Histórico"],
                       horizontal=True, key="val_base")

    def acc(d):
        out = {}; s = 0.0
        for m in range(1, 13): s += d[m]; out[m] = s
        return out
    P = acc(plan) if modo == "Acumulado" else plan
    R = acc(real) if modo == "Acumulado" else real
    H = acc(hist) if modo == "Acumulado" else hist

    if base == "Planejado × Realizado":   A, B = P, R
    elif base == "Planejado × Histórico": A, B = P, H
    else:                                  A, B = R, H

    def cor_var(v):
        return VERMELHO if v < -0.005 else (VERDE if v > 0.005 else CINZA_TXT)

    th = ("<th style='text-align:left'>Ano/Mês</th><th>Planejado</th><th>Realizado</th>"
          "<th>Histórico</th><th>Variação (R$)</th><th>Variação (%)</th>")
    corpo = ""
    for m in range(1, 13):
        var = A[m] - B[m]
        vpct = (var / B[m] * 100) if abs(B[m]) > 0.005 else 0.0
        co = cor_var(var)
        corpo += (f"<tr><td style='text-align:left'>{ano} / {MABREV[m]}</td>"
                  f"<td>{brl(P[m])}</td><td>{brl(R[m])}</td><td>{brl(H[m])}</td>"
                  f"<td style='color:{co}'>{brl(var)}</td><td style='color:{co}'>{pct_txt(vpct)}</td></tr>")
    # Total do ano (soma dos 12, independe do modo)
    tP, tR, tH = sum(plan.values()), sum(real.values()), sum(hist.values())
    tA = tP if base.startswith("Planejado") else tR
    tB = tR if base == "Planejado × Realizado" else tH
    tvar = tA - tB; tpct = (tvar / tB * 100) if abs(tB) > 0.005 else 0.0
    co = cor_var(tvar)
    corpo += (f"<tr class='mark'><td style='text-align:left'><b>Total</b></td>"
              f"<td><b>{brl(tP)}</b></td><td><b>{brl(tR)}</b></td><td><b>{brl(tH)}</b></td>"
              f"<td style='color:{co}'><b>{brl(tvar)}</b></td><td style='color:{co}'><b>{pct_txt(tpct)}</b></td></tr>")
    st.markdown(f"<div class='scroll'><table class='lle'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)
    st.caption("Histórico = realizado do ano anterior (mesmos centros de resultado/contas do filtro). "
               "Variação = (primeira coluna − segunda) e % sobre a segunda.")

def aviso_estouros(d_mes, banda, mes, ano):
    """Aviso no topo do Acompanhamento: contas que estouraram a faixa no mês (desfavoráveis)."""
    estouros = []
    for _, v in d_mes.iterrows():
        raw, pct = var_de(v["valor_planejado"], v["valor_realizado"])
        lab, _ = classifica(raw, pct, v["conta_cod"], banda)
        if lab == "Desfavorável":
            estouros.append((abs(raw), raw, pct, v))
    if not estouros:
        return
    estouros.sort(key=lambda x: x[0], reverse=True)
    total = sum(e[0] for e in estouros)
    st.markdown(
        f"<div style='background:#FDECEA;border-left:4px solid {VERMELHO};padding:10px 14px;border-radius:6px;margin-bottom:8px;'>"
        f"<b style='color:{VERMELHO}'>⚠️ {len(estouros)} conta(s) estouraram a faixa em {MESES[mes]}/{ano}</b>"
        f" &nbsp;·&nbsp; impacto desfavorável total: <b>{brl(total)}</b></div>", unsafe_allow_html=True)
    th = ("<th style='text-align:left'>Conta</th><th style='text-align:left'>Centro de resultado</th>"
          "<th>Variação (R$)</th><th>Variação (%)</th>")
    def _linhas(lista):
        return "".join(
            f"<tr><td style='text-align:left'>{int(v['conta_cod'])} · {v.get('conta_desc','')}</td>"
            f"<td style='text-align:left'>{v.get('cr_nome','')} ({v.get('unidade','')})</td>"
            f"<td style='color:{VERMELHO}'>{brl(raw)}</td><td style='color:{VERMELHO}'>{pct_txt(pct)}</td></tr>"
            for _, raw, pct, v in lista)
    st.markdown(f"<div class='scroll'><table class='lle'><tr>{th}</tr>{_linhas(estouros[:5])}</table></div>", unsafe_allow_html=True)
    if len(estouros) > 5:
        with st.expander(f"Ver todas as {len(estouros)} contas estouradas"):
            st.markdown(f"<div class='scroll'><table class='lle'><tr>{th}</tr>{_linhas(estouros)}</table></div>", unsafe_allow_html=True)

@fragment
def tela_acompanhamento(c, prof, banda, df_orc, cg, is_ctrl, ano, mes, mostrar_justif=True):
    st.markdown("<div class='modtag'>Módulo Acompanhamento de Despesas — Orçado x Realizado</div>", unsafe_allow_html=True)

    if df_orc.empty:
        st.info("Nenhum dado carregado ainda." + (" Use a aba 'Importar dados'." if is_ctrl else " Fale com a controladoria."))
        return
    df = df_orc.copy()
    if is_ctrl and cg:
        df["_resp"] = df.apply(lambda r: cg.get((int(r["uni_cod"]), int(r["cr_cod"])), ("—", ""))[0], axis=1)

    tem_resp = is_ctrl and "_resp" in df.columns
    # ---------- filtros horizontais (o período vem do seletor global Ano/Mês) ----------
    fcols = st.columns([1.5, 1.2, 1.7, 1.7] if tem_resp else [1.2, 1.7, 1.7])
    i = 0
    if tem_resp:
        resps = ["Todos"] + sorted([r for r in df["_resp"].dropna().unique().tolist() if r])
        f_resp = fcols[i].selectbox("Gestor", resps, key="acomp_gestor"); i += 1
        if f_resp != "Todos": df = df[df["_resp"] == f_resp]
    unis = ["Todas"] + sorted(df["unidade"].dropna().unique().tolist())
    f_uni = fcols[i].selectbox("Unidade", unis, key="acomp_uni"); i += 1
    if f_uni != "Todas": df = df[df["unidade"] == f_uni]
    crs = ["Todos"] + sorted(df["cr_nome"].dropna().unique().tolist())
    f_cr = fcols[i].selectbox("Centro de resultado", crs, key="acomp_cr"); i += 1
    if f_cr != "Todos": df = df[df["cr_nome"] == f_cr]
    contas = ["Todas"] + sorted(df["conta_desc"].dropna().unique().tolist())
    f_conta = fcols[i].selectbox("Conta", contas, key="acomp_conta")
    if f_conta != "Todas": df = df[df["conta_desc"] == f_conta]

    # SEM consolidação: cada empresa (PISA/KING) vira uma linha própria por CR+conta
    d_mes = df[df["mes"] == mes]
    d_ytd = df[df["mes"] <= mes]

    # aviso de contas estouradas no mês (informativo, no topo)
    aviso_estouros(d_mes, banda, mes, ano)

    # aviso do gestor
    if not is_ctrl and mes < get_cobranca(c):
        st.info(f"{MESES[mes]}/{ano} não está sujeito à cobrança de justificativa.")
    elif not is_ctrl:
        js = carregar_justificativas(ano, mes)
        enviadas = {(int(j["uni_cod"]), int(j["cr_cod"]), int(j["conta_cod"])) for j in js if j.get("status") in ("JUSTIFICADO", "EM_REVISAO", "APROVADO")}
        pend = 0
        for _, v in d_mes.iterrows():
            raw, pct = var_de(v["valor_planejado"], v["valor_realizado"])
            lab, _ = classifica(raw, pct, v["conta_cod"], banda)
            if lab == "Desfavorável" and (int(v["uni_cod"]), int(v["cr_cod"]), int(v["conta_cod"])) not in enviadas:
                pend += 1
        if pend:
            st.warning(f"Você tem {pend} conta(s) a justificar em {MESES[mes]}/{ano} — acesse o menu Justificativas.")
        else:
            st.success(f"Nenhuma justificativa pendente em {MESES[mes]}/{ano}.")

    # ---------- submenu de seções (uma por vez -> menos rolagem, menos carga) ----------
    secoes_int = ["📌 Resumo", "📈 Evolução mensal", "📅 Valores (mês a mês)", "🔎 Desvios por CR"]
    if mostrar_justif:
        secoes_int.append("📝 Justificativas")
    secao = st.radio("Seção", secoes_int, horizontal=True, key="acomp_secao", label_visibility="collapsed")
    st.divider()

    if secao.endswith("Resumo"):
        resumo_colunas(d_mes, d_ytd, banda, mes, ano)
        contadores(d_mes, banda)
        st.caption("Convenção: contas de receita/dedução (código 3 ou 6) têm sinal invertido — a variação mede o IMPACTO no resultado. "
                   "Verde = favorável, vermelho = desfavorável, cinza = dentro da faixa neutra (±"
                   + f"{banda:.1f}".replace(".", ",") + "%).")
    elif "Evolução" in secao:
        st.markdown("#### Evolução mensal — Jan a Dez (mês e acumulado YTD)")
        hist_m = None
        if st.checkbox(f"Mostrar realizado {ano-1} (mês a mês)", value=False, key="evo_hist",
                       help="Adiciona, ao lado de Realizado, o realizado do mesmo mês no ano anterior — no mesmo recorte de filtros."):
            keys = {(int(r["uni_cod"]), int(r["cr_cod"]), int(r["conta_cod"])) for _, r in df.iterrows()}
            hist_m = {m: 0.0 for m in range(1, 13)}
            for r in (carregar_orc(ano - 1) or []):
                if (int(r.get("uni_cod", 0) or 0), int(r.get("cr_cod", 0) or 0), int(r.get("conta_cod", 0) or 0)) in keys:
                    mm = int(r.get("mes", 0) or 0)
                    if 1 <= mm <= 12:
                        hist_m[mm] += float(r.get("valor_realizado") or 0)
        tabela_evolucao(df, banda, mes, hist_m, f"Realizado {ano-1}")
    elif "Valores" in secao:
        st.markdown(f"#### Valores mês a mês — Planejado × Realizado × Histórico ({ano-1})")
        secao_valores(c, df, ano, mes)
    elif "Desvios" in secao:
        st.markdown(f"#### Desvios por centro de resultado — {MESES[mes]}/{ano}")
        drill_desvios(d_mes, banda, mes)
    else:
        st.markdown(f"#### Justificativas · {MESES[mes]}/{ano}")
        secao_justificativas(c, prof, d_mes, mes, is_ctrl, banda, ano)

# ---------------------------------------------------------------- planejamento (orçamento pelo gestor)
@fragment
def _plan_grid_frag(c, prof, ano, uni_cod, cr_cod, cr_nome, contas, plan_rows, editavel, hist=None, membros_key=None):
    hist = hist or {}
    if not contas:
        st.info("Nenhuma conta disponível para lançar. Importe/atualize o plano de contas.")
        return
    idx = {int(r["conta_cod"]): r for r in plan_rows
           if int(r.get("uni_cod", 0) or 0) == uni_cod and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("conta_cod") is not None}
    hm_on = st.checkbox(f"Histórico mês a mês ({ano-1})", value=False, key=f"plan_hm_{uni_cod}_{cr_cod}",
                        help="Mostra, ao lado de cada mês, o realizado do mesmo mês no ano anterior (referência, não editável).")
    data = {"Conta": [f"{cod} · {desc}" for cod, desc in contas],
            f"Histórico {ano-1}": [int(round(float(hist.get(cod, {}).get("total", 0.0)))) for cod, _ in contas]}
    ant_cols = []
    for mi in range(1, 13):
        data[MABREV[mi]] = [int(round(float((idx.get(cod, {}) or {}).get(f"m{mi}") or 0))) for cod, _ in contas]
        if hm_on:
            an = f"{MABREV[mi]} ant."
            ant_cols.append(an)
            data[an] = [int(round(float((hist.get(cod, {}).get("m", {}) or {}).get(mi, 0.0)))) for cod, _ in contas]
    if editavel:
        data["Remover"] = [False for _ in contas]
    df = pd.DataFrame(data)
    for cn in [f"Histórico {ano-1}"] + [MABREV[mi] for mi in range(1, 13)] + ant_cols:
        df[cn] = pd.to_numeric(df[cn], errors="coerce").fillna(0).astype(int)
    if editavel:
        df["Remover"] = df["Remover"].astype(bool)
    colcfg = {f"Histórico {ano-1}": st.column_config.NumberColumn(f"Histórico {ano-1} (R$)", format="%.0f", help="Realizado do ano anterior (total do ano) — referência, não editável")}
    colcfg.update({MABREV[mi]: st.column_config.NumberColumn(f"{MABREV[mi]} (R$)", format="%.0f", step=1, help="Digite o valor planejado deste mês") for mi in range(1, 13)})
    colcfg.update({an: st.column_config.NumberColumn(f"{an} (R$)", format="%.0f", help=f"Realizado deste mês em {ano-1} — referência, não editável") for an in ant_cols})
    if editavel:
        colcfg["Remover"] = st.column_config.CheckboxColumn("Remover", help="Marque para tirar a conta da grade: zera os 12 meses e some ao salvar.")
    disabled = ["Conta", f"Histórico {ano-1}"] + ant_cols + ([] if editavel else [MABREV[mi] for mi in range(1, 13)])
    if editavel:
        st.markdown(
            "<div style='background:#EEF2F8;border-left:4px solid " + AZUL_CORP + ";padding:9px 13px;"
            "border-radius:6px;margin:2px 0 8px;font-size:14px;color:#1f2b45;'>"
            f"✍️ <b>Digite aqui:</b> em cada linha (conta), preencha o valor planejado de <b>Janeiro a Dezembro</b>. "
            f"A coluna <b>Histórico {ano-1}</b> mostra o realizado do ano anterior como referência (não editável); "
            "as colunas dos meses (fundo branco) são editáveis.</div>",
            unsafe_allow_html=True)
    else:
        st.caption("🔒 Somente leitura (janela fechada ou já enviado/aprovado).")
    ed = st.data_editor(df, key=f"plan_{ano}_{uni_cod}_{cr_cod}_{len(contas)}", hide_index=True, use_container_width=True,
                        num_rows="fixed", disabled=disabled, column_config=colcfg)
    if not editavel:
        return
    b = st.columns([1, 1.3, 1.5, 2])
    salvar = b[0].button("💾 Salvar rascunho", key=f"plan_sv_{uni_cod}_{cr_cod}")
    enviar = b[1].button("📤 Finalizar e enviar este CR", key=f"plan_en_{uni_cod}_{cr_cod}", type="primary")
    copiar = b[2].button(f"📋 Copiar {ano-1}", key=f"plan_cp_{uni_cod}_{cr_cod}",
                         help=f"Preenche os 12 meses com o realizado de {ano-1} (das contas que têm histórico). Depois é só ajustar.")
    if copiar:
        ok = False
        try:
            n = 0
            for cod, desc in contas:
                hm = hist.get(cod, {}).get("m", {})
                if not any(abs(float(v or 0)) > 0.005 for v in hm.values()):
                    continue  # sem histórico -> não cria linha
                row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                       "conta_cod": int(cod), "conta_desc": desc, "atualizado_por": prof.get("nome", "")}
                for mi in range(1, 13):
                    row[f"m{mi}"] = round(float(hm.get(mi, 0) or 0))
                c.table("orc_plan").upsert(row, on_conflict="ano,uni_cod,cr_cod,conta_cod").execute(); n += 1
            c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "status": "RASCUNHO",
                                               "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
            ok = True
        except Exception:
            st.error("Não foi possível copiar o ano anterior agora.")
        if ok:
            limpar_cache(); st.success(f"{ano-1} copiado para {n} conta(s). Ajuste os meses e envie."); st.rerun()
    if salvar or enviar:
        gravou = 0; removidos = []
        for i, (cod, desc) in enumerate(contas):
            remover = bool(ed["Remover"].iloc[i]) if "Remover" in ed.columns else False
            if remover:
                row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                       "conta_cod": int(cod), "conta_desc": desc, "atualizado_por": prof.get("nome", ""), "ativo": False}
                for mi in range(1, 13):
                    row[f"m{mi}"] = 0
                _upsert_soft(c, "orc_plan", row, "ano,uni_cod,cr_cod,conta_cod")
                removidos.append(int(cod))
                continue
            vals = {f"m{mi}": round(float(ed[MABREV[mi]].iloc[i] or 0)) for mi in range(1, 13)}
            tem_valor = any(v != 0 for v in vals.values())
            ja_existe = int(cod) in idx  # conta já lançada antes neste CR
            if not tem_valor and not ja_existe:
                continue  # não cria linha zerada para conta nunca usada (mantém o orc_plan enxuto)
            row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                   "conta_cod": int(cod), "conta_desc": desc, "atualizado_por": prof.get("nome", ""), "ativo": True, **vals}
            _upsert_soft(c, "orc_plan", row, "ano,uni_cod,cr_cod,conta_cod")
            gravou += 1
        pend = []
        if enviar:
            try:
                cols = "conta_cod,conta_desc,justificativa," + ",".join(f"m{mi}" for mi in range(1, 13))
                fresh = (c.table("orc_plan").select(cols).eq("ano", ano).eq("uni_cod", uni_cod)
                         .eq("cr_cod", cr_cod).execute().data or [])
            except Exception:
                fresh = []
            for r in fresh:
                planned = sum(float(r.get(f"m{mi}") or 0) for mi in range(1, 13))
                histv = float((hist.get(int(r.get("conta_cod") or 0), {}) or {}).get("total", 0.0))
                if _exige_justif_orc(planned, histv) and not str(r.get("justificativa") or "").strip():
                    pend.append(f"{int(r.get('conta_cod') or 0)} · {r.get('conta_desc','')}")
        if enviar and pend:
            # valores salvos, mas NÃO envia: mantém rascunho e cobra as justificativas obrigatórias
            c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "status": "RASCUNHO",
                                               "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
            limpar_cache()
            st.error("Não é possível finalizar: contas com orçado acima de +5% do histórico sem justificativa — "
                     + "; ".join(pend) + ". Vá em “2) Justificativas por conta”, justifique e reenvie.")
            st.stop()
        novo = "ENVIADO" if enviar else "RASCUNHO"
        c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "status": novo,
                                           "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        if enviar:
            st.success("CR finalizado e enviado à controladoria. Para substituir algum valor agora, "
                       "a controladoria precisa devolver o CR — aí você ajusta e reenvia.")
        else:
            st.success("Rascunho salvo.")
        st.rerun()

def secao_justif_orcamento(c, prof, ano, uni_cod, cr_cod, plan_rows, hist, editavel):
    """Justificativas do orçamento por conta contábil (mesma cara do fluxo de justificativas).
    Sobem junto no orc_plan.justificativa. Obrigatória quando o orçado excede +5% do histórico."""
    rows = [r for r in plan_rows if int(r.get("uni_cod", 0) or 0) == uni_cod
            and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("conta_cod") is not None]
    itens = []
    for r in rows:
        planned = sum(float(r.get(f"m{mi}") or 0) for mi in range(1, 13))
        if planned == 0:
            continue
        cod = int(r["conta_cod"]); histv = float((hist.get(cod, {}) or {}).get("total", 0.0))
        itens.append((cod, r.get("conta_desc", "") or "", planned, histv, str(r.get("justificativa") or "")))
    if not itens:
        st.info("Nenhuma conta com valor orçado neste centro ainda. Preencha os valores na etapa “Valores” primeiro.")
        return
    itens.sort(key=lambda x: -x[2])
    obrig_pend = [it for it in itens if _exige_justif_orc(it[2], it[3]) and not it[4].strip()]
    st.caption("Justifique as despesas orçadas por conta. A justificativa é obrigatória quando o orçado excede em "
               f"mais de 5% o realizado de {ano-1}. Ela sobe junto para a controladoria analisar na aprovação.")
    if obrig_pend:
        st.warning(f"{len(obrig_pend)} conta(s) acima de +5% do histórico ainda sem justificativa — obrigatórias para enviar o CR.")
    for cod, desc, planned, histv, just in itens:
        req = _exige_justif_orc(planned, histv)
        varpct = ((planned - histv) / histv * 100) if histv else (100.0 if planned else 0.0)
        if just.strip():
            tag = chip("Justificada", VERDE)
        elif req:
            tag = chip("Obrigatória", VERMELHO)
        else:
            tag = chip("Opcional", CINZA_TXT)
        titulo = f"{cod} · {desc} — Orçado {brl(planned)} · Hist {ano-1} {brl(histv)} ({pct_txt(varpct)})"
        with st.expander(titulo):
            st.markdown(tag, unsafe_allow_html=True)
            if editavel:
                txt = st.text_area("Justificativa da conta", value=just, key=f"ojust_{uni_cod}_{cr_cod}_{cod}")
                if st.button("Salvar justificativa", key=f"ojust_sv_{uni_cod}_{cr_cod}_{cod}"):
                    try:
                        c.table("orc_plan").update({"justificativa": txt}).match(
                            {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "conta_cod": cod}).execute()
                        limpar_cache(); st.success("Justificativa salva."); st.rerun()
                    except Exception as e:
                        st.error(f"Não foi possível salvar (a coluna 'justificativa' existe no orc_plan?): {e}")
            else:
                st.info(just or "— (sem justificativa)")
                st.caption("🔒 Não editável (janela fechada ou já enviado/aprovado).")

def tela_planejamento_gestor(c, prof, ano):
    st.markdown(f"<div class='modtag'>Planejamento do Orçamento {ano}</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Preencha o orçamento por conta e mês nos seus centros de resultado. Salve rascunho e envie para a controladoria.</div>", unsafe_allow_html=True)
    anos_hab = get_plan_anos(c)
    if ano not in anos_hab:
        st.info(f"O planejamento está habilitado apenas para: {', '.join(map(str, anos_hab))}. "
                "Selecione um desses anos no seletor **Ano**, no topo da página.")
        return
    aberta = get_plan_janela(c, ano)
    EMP = {1: "PISA", 2: "KING"}
    ref = carregar_orc(ano - 1) or []
    orc_cur = carregar_orc(ano) or []
    uni_nome = {int(r["uni_cod"]): (r.get("unidade") or "") for r in (ref + orc_cur) if r.get("uni_cod") is not None}
    # CRs do gestor a partir do CADASTRO (cr_gestor): SOMENTE os vinculados ao gestor logado
    meu = str(prof.get("gestor_codigo"))
    crmap = {}
    for r in (carregar_cr_gestor() or []):
        if r.get("cr_cod") is not None and r.get("uni_cod") is not None and str(r.get("gestor_codigo")) == meu:
            crmap[(int(r["uni_cod"]), int(r["cr_cod"]))] = r.get("cr_nome", "") or ""
    crs = sorted(crmap.keys())
    if not crs:
        st.info("Você não tem centros de resultado vinculados. Fale com a controladoria (cadastro de CRs por gestor).")
        return
    # 1) Empresa  ->  2) Centro de resultado (da empresa escolhida)
    unis_disp = sorted({k[0] for k in crs})
    cse = st.columns([1.2, 3])
    uni_cod = cse[0].selectbox("1) Empresa", unis_disp,
                               format_func=lambda u: (uni_nome.get(u) or EMP.get(u, str(u))), key="plan_uni")
    crs_emp = [k for k in crs if k[0] == uni_cod]
    if not crs_emp:
        st.info("Nenhum centro de resultado atribuído a você nesta empresa.")
        return
    cr_opt = cse[1].selectbox("2) Centro de resultado", crs_emp, format_func=lambda k: f"{k[1]} · {crmap[k]}", key="plan_cr")
    uni_cod, cr_cod = cr_opt
    cr_nome = crmap[cr_opt]
    # CONTAS = união de TODAS as contas já vistas (plano_contas + orçado do ano e do ano-1)
    contas_dict = {}
    for x in (carregar_plano_contas() or []):
        if x.get("conta_cod") is not None and x.get("ativo", True):
            contas_dict[int(x["conta_cod"])] = x.get("conta_desc", "") or ""
    for r in (orc_cur + ref):
        cd = r.get("conta_cod")
        if cd is not None:
            contas_dict.setdefault(int(cd), r.get("conta_desc", "") or "")
    contas_full = sorted((cod, desc) for cod, desc in contas_dict.items())
    plan_rows = carregar_orc_plan(ano)
    stt = {(int(s["uni_cod"]), int(s["cr_cod"])): s.get("status", "RASCUNHO") for s in carregar_orc_plan_status(ano)}
    status = stt.get((uni_cod, cr_cod), "RASCUNHO")
    cor = {"RASCUNHO": CINZA_TXT, "ENVIADO": AZUL_CORP, "APROVADO": VERDE, "DEVOLVIDO": VERMELHO}.get(status, CINZA_TXT)
    st.markdown(f"Selecionado: <b>{EMP.get(uni_cod, uni_cod)} · {cr_cod} · {cr_nome}</b> &nbsp;·&nbsp; Situação: {chip(status, cor)}", unsafe_allow_html=True)
    if not aberta and status != "APROVADO":
        st.info(f"A janela de preenchimento do orçamento {ano} está fechada. Você pode consultar, mas não editar.")
    if status == "APROVADO":
        st.success("Este centro de resultado já foi aprovado — em modo leitura.")
    if status == "DEVOLVIDO":
        st.warning("Devolvido pela controladoria para ajuste. Corrija e envie novamente.")
    editavel = aberta and status not in ("APROVADO", "ENVIADO")
    # histórico do ano anterior (realizado) por conta, para este CR
    hist = {}
    for r in ref:
        if int(r.get("uni_cod", 0) or 0) == uni_cod and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("conta_cod") is not None:
            cod = int(r["conta_cod"]); mm = int(r.get("mes", 0) or 0)
            d = hist.setdefault(cod, {"m": {k: 0.0 for k in range(1, 13)}})
            if 1 <= mm <= 12:
                d["m"][mm] += float(r.get("valor_realizado") or 0)
    for cod in hist:
        hist[cod]["total"] = sum(hist[cod]["m"].values())
    # histórico (realizado ano-1) por (empresa, CR, conta) — para a regra de obrigatoriedade no envio em lote
    hist_tot = {}
    for r in ref:
        if r.get("conta_cod") is not None:
            k = (int(r.get("uni_cod", 0) or 0), int(r.get("cr_cod", 0) or 0), int(r["conta_cod"]))
            hist_tot[k] = hist_tot.get(k, 0.0) + float(r.get("valor_realizado") or 0)

    # ----- grade em uso + adicionar conta (baseada no banco, igual ao QLP) -----
    _plan_cr = [r for r in plan_rows if int(r.get("uni_cod", 0) or 0) == uni_cod
                and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("conta_cod") is not None]
    _removidas_c = {int(r["conta_cod"]) for r in _plan_cr if r.get("ativo") is False}
    _ativas_c = {int(r["conta_cod"]) for r in _plan_cr if r.get("ativo") is not False}
    membros_c = sorted((_ativas_c | set(hist.keys())) - _removidas_c)
    if editavel:
        faltantes = sorted([cod for cod, _ in contas_full if cod not in set(membros_c)])
        selkey = f"plan_add_sel_{uni_cod}_{cr_cod}"

        def _plan_add():
            s = st.session_state.get(selkey, "")
            if s != "" and int(s) not in set(membros_c):
                row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                       "conta_cod": int(s), "conta_desc": contas_dict.get(int(s), ""), "atualizado_por": prof.get("nome", ""), "ativo": True}
                for mi in range(1, 13):
                    row[f"m{mi}"] = 0
                try:
                    _upsert_soft(c, "orc_plan", row, "ano,uni_cod,cr_cod,conta_cod")
                    limpar_cache()
                except Exception:
                    pass
            st.session_state[selkey] = ""

        addc = st.columns([3.4, 1.1])
        addc[0].selectbox("➕ Adicionar conta à grade (plano de contas completo)", [""] + faltantes,
                          format_func=lambda cc: "— selecione uma conta —" if cc == "" else f"{cc} · {contas_dict.get(cc, '')}",
                          key=selkey,
                          help="Todas as contas da base, em ordem. Escolha e clique em Adicionar para incluir na grade.")
        addc[1].button("Adicionar conta", key=f"plan_add_btn_{uni_cod}_{cr_cod}", on_click=_plan_add)
        if not faltantes:
            st.caption("Todas as contas da base já estão na grade.")
    q = st.text_input("Filtrar conta (código ou nome) — opcional", key="plan_filtro")
    contas = [(cod, contas_dict.get(cod, "")) for cod in membros_c if (not q) or q.lower() in f"{cod} {contas_dict.get(cod,'')}".lower()]

    etapa = st.radio("Etapa", ["1) Valores", "2) Justificativas por conta"], horizontal=True, key="plan_etapa")
    if etapa.startswith("1"):
        st.markdown(f"<div style='font-weight:600;color:{AZUL_PROFUNDO};margin-top:6px'>Preencha os valores por mês</div>"
                    f"<div style='font-size:12px;color:{CINZA_TXT}'>A grade começa com as contas em uso e o histórico de {ano-1}. "
                    f"Use <b>➕ Adicionar conta</b> (acima) para trazer qualquer conta da base ({len(contas_full)} no total) e "
                    f"<b>Remover</b> para tirar da grade. <b>Histórico {ano-1}</b> = realizado do ano anterior. "
                    f"Contas em zero não geram lançamento. Depois vá em “2) Justificativas por conta”.</div>", unsafe_allow_html=True)
        _plan_grid_frag(c, prof, ano, uni_cod, cr_cod, cr_nome, contas, plan_rows, editavel, hist)
    else:
        secao_justif_orcamento(c, prof, ano, uni_cod, cr_cod, plan_rows, hist, editavel)

    # ---------- Finalizar orçamento: resumo dos CRs do gestor + envio em lote ----------
    st.divider()
    with st.container(border=True):
        st.markdown(f"<div style='font-weight:700;color:{AZUL_PROFUNDO}'>✅ Finalizar orçamento {ano}</div>", unsafe_allow_html=True)
        st.caption("Resumo de todos os seus centros de resultado. Você pode enviar de uma vez todos os que estão em "
                   "rascunho ou devolvidos (e já têm valores lançados), ou usar o botão de enviar de cada CR acima. "
                   "Depois de finalizado, para substituir um valor a controladoria precisa devolver o CR — aí você ajusta e reenvia.")
        # total orçado por (uni, cr) a partir do orc_plan já salvo
        tot_plan = {}
        for rr in plan_rows:
            k = (int(rr.get("uni_cod", 0) or 0), int(rr.get("cr_cod", 0) or 0))
            tot_plan[k] = tot_plan.get(k, 0.0) + sum(float(rr.get(f"m{mi}") or 0) for mi in range(1, 13))
        nome_cr = dict(crmap)
        for rr in plan_rows:
            k = (int(rr.get("uni_cod", 0) or 0), int(rr.get("cr_cod", 0) or 0))
            nome_cr.setdefault(k, rr.get("cr_nome", ""))
        def _fmt2(k):
            emp = uni_nome.get(k[0]) or EMP.get(k[0], str(k[0]))
            return f"{emp} · {k[1]} · {nome_cr.get(k, '')}"
        cor_st = {"RASCUNHO": CINZA_TXT, "ENVIADO": AZUL_CORP, "APROVADO": VERDE,
                  "DEVOLVIDO": VERMELHO, "SEM LANÇAMENTO": CINZA_TXT}
        todos_cr = sorted(set(crs) | set(tot_plan.keys()))
        linhas = ""; finalizaveis = []
        for k in todos_cr:
            s = stt.get(k, "RASCUNHO")
            tot = tot_plan.get(k, 0.0)
            tem = abs(tot) > 0.005 or k in tot_plan
            s_disp = s if tem else "SEM LANÇAMENTO"
            if tem and s in ("RASCUNHO", "DEVOLVIDO"):
                finalizaveis.append(k)
            linhas += (f"<tr><td style='text-align:left'>{_fmt2(k)}</td>"
                       f"<td style='text-align:center'>{chip(s_disp, cor_st.get(s_disp, CINZA_TXT))}</td>"
                       f"<td>{brl(tot)}</td></tr>")
        st.markdown(f"<div class='scroll'><table class='lle'><tr><th style='text-align:left'>Centro de resultado</th>"
                    f"<th style='text-align:center'>Situação</th><th>Total orçado</th></tr>{linhas}</table></div>",
                    unsafe_allow_html=True)
        n_fin = len(finalizaveis)
        if not aberta:
            st.info("Janela de preenchimento fechada — não é possível finalizar/enviar agora.")
        elif n_fin == 0:
            st.caption("Nenhum CR em rascunho/devolvido com valores lançados para finalizar. "
                       "(CRs sem lançamento ou já enviados/aprovados são ignorados.)")
        else:
            if st.button(f"📤 Finalizar e enviar {n_fin} CR(s) para a controladoria", key="plan_fin_todos", type="primary"):
                enviados = []; retidos = []
                for (u, cc) in finalizaveis:
                    pend = 0
                    for r in plan_rows:
                        if int(r.get("uni_cod", 0) or 0) == u and int(r.get("cr_cod", 0) or 0) == cc and r.get("conta_cod") is not None:
                            planned = sum(float(r.get(f"m{mi}") or 0) for mi in range(1, 13))
                            histv = hist_tot.get((u, cc, int(r["conta_cod"])), 0.0)
                            if _exige_justif_orc(planned, histv) and not str(r.get("justificativa") or "").strip():
                                pend += 1
                    if pend:
                        retidos.append(f"{cc} · {nome_cr.get((u, cc), '')} ({pend} sem justificativa)")
                        continue
                    c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": u, "cr_cod": cc, "status": "ENVIADO",
                                                       "atualizado_por": prof.get("nome", "")},
                                                      on_conflict="ano,uni_cod,cr_cod").execute()
                    enviados.append(f"{cc} · {nome_cr.get((u, cc), '')}")
                limpar_cache()
                if enviados:
                    st.success("Finalizado(s) e enviado(s): " + "; ".join(enviados) +
                               ". Para substituir algo, peça à controladoria para devolver o CR.")
                if retidos:
                    st.error("Não enviados (faltam justificativas obrigatórias em contas acima de +5% do histórico): "
                             + "; ".join(retidos) + ". Justifique na etapa “2) Justificativas por conta” de cada CR.")
                st.rerun()

def _consolidar_plan(c, ano, uni, cr, cr_nome, plan_cr, ref):
    """Grava o planejamento aprovado em orc_realizado[valor_planejado] (12 linhas por conta).
       Preserva valor_realizado (não é enviado no payload)."""
    EMP = {1: "PISA", 2: "KING"}
    unidade = next((x.get("unidade", "") for x in ref if int(x.get("uni_cod", 0) or 0) == uni and x.get("unidade")), EMP.get(uni, str(uni)))
    cr_grupo = next((x.get("cr_grupo", "") for x in ref if int(x.get("uni_cod", 0) or 0) == uni and int(x.get("cr_cod", 0) or 0) == cr), "")
    meta = {int(x["conta_cod"]): {"tipo_conta": x.get("tipo_conta", ""), "classificacao": x.get("classificacao", "")}
            for x in ref if x.get("conta_cod") is not None}
    payloads = []
    for r in plan_cr:
        cod = int(r["conta_cod"]); m = meta.get(cod, {})
        for mi in range(1, 13):
            payloads.append(dict(ano=ano, mes=mi, uni_cod=uni, unidade=unidade, cr_cod=cr, cr_nome=cr_nome,
                                 cr_grupo=cr_grupo, conta_cod=cod, conta_desc=r.get("conta_desc", ""),
                                 tipo_conta=m.get("tipo_conta", ""), classificacao=m.get("classificacao", ""),
                                 valor_planejado=round(float(r.get(f"m{mi}") or 0))))
    if payloads:
        c.table("orc_realizado").upsert(payloads, on_conflict="ano,mes,uni_cod,cr_cod,conta_cod").execute()

def _painel_finalizacao_gestor(status_rows, ano, modulo):
    """Painel da controladoria: por gestor, quantos CRs foram finalizados (ENVIADO/APROVADO)
       e quantos ainda faltam. Cruza cr_gestor (universo esperado) com o status do módulo."""
    stt = {(int(s["uni_cod"]), int(s["cr_cod"])): s.get("status", "") for s in (status_rows or [])}
    _EMP = {1: "PISA", 2: "KING"}
    porg = {}
    for r in (carregar_cr_gestor() or []):
        u = int(r.get("uni_cod", 0) or 0); crc = int(r.get("cr_cod", 0) or 0)
        gn = (r.get("gestor") or {}).get("nome", "") if isinstance(r.get("gestor"), dict) else ""
        gn = gn or "— (sem gestor)"
        s = stt.get((u, crc), "NÃO INICIADO")
        d = porg.setdefault(gn, {"crs": [], "fin": 0, "pend": 0})
        d["crs"].append((u, crc, r.get("cr_nome", ""), s))
        if s in ("ENVIADO", "APROVADO"):
            d["fin"] += 1
        else:
            d["pend"] += 1
    if not porg:
        st.caption("Nenhum vínculo gestor↔CR cadastrado para acompanhar a finalização.")
        return
    n_ok = sum(1 for g in porg if porg[g]["pend"] == 0)
    kc = st.columns(3)
    kc[0].metric("Gestores", len(porg))
    kc[1].metric(f"Finalizaram o {modulo}", n_ok)
    kc[2].metric("Ainda faltam", len(porg) - n_ok)
    cor_st = {"APROVADO": VERDE, "ENVIADO": AZUL_CORP, "RASCUNHO": "#EF9F27",
              "DEVOLVIDO": VERMELHO, "NÃO INICIADO": VERMELHO}
    for gn in sorted(porg, key=lambda g: (porg[g]["pend"] == 0, g)):
        d = porg[gn]
        tag = "✅ finalizou tudo" if d["pend"] == 0 else f"⏳ faltam {d['pend']} de {len(d['crs'])}"
        with st.expander(f"{gn} — {d['fin']}/{len(d['crs'])} CR(s) finalizado(s) · {tag}"):
            linhas = ""
            for u, crc, crn, s in sorted(d["crs"], key=lambda x: x[1]):
                linhas += (f"<tr><td style='text-align:left'>{_EMP.get(u, u)} · {crc} · {crn}</td>"
                           f"<td style='text-align:center'>{chip(s, cor_st.get(s, CINZA_TXT))}</td></tr>")
            st.markdown(f"<div class='scroll'><table class='lle'><tr><th style='text-align:left'>Centro de resultado</th>"
                        f"<th style='text-align:center'>Situação</th></tr>{linhas}</table></div>", unsafe_allow_html=True)

def tela_planejamento_ctrl(c, prof, ano):
    st.markdown(f"<div class='modtag'>Planejamento do Orçamento {ano}</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Abra/feche a janela, revise os envios dos gestores, aprove (consolida no orçado) ou devolva para ajuste.</div>", unsafe_allow_html=True)
    anos_hab = get_plan_anos(c)
    with st.expander("⚙️ Anos habilitados para planejamento"):
        st.caption("Para quais anos os gestores podem preencher o orçamento. Ex.: 2027. No próximo ano, inclua 2028 (e assim por diante).")
        txt = st.text_input("Anos habilitados (separados por vírgula)", value=", ".join(map(str, anos_hab)), key="plan_anos_txt")
        if st.button("Salvar anos habilitados", key="plan_anos_save"):
            novos = set_plan_anos(c, txt)
            st.success(f"Habilitado(s): {', '.join(map(str, novos))}."); st.rerun()
    if ano not in anos_hab:
        st.info(f"O planejamento está habilitado apenas para: {', '.join(map(str, anos_hab))}. "
                "Selecione um ano habilitado no topo (**Ano**) para abrir a janela e revisar, "
                "ou inclua o ano acima em **Anos habilitados**.")
        return
    aberta = get_plan_janela(c, ano)
    jc = st.columns([3, 1.4])
    jc[0].markdown(f"<div style='padding-top:8px'>Janela de preenchimento {ano}: "
                   f"<b style='color:{VERDE if aberta else VERMELHO}'>{'ABERTA' if aberta else 'FECHADA'}</b></div>", unsafe_allow_html=True)
    if jc[1].button(("🔒 Fechar janela" if aberta else "🔓 Abrir janela"), key="plan_toggle", use_container_width=True):
        set_plan_janela(c, ano, not aberta); st.rerun()
    if st.button("🔄 Sincronizar plano de contas", key="plan_sync"):
        vistos = {}
        for yr in (ano, ano - 1):
            for x in (carregar_orc(yr) or []):
                if x.get("conta_cod") is not None:
                    vistos[int(x["conta_cod"])] = x.get("conta_desc", "") or ""
        if not vistos:
            st.info("Não há contas em orc_realizado para sincronizar.")
        else:
            try:
                # não depende de constraint UNIQUE: lê o que já existe e insere só o que falta
                existentes = {int(x["conta_cod"]) for x in (carregar_plano_contas() or []) if x.get("conta_cod") is not None}
                novos = [{"conta_cod": k, "conta_desc": v} for k, v in vistos.items() if k not in existentes]
                for ch in chunks(novos, 500):
                    c.table("plano_contas").insert(ch).execute()
                limpar_cache()
                st.success(f"Catálogo sincronizado: {len(novos)} conta(s) nova(s) adicionada(s); {len(existentes)} já existiam.")
                st.rerun()
            except Exception as e:
                st.error(f"Não foi possível sincronizar o plano de contas: {e}")
    st.divider()

    status_rows = carregar_orc_plan_status(ano)
    with st.expander("📋 Finalização por gestor (orçamento)", expanded=True):
        _painel_finalizacao_gestor(status_rows, ano, "orçamento")
    st.divider()
    if not status_rows:
        st.info("Nenhum gestor iniciou o preenchimento ainda.")
        return
    stt = {(int(s["uni_cod"]), int(s["cr_cod"])): s for s in status_rows}
    corcls = {"RASCUNHO": CINZA_TXT, "ENVIADO": AZUL_CORP, "APROVADO": VERDE, "DEVOLVIDO": VERMELHO}

    ordem = {"ENVIADO": 0, "DEVOLVIDO": 1, "RASCUNHO": 2, "APROVADO": 3}
    opts = sorted(stt.keys(), key=lambda k: (ordem.get(stt[k].get("status", ""), 9), k[1]))
    sel = st.selectbox("Centro de resultado", opts,
                       format_func=lambda k: f"{k[1]} · {stt[k].get('status','')}", key="plan_rev_cr")
    uni, cr = sel
    srow = stt[sel]; status = srow.get("status", "RASCUNHO")
    cr_nome = ""
    plan_rows = carregar_orc_plan(ano)
    plan_cr = [r for r in plan_rows if int(r.get("uni_cod", 0) or 0) == uni and int(r.get("cr_cod", 0) or 0) == cr]
    if plan_cr: cr_nome = plan_cr[0].get("cr_nome", "")

    st.markdown(f"Situação: {chip(status, corcls.get(status, CINZA_TXT))}"
                + (f" &nbsp;·&nbsp; <span style='color:{CINZA_TXT}'>atualizado por {srow.get('atualizado_por','')}</span>" if srow.get('atualizado_por') else ""),
                unsafe_allow_html=True)
    if srow.get("comentario"):
        st.caption(f"Comentário registrado: {srow.get('comentario')}")

    # grade enviada (somente leitura) com total do ano
    if not plan_cr:
        st.info("Este CR ainda não tem valores lançados.")
    else:
        th = "<th style='text-align:left'>Conta</th>" + "".join(f"<th>{MABREV[m]}</th>" for m in range(1, 13)) + "<th>Total</th>"
        corpo = ""
        tot_geral = 0.0
        for r in sorted(plan_cr, key=lambda x: int(x.get("conta_cod", 0) or 0)):
            tot = sum(float(r.get(f"m{m}") or 0) for m in range(1, 13)); tot_geral += tot
            corpo += (f"<tr><td style='text-align:left'>{int(r['conta_cod'])} · {r.get('conta_desc','')}</td>"
                      + "".join(f"<td>{brl(float(r.get(f'm{m}') or 0))}</td>" for m in range(1, 13))
                      + f"<td><b>{brl(tot)}</b></td></tr>")
        corpo += (f"<tr class='mark'><td style='text-align:left'><b>Total do CR</b></td>"
                  + "".join("<td></td>" for _ in range(1, 13)) + f"<td><b>{brl(tot_geral)}</b></td></tr>")
        st.markdown(f"<div class='scroll'><table class='lle matrix'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)

        # justificativas do gestor por conta (subiram junto com o orçamento)
        just_rows = [r for r in plan_cr if str(r.get("justificativa") or "").strip()]
        with st.expander(f"📝 Justificativas do gestor por conta ({len(just_rows)})", expanded=bool(just_rows)):
            if not just_rows:
                st.caption("Nenhuma justificativa registrada pelo gestor neste CR.")
            else:
                for r in sorted(just_rows, key=lambda x: int(x.get("conta_cod", 0) or 0)):
                    tot = sum(float(r.get(f"m{m}") or 0) for m in range(1, 13))
                    st.markdown(f"<b>{int(r['conta_cod'])} · {r.get('conta_desc','')}</b> — orçado {brl(tot)}", unsafe_allow_html=True)
                    st.caption(str(r.get("justificativa") or ""))
    ref = carregar_orc(ano - 1) or carregar_orc(ano) or []
    orc_ano = carregar_orc(ano) or []
    conflito = any(int(x.get("uni_cod", 0) or 0) == uni and int(x.get("cr_cod", 0) or 0) == cr
                   and float(x.get("valor_planejado") or 0) != 0 for x in orc_ano)

    st.divider()
    if status == "APROVADO":
        st.success("CR já aprovado e consolidado no orçado. Se precisar reabrir, use Devolver.")
    ca = st.columns([1.4, 1.4, 3])
    ok_over = True
    if conflito and status != "APROVADO":
        st.warning(f"Já existe orçado lançado para o CR {cr} em {ano}. Aprovar vai SOBRESCREVER o planejado (o realizado é preservado).")
        ok_over = st.checkbox("Confirmo a sobrescrita", key=f"plan_over_{uni}_{cr}")
    aprovar = ca[0].button("✅ Aprovar e consolidar", key=f"plan_apr_{uni}_{cr}",
                           type="primary", disabled=(status == "APROVADO") or (conflito and not ok_over) or not plan_cr)
    coment = ca[2].text_input("Comentário (para devolução)", key=f"plan_com_{uni}_{cr}")
    devolver = ca[1].button("↩️ Devolver", key=f"plan_dev_{uni}_{cr}", disabled=not plan_cr)

    if aprovar:
        _consolidar_plan(c, ano, uni, cr, cr_nome, plan_cr, ref)
        c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": uni, "cr_cod": cr, "status": "APROVADO",
                                           "comentario": "", "atualizado_por": prof.get("nome", "")},
                                          on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        st.success(f"CR {cr} aprovado e consolidado no orçado de {ano}."); st.rerun()
    if devolver:
        c.table("orc_plan_status").upsert({"ano": ano, "uni_cod": uni, "cr_cod": cr, "status": "DEVOLVIDO",
                                           "comentario": coment, "atualizado_por": prof.get("nome", "")},
                                          on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        st.success(f"CR {cr} devolvido para ajuste."); st.rerun()

# ---------------------------------------------------------------- QLP (planejamento de pessoal / headcount)
def _qlp_grid_frag(c, prof, ano, uni_cod, cr_cod, cr_nome, cargos, plan_rows, editavel, hist=None, membros_key=None):
    """Grade de headcount (quantidade) por cargo × 12 meses. Espelha o planejamento do orçamento.
       A grade começa com a estrutura do ano anterior + rascunho. A coluna 'Remover' zera os 12 meses
       e tira o cargo da grade (regra: quantidade > 0 = está no quadro; zerado = fora). Não apaga histórico."""
    hist = hist or {}
    if not cargos:
        # grade sem cargos: não renderiza o editor (colunas vazias quebram o data_editor)
        return
    idx = {str(r["cargo_cod"]): r for r in plan_rows
           if int(r.get("uni_cod", 0) or 0) == uni_cod and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("cargo_cod") is not None}
    data = {"Cargo": [f"{cod} · {nome}" for cod, nome in cargos],
            f"Hist. {ano-1}": [int(round(float(hist.get(str(cod), {}).get("total", 0.0)))) for cod, _ in cargos]}
    for mi in range(1, 13):
        data[MABREV[mi]] = [int(round(float((idx.get(str(cod), {}) or {}).get(f"m{mi}") or 0))) for cod, _ in cargos]
    if editavel:
        data["Remover"] = [False for _ in cargos]
    df = pd.DataFrame(data)
    # garante dtypes corretos (evita StreamlitAPIException de incompatibilidade de tipo)
    for cn in [f"Hist. {ano-1}"] + [MABREV[mi] for mi in range(1, 13)]:
        df[cn] = pd.to_numeric(df[cn], errors="coerce").fillna(0).astype(int)
    if editavel:
        df["Remover"] = df["Remover"].astype(bool)
    colcfg = {f"Hist. {ano-1}": st.column_config.NumberColumn(f"Hist. {ano-1} (HC)", format="%d", help="Headcount realizado do ano anterior (total do ano) — referência, não editável")}
    colcfg.update({MABREV[mi]: st.column_config.NumberColumn(f"{MABREV[mi]}", format="%d", step=1, min_value=0, help="Quantidade de funcionários planejada neste mês") for mi in range(1, 13)})
    if editavel:
        colcfg["Remover"] = st.column_config.CheckboxColumn("Remover", help="Marque para tirar o cargo do quadro: zera os 12 meses e some da grade. Não apaga o histórico.")
    disabled = ["Cargo", f"Hist. {ano-1}"] + ([] if editavel else [MABREV[mi] for mi in range(1, 13)])
    if editavel:
        st.markdown(
            "<div style='background:#EEF2F8;border-left:4px solid " + AZUL_CORP + ";padding:9px 13px;"
            "border-radius:6px;margin:2px 0 8px;font-size:14px;color:#1f2b45;'>"
            f"✍️ <b>Digite aqui:</b> em cada linha (cargo), preencha a <b>quantidade de funcionários</b> de Janeiro a Dezembro. "
            f"A coluna <b>Hist. {ano-1}</b> é referência (headcount do ano anterior). Para <b>tirar</b> um cargo do quadro, "
            "marque <b>Remover</b> (zera e some ao salvar). Para <b>incluir</b> um cargo novo, use o seletor acima da grade.</div>",
            unsafe_allow_html=True)
    else:
        st.caption("🔒 Somente leitura (janela fechada ou já enviado/aprovado).")
    ed = st.data_editor(df, key=f"qlp_{ano}_{uni_cod}_{cr_cod}_{len(cargos)}", hide_index=True, use_container_width=True,
                        num_rows="fixed", disabled=disabled, column_config=colcfg)
    if not editavel:
        return
    b = st.columns([1, 1.3, 1.5, 2])
    salvar = b[0].button("💾 Salvar rascunho", key=f"qlp_sv_{uni_cod}_{cr_cod}")
    enviar = b[1].button("📤 Finalizar e enviar este CR", key=f"qlp_en_{uni_cod}_{cr_cod}", type="primary")
    copiar = b[2].button(f"📋 Copiar {ano-1}", key=f"qlp_cp_{uni_cod}_{cr_cod}",
                         help=f"Preenche os 12 meses com o headcount realizado de {ano-1} (dos cargos que têm histórico).")
    if copiar:
        ok = False
        try:
            n = 0
            for cod, nome in cargos:
                hm = hist.get(str(cod), {}).get("m", {})
                if not any(int(round(float(v or 0))) > 0 for v in hm.values()):
                    continue
                row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                       "cargo_cod": str(cod), "cargo_nome": nome, "atualizado_por": prof.get("nome", "")}
                for mi in range(1, 13):
                    row[f"m{mi}"] = int(round(float(hm.get(mi, 0) or 0)))
                c.table("qlp_plan").upsert(row, on_conflict="ano,uni_cod,cr_cod,cargo_cod").execute(); n += 1
            c.table("qlp_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "status": "RASCUNHO",
                                               "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
            ok = True
        except Exception:
            st.error("Não foi possível copiar o ano anterior agora.")
        if ok:
            limpar_cache(); st.success(f"{ano-1} copiado para {n} cargo(s). Ajuste e envie."); st.rerun()
    if salvar or enviar:
        removidos = []
        for i, (cod, nome) in enumerate(cargos):
            remover = bool(ed["Remover"].iloc[i]) if "Remover" in ed.columns else False
            row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                   "cargo_cod": str(cod), "cargo_nome": nome, "atualizado_por": prof.get("nome", "")}
            if remover:
                for mi in range(1, 13):
                    row[f"m{mi}"] = 0
                row["ativo"] = False
                removidos.append(str(cod))
            else:
                for mi in range(1, 13):
                    row[f"m{mi}"] = int(round(float(ed[MABREV[mi]].iloc[i] or 0)))
                row["ativo"] = True
            _upsert_soft(c, "qlp_plan", row, "ano,uni_cod,cr_cod,cargo_cod")
        novo = "ENVIADO" if enviar else "RASCUNHO"
        c.table("qlp_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "status": novo,
                                           "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        if enviar:
            msg = ("QLP do CR finalizado e enviado à controladoria. Para substituir algo agora, "
                   "a controladoria precisa devolver o CR — aí você ajusta e reenvia.")
        else:
            msg = "Rascunho salvo."
        if removidos:
            msg += f" {len(removidos)} cargo(s) removido(s) do quadro."
        st.success(msg); st.rerun()

def tela_qlp_gestor(c, prof, ano):
    st.markdown(f"<div class='modtag'>Planejamento de Pessoal — QLP {ano}</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Quadro de Lotação de Pessoal: planeje a quantidade de funcionários por cargo e mês nos seus centros de resultado. Salve rascunho e envie para a controladoria.</div>", unsafe_allow_html=True)
    anos_hab = get_plan_anos(c)
    if ano not in anos_hab:
        st.info(f"O QLP está habilitado apenas para: {', '.join(map(str, anos_hab))}. Selecione um desses anos no seletor **Ano**, no topo.")
        return
    aberta = get_qlp_janela(c, ano)
    EMP = {1: "PISA", 2: "KING"}
    ref = carregar_hc_quadro(ano - 1) or []
    cur_hc = carregar_hc_quadro(ano) or []
    uni_nome = {int(r["uni_cod"]): (r.get("unidade") or "") for r in (ref + cur_hc) if r.get("uni_cod") is not None}
    # CRs do gestor a partir do CADASTRO (cr_gestor): SOMENTE os vinculados ao gestor logado
    meu = str(prof.get("gestor_codigo"))
    crmap = {}
    for r in (carregar_cr_gestor() or []):
        if r.get("cr_cod") is not None and r.get("uni_cod") is not None and str(r.get("gestor_codigo")) == meu:
            crmap[(int(r["uni_cod"]), int(r["cr_cod"]))] = r.get("cr_nome", "") or ""
    crs = sorted(crmap.keys())
    if not crs:
        st.info("Você não tem centros de resultado vinculados. Fale com a controladoria (cadastro de CRs por gestor).")
        return
    unis_disp = sorted({k[0] for k in crs})
    cse = st.columns([1.2, 3])
    uni_cod = cse[0].selectbox("1) Empresa", unis_disp, format_func=lambda u: (uni_nome.get(u) or EMP.get(u, str(u))), key="qlp_uni")
    crs_emp = [k for k in crs if k[0] == uni_cod]
    if not crs_emp:
        st.info("Nenhum centro de resultado atribuído a você nesta empresa.")
        return
    cr_opt = cse[1].selectbox("2) Centro de resultado", crs_emp, format_func=lambda k: f"{k[1]} · {crmap[k]}", key="qlp_cr")
    uni_cod, cr_cod = cr_opt
    cr_nome = crmap[cr_opt]
    # Catálogo do seletor "Adicionar cargo" = TODOS os cargos já existentes no sistema
    # (tabela de cargos + histórico de headcount de qualquer ano + planos). A grade COMEÇA com os
    # cargos deste CR; o seletor permite trazer qualquer cargo da empresa (lista de seleção ordenada).
    plan_rows = carregar_qlp_plan(ano)
    catalogo = {}
    for x in (carregar_hc_cargo() or []):
        if x.get("cargo_cod") is not None and x.get("ativo", True):
            catalogo[str(x["cargo_cod"])] = x.get("cargo_nome", "") or str(x["cargo_cod"])
    for cc, nm in (carregar_cargos_todos() or {}).items():
        catalogo.setdefault(cc, nm)
    for r in plan_rows:
        if r.get("cargo_cod") is not None:
            catalogo.setdefault(str(r["cargo_cod"]), (r.get("cargo_nome") or str(r["cargo_cod"])))
    cargos_cat = catalogo  # compatibilidade com o restante da função
    nomes = dict(catalogo)
    stt = {(int(s["uni_cod"]), int(s["cr_cod"])): s.get("status", "RASCUNHO") for s in carregar_qlp_status(ano)}
    status = stt.get((uni_cod, cr_cod), "RASCUNHO")
    cor = {"RASCUNHO": CINZA_TXT, "ENVIADO": AZUL_CORP, "APROVADO": VERDE, "DEVOLVIDO": VERMELHO}.get(status, CINZA_TXT)
    st.markdown(f"CR selecionado: <b>{cr_cod} · {cr_nome}</b> &nbsp;·&nbsp; Situação: {chip(status, cor)}", unsafe_allow_html=True)
    if not aberta and status != "APROVADO":
        st.info(f"A janela do QLP {ano} está fechada. Você pode consultar, mas não editar.")
    if status == "APROVADO":
        st.success("Este centro de resultado já foi aprovado — em modo leitura.")
    if status == "DEVOLVIDO":
        st.warning("Devolvido pela controladoria para ajuste. Corrija e envie novamente.")
    editavel = aberta and status not in ("APROVADO", "ENVIADO")

    # histórico do ano anterior (headcount realizado) por cargo, para este CR
    hist = {}
    for r in ref:
        if int(r.get("uni_cod", 0) or 0) == uni_cod and int(r.get("cr_cod", 0) or 0) == cr_cod and r.get("cargo_cod") is not None:
            cod = str(r["cargo_cod"]); mm = int(r.get("mes", 0) or 0)
            d = hist.setdefault(cod, {"m": {k: 0.0 for k in range(1, 13)}})
            if 1 <= mm <= 12:
                d["m"][mm] += float(r.get("qtd_realizada") or 0)
    for cod in hist:
        # headcount por mês é uma quantidade pontual; o "total do ano" aqui é o pico do ano (máximo mensal)
        hist[cod]["total"] = max(hist[cod]["m"].values()) if hist[cod]["m"] else 0.0

    # ----- membership da grade (SEM estado de sessão) -----
    # Se o CR já tem plano lançado, a grade reflete o PLANO; senão, começa pela estrutura do histórico.
    plan_cods = [str(r["cargo_cod"]) for r in plan_rows
                 if int(r.get("uni_cod", 0) or 0) == uni_cod and int(r.get("cr_cod", 0) or 0) == cr_cod
                 and r.get("cargo_cod") is not None and (r.get("ativo") is not False)]
    if plan_cods:
        membros = sorted(set(plan_cods), key=lambda x: ((nomes.get(x, x) or "").lower(), x))
    else:
        membros = sorted(set(str(cod) for cod in hist), key=lambda x: ((nomes.get(x, x) or "").lower(), x))

    # ----- adicionar cargo do catálogo/histórico à grade -----
    if editavel:
        faltantes = sorted([cod for cod in cargos_cat if cod not in set(membros)],
                           key=lambda cc: ((nomes.get(cc, cc) or "").lower(), cc))
        selkey = f"qlp_add_sel_{uni_cod}_{cr_cod}"

        def _qlp_add():
            s = st.session_state.get(selkey, "")
            if not s:
                st.session_state["qlp_add_msg"] = ("warn", "Nenhum cargo selecionado no seletor.")
                return
            if s in membros:
                st.session_state["qlp_add_msg"] = ("warn", f"O cargo {s} já está na grade.")
                st.session_state[selkey] = ""
                return
            row = {"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod, "cr_nome": cr_nome,
                   "cargo_cod": str(s), "cargo_nome": nomes.get(s, s), "atualizado_por": prof.get("nome", ""), "ativo": True}
            for mi in range(1, 13):
                row[f"m{mi}"] = 0
            try:
                _upsert_soft(c, "qlp_plan", row, "ano,uni_cod,cr_cod,cargo_cod")
                # garante status do CR (para aparecer no controle) sem sobrescrever se já existir
                try:
                    c.table("qlp_plan_status").upsert({"ano": ano, "uni_cod": uni_cod, "cr_cod": cr_cod,
                                                       "status": (stt.get((uni_cod, cr_cod)) or "RASCUNHO"),
                                                       "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
                except Exception:
                    pass
                limpar_cache()
                st.session_state["qlp_add_msg"] = ("ok", f"Cargo {s} · {nomes.get(s, s)} adicionado. Preencha o headcount e salve.")
            except Exception as e:
                st.session_state["qlp_add_msg"] = ("err", f"Falha ao gravar no qlp_plan: {e}")
            st.session_state[selkey] = ""

        addc = st.columns([3.4, 1.1])
        addc[0].selectbox("➕ Adicionar cargo à grade (todos os cargos da empresa)", [""] + faltantes,
                          format_func=lambda cc: "— selecione um cargo —" if cc == "" else f"{cc} · {nomes.get(cc, cc)}",
                          key=selkey,
                          help="Lista com todos os cargos existentes no sistema, em ordem alfabética. Jr/Pleno/Sr têm códigos próprios e aparecem separados. Escolha o que se aplica ao seu CR.")
        addc[1].button("Adicionar", key=f"qlp_add_btn_{uni_cod}_{cr_cod}", on_click=_qlp_add)
        _m = st.session_state.pop("qlp_add_msg", None)
        if _m:
            {"ok": st.success, "warn": st.warning, "err": st.error}.get(_m[0], st.info)(_m[1])
        if not faltantes:
            st.caption("Todos os cargos conhecidos já estão na grade. Cargos novos entram na base pela importação de pessoal.")

    q = st.text_input("Filtrar cargo (código ou nome) — opcional", key="qlp_filtro")
    cargos = [(cod, nomes.get(cod, cod)) for cod in membros if (not q) or q.lower() in f"{cod} {nomes.get(cod, '')}".lower()]
    st.markdown(f"<div style='font-weight:600;color:{AZUL_PROFUNDO};margin-top:6px'>2) Preencha o headcount por mês e 3) envie</div>"
                f"<div style='font-size:12px;color:{CINZA_TXT}'>Quadro atual — {len(cargos)} cargo(s) na grade "
                f"(catálogo disponível: {len(cargos_cat)}). A grade começa com a estrutura de {ano-1}; adicione cargos pelo seletor "
                f"e remova marcando <b>Remover</b>. <b>Hist. {ano-1}</b> = pico de headcount do ano anterior (referência).</div>", unsafe_allow_html=True)
    if not cargos:
        st.caption("Nenhum cargo na grade ainda. Use o seletor acima para adicionar cargos do catálogo.")
    _qlp_grid_frag(c, prof, ano, uni_cod, cr_cod, cr_nome, cargos, plan_rows, editavel, hist)

def _consolidar_qlp(c, ano, uni, cr, cr_nome, plan_cr, ref):
    """Grava o QLP aprovado em hc_quadro[qtd_orcada] (12 linhas por cargo). Preserva qtd_realizada."""
    EMP = {1: "PISA", 2: "KING"}
    unidade = next((x.get("unidade", "") for x in ref if int(x.get("uni_cod", 0) or 0) == uni and x.get("unidade")), EMP.get(uni, str(uni)))
    payloads = []
    for r in plan_cr:
        cod = str(r["cargo_cod"]); nome = r.get("cargo_nome", "")
        for mi in range(1, 13):
            payloads.append(dict(ano=ano, mes=mi, uni_cod=uni, unidade=unidade, cr_cod=cr, cr_nome=cr_nome,
                                 cargo_cod=cod, cargo_nome=nome,
                                 qtd_orcada=int(round(float(r.get(f"m{mi}") or 0)))))
    if payloads:
        c.table("hc_quadro").upsert(payloads, on_conflict="ano,mes,uni_cod,cr_cod,cargo_cod").execute()

def tela_qlp_ctrl(c, prof, ano):
    st.markdown(f"<div class='modtag'>Planejamento de Pessoal — QLP {ano}</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Abra/feche a janela do QLP, revise os envios, aprove (consolida no headcount orçado) ou devolva.</div>", unsafe_allow_html=True)
    anos_hab = get_plan_anos(c)
    if ano not in anos_hab:
        st.info(f"O QLP usa os mesmos anos habilitados do orçamento: {', '.join(map(str, anos_hab))}. "
                "Selecione um ano habilitado no topo (**Ano**) — os anos são geridos na tela de Planejamento (orçamento).")
        return
    aberta = get_qlp_janela(c, ano)
    jc = st.columns([3, 1.4])
    jc[0].markdown(f"<div style='padding-top:8px'>Janela do QLP {ano}: "
                   f"<b style='color:{VERDE if aberta else VERMELHO}'>{'ABERTA' if aberta else 'FECHADA'}</b></div>", unsafe_allow_html=True)
    if jc[1].button(("🔒 Fechar janela" if aberta else "🔓 Abrir janela"), key="qlp_toggle", use_container_width=True):
        set_qlp_janela(c, ano, not aberta); st.rerun()
    if st.button("🔄 Sincronizar cargos (catálogo da empresa)", key="qlp_sync_cargos",
                 help="Preenche o catálogo de cargos com TODOS os cargos já vistos no headcount de qualquer CR/ano. "
                      "Depois disso, os gestores passam a ver a empresa inteira no seletor de adicionar cargo."):
        todos = carregar_cargos_todos() or {}   # controladoria enxerga todos os CRs
        if not todos:
            st.info("Nenhum cargo encontrado no headcount para sincronizar.")
        else:
            try:
                existentes = {str(x["cargo_cod"]) for x in (carregar_hc_cargo() or []) if x.get("cargo_cod") is not None}
                novos = [{"cargo_cod": k, "cargo_nome": v, "ativo": True} for k, v in todos.items() if k not in existentes]
                for ch in chunks(novos, 500):
                    c.table("hc_cargo").insert(ch).execute()
                limpar_cache()
                st.success(f"Catálogo sincronizado: {len(novos)} cargo(s) novo(s); {len(existentes)} já existiam. "
                           "Os gestores já veem a empresa inteira no seletor.")
                st.rerun()
            except Exception as e:
                st.error(f"Não foi possível sincronizar os cargos: {e}")
    with st.expander("🗂️ Catálogo de cargos da empresa (alimenta o seletor de TODOS os gestores)"):
        cats = carregar_hc_cargo() or []
        st.caption(f"{len(cats)} cargo(s) no catálogo. Os gestores só conseguem adicionar cargos que estejam AQUI "
                   "(por segurança, eles não leem os cargos de CRs que não são deles). Mantenha aqui a lista completa da empresa.")
        st.markdown("**Adicionar um cargo**")
        ca = st.columns([1.2, 3, 1.1])
        ncod = ca[0].text_input("Código", key="cadcargo_cod")
        nnome = ca[1].text_input("Nome do cargo", key="cadcargo_nome")
        if ca[2].button("Adicionar", key="cadcargo_add", use_container_width=True):
            if ncod.strip() and nnome.strip():
                try:
                    c.table("hc_cargo").upsert({"cargo_cod": str(ncod).strip(), "cargo_nome": nnome.strip(), "ativo": True},
                                               on_conflict="cargo_cod").execute()
                    limpar_cache(); st.success(f"Cargo {ncod} · {nnome} salvo no catálogo."); st.rerun()
                except Exception as e:
                    st.error(f"Falha ao salvar: {e}")
            else:
                st.warning("Informe o código e o nome do cargo.")
        st.markdown("**Adicionar vários de uma vez** — cole um por linha no formato `código;nome`:")
        bulk = st.text_area("Colar lista de cargos", key="cadcargo_bulk", height=140,
                            placeholder="1001;GERENTE COMERCIAL\n1002;VENDEDOR\n1003;ASSISTENTE ADMINISTRATIVO")
        if st.button("Importar lista colada", key="cadcargo_bulk_btn"):
            registros = {}
            for linha in (bulk or "").splitlines():
                if not linha.strip():
                    continue
                partes = re.split(r"[;\t]", linha.strip(), maxsplit=1)
                if len(partes) == 2 and partes[0].strip():
                    registros[str(partes[0]).strip()] = partes[1].strip()
            if not registros:
                st.warning("Nada para importar. Use o formato `código;nome`, um por linha.")
            else:
                try:
                    payload = [{"cargo_cod": k, "cargo_nome": v, "ativo": True} for k, v in registros.items()]
                    for ch in chunks(payload, 500):
                        c.table("hc_cargo").upsert(ch, on_conflict="cargo_cod").execute()
                    limpar_cache(); st.success(f"{len(registros)} cargo(s) importado(s)/atualizado(s) no catálogo."); st.rerun()
                except Exception as e:
                    st.error(f"Falha ao importar: {e}")
        if cats:
            st.markdown("**Catálogo atual**")
            st.dataframe(pd.DataFrame([{"Código": x.get("cargo_cod"), "Cargo": x.get("cargo_nome"),
                                        "Ativo": x.get("ativo", True)} for x in cats]),
                         hide_index=True, use_container_width=True)
    st.divider()

    status_rows = carregar_qlp_status(ano)
    stt = {(int(s["uni_cod"]), int(s["cr_cod"])): s for s in status_rows}

    # ---- painel: quais gestores/CRs ainda faltam enviar o QLP ----
    vinc = carregar_cr_gestor()
    _EMP = {1: "PISA", 2: "KING"}
    universo = []
    for r in vinc:
        u = int(r.get("uni_cod", 0) or 0); crc = int(r.get("cr_cod", 0) or 0)
        gn = (r.get("gestor") or {}).get("nome", "") if isinstance(r.get("gestor"), dict) else ""
        universo.append((u, crc, r.get("cr_nome", ""), gn))
    def _situ(u, crc):
        return stt.get((u, crc), {}).get("status", "NÃO INICIADO")
    ok_list = [x for x in universo if _situ(x[0], x[1]) in ("ENVIADO", "APROVADO")]
    pend_list = [x for x in universo if _situ(x[0], x[1]) not in ("ENVIADO", "APROVADO")]
    kc = st.columns(3)
    kc[0].metric("CRs esperados", len(universo))
    kc[1].metric("Enviados/aprovados", len(ok_list))
    kc[2].metric("Faltam enviar", len(pend_list))
    if universo and pend_list:
        st.markdown(f"<div style='background:#FDECEA;border-left:4px solid {VERMELHO};padding:9px 13px;border-radius:6px;margin:4px 0 8px'>"
                    f"<b style='color:{VERMELHO}'>{len(pend_list)} de {len(universo)} CR(s) ainda não enviaram o QLP {ano}.</b></div>", unsafe_allow_html=True)
        ordp = {"DEVOLVIDO": 0, "RASCUNHO": 1, "NÃO INICIADO": 2}
        linhas = ""
        for u, crc, crn, gn in sorted(pend_list, key=lambda x: (ordp.get(_situ(x[0], x[1]), 9), x[1])):
            s = _situ(u, crc); cs = "#EF9F27" if s == "RASCUNHO" else VERMELHO
            linhas += (f"<tr><td style='text-align:left'>{gn or '—'}</td>"
                       f"<td style='text-align:left'>{_EMP.get(u, u)} · {crc} · {crn}</td>"
                       f"<td style='color:{cs}'>{s}</td></tr>")
        st.markdown(f"<div class='scroll'><table class='lle'><tr><th style='text-align:left'>Gestor</th>"
                    f"<th style='text-align:left'>Centro de resultado</th><th>Situação</th></tr>{linhas}</table></div>", unsafe_allow_html=True)
    elif universo:
        st.success(f"Todos os {len(universo)} CRs enviaram o QLP {ano}.")
    else:
        st.caption("Nenhum vínculo gestor↔CR cadastrado para comparar pendências.")

    with st.expander("📋 Finalização por gestor (QLP)", expanded=False):
        _painel_finalizacao_gestor(status_rows, ano, "QLP")

    # ---- extrato do QLP (planilha para enviar ao RH avaliar/consolidar) ----
    plan_all = carregar_qlp_plan(ano)
    if plan_all:
        try:
            import io
            EMP = {1: "PISA", 2: "KING"}
            ref_u = {int(x["uni_cod"]): (x.get("unidade") or "") for x in (carregar_hc_quadro(ano) or []) + (carregar_hc_quadro(ano - 1) or []) if x.get("uni_cod") is not None}
            regs = []
            for r in sorted(plan_all, key=lambda x: (int(x.get("uni_cod", 0) or 0), int(x.get("cr_cod", 0) or 0), str(x.get("cargo_cod", "")))):
                u = int(r.get("uni_cod", 0) or 0); crc = int(r.get("cr_cod", 0) or 0)
                situ = stt.get((u, crc), {}).get("status", "")
                meses_v = {MABREV[mi]: int(round(float(r.get(f"m{mi}") or 0))) for mi in range(1, 13)}
                reg = {"Empresa": ref_u.get(u) or EMP.get(u, u), "CR (cód)": crc, "Centro de resultado": r.get("cr_nome", ""),
                       "Cargo (cód)": r.get("cargo_cod", ""), "Cargo": r.get("cargo_nome", ""), "Situação": situ}
                reg.update(meses_v)
                reg["Pico do ano"] = max(meses_v.values()) if meses_v else 0
                reg["Parecer RH"] = ""  # janela de preenchimento do RH
                reg["Qtd consolidada (RH)"] = ""
                regs.append(reg)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as x:
                pd.DataFrame(regs).to_excel(x, index=False, sheet_name=f"QLP {ano}")
            st.download_button(f"📥 Baixar extrato do QLP {ano} (para o RH)", data=buf.getvalue(),
                               file_name=f"QLP_{ano}_extrato_RH.xlsx", mime=XLSX_MIME, key="qlp_extrato_dl",
                               help="Planilha com a projeção de headcount de todos os CRs, com colunas em branco para o RH avaliar e consolidar.")
            st.caption("Extrato de todos os CRs (projeção do gestor por cargo/mês) com colunas em branco para o RH preencher a avaliação/consolidação.")
        except Exception:
            pass
    st.divider()

    if not status_rows:
        st.info("Ainda não há envios do QLP para revisar.")
        return
    corcls = {"RASCUNHO": CINZA_TXT, "ENVIADO": AZUL_CORP, "APROVADO": VERDE, "DEVOLVIDO": VERMELHO}
    ordem = {"ENVIADO": 0, "DEVOLVIDO": 1, "RASCUNHO": 2, "APROVADO": 3}
    opts = sorted(stt.keys(), key=lambda k: (ordem.get(stt[k].get("status", ""), 9), k[1]))
    sel = st.selectbox("Centro de resultado", opts, format_func=lambda k: f"{k[1]} · {stt[k].get('status','')}", key="qlp_rev_cr")
    uni, cr = sel
    srow = stt[sel]; status = srow.get("status", "RASCUNHO")
    plan_rows = carregar_qlp_plan(ano)
    plan_cr = [r for r in plan_rows if int(r.get("uni_cod", 0) or 0) == uni and int(r.get("cr_cod", 0) or 0) == cr]
    cr_nome = plan_cr[0].get("cr_nome", "") if plan_cr else ""
    st.markdown(f"Situação: {chip(status, corcls.get(status, CINZA_TXT))}"
                + (f" &nbsp;·&nbsp; <span style='color:{CINZA_TXT}'>atualizado por {srow.get('atualizado_por','')}</span>" if srow.get('atualizado_por') else ""),
                unsafe_allow_html=True)
    if srow.get("comentario"):
        st.caption(f"Comentário registrado: {srow.get('comentario')}")

    if not plan_cr:
        st.info("Este CR ainda não tem headcount lançado.")
    else:
        th = "<th style='text-align:left'>Cargo</th>" + "".join(f"<th>{MABREV[m]}</th>" for m in range(1, 13)) + "<th>Pico</th>"
        corpo = ""
        for r in sorted(plan_cr, key=lambda x: str(x.get("cargo_cod", ""))):
            vals = [int(round(float(r.get(f"m{m}") or 0))) for m in range(1, 13)]
            corpo += (f"<tr><td style='text-align:left'>{r.get('cargo_cod','')} · {r.get('cargo_nome','')}</td>"
                      + "".join(f"<td style='text-align:center'>{v}</td>" for v in vals)
                      + f"<td style='text-align:center'><b>{max(vals) if vals else 0}</b></td></tr>")
        st.markdown(f"<div class='scroll'><table class='lle matrix'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)

    ref = carregar_hc_quadro(ano - 1) or carregar_hc_quadro(ano) or []
    hc_ano = carregar_hc_quadro(ano) or []
    conflito = any(int(x.get("uni_cod", 0) or 0) == uni and int(x.get("cr_cod", 0) or 0) == cr
                   and int(round(float(x.get("qtd_orcada") or 0))) != 0 for x in hc_ano)
    st.divider()
    if status == "APROVADO":
        st.success("CR já aprovado e consolidado no headcount orçado. Para reabrir, use Devolver.")
    ca = st.columns([1.4, 1.4, 3])
    ok_over = True
    if conflito and status != "APROVADO":
        st.warning(f"Já existe headcount orçado para o CR {cr} em {ano}. Aprovar vai SOBRESCREVER (o realizado é preservado).")
        ok_over = st.checkbox("Confirmo a sobrescrita", key=f"qlp_over_{uni}_{cr}")
    aprovar = ca[0].button("✅ Aprovar e consolidar", key=f"qlp_apr_{uni}_{cr}",
                           type="primary", disabled=(status == "APROVADO") or (conflito and not ok_over) or not plan_cr)
    coment = ca[2].text_input("Comentário (para devolução)", key=f"qlp_com_{uni}_{cr}")
    devolver = ca[1].button("↩️ Devolver", key=f"qlp_dev_{uni}_{cr}", disabled=not plan_cr)
    if aprovar:
        _consolidar_qlp(c, ano, uni, cr, cr_nome, plan_cr, ref)
        c.table("qlp_plan_status").upsert({"ano": ano, "uni_cod": uni, "cr_cod": cr, "status": "APROVADO",
                                           "comentario": "", "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        st.success(f"QLP do CR {cr} aprovado e consolidado no headcount orçado de {ano}."); st.rerun()
    if devolver:
        c.table("qlp_plan_status").upsert({"ano": ano, "uni_cod": uni, "cr_cod": cr, "status": "DEVOLVIDO",
                                           "comentario": coment, "atualizado_por": prof.get("nome", "")}, on_conflict="ano,uni_cod,cr_cod").execute()
        limpar_cache()
        st.success(f"QLP do CR {cr} devolvido para ajuste."); st.rerun()

# ---------------------------------------------------------------- main
# ============================================================================
# MÓDULO FORECAST / CENÁRIOS  (paralelo, só leitura sobre o orçamento oficial)
# O cenário NUNCA grava nas tabelas do orçamento. Guarda apenas metadados
# (forecast_cenario) e fatores de ajuste em % (forecast_fator). O valor do
# cenário é calculado ao vivo: base (dados originais, só leitura) × (1 + %).
# ============================================================================
@st.cache_data(ttl=30, show_spinner=False)
def _q_fc_cen(tok, rtok, ano):
    cc = _cli_tok(tok, rtok)
    try:
        return cc.table("forecast_cenario").select("*").eq("ano", ano).eq("ativo", True).order("id", desc=True).execute().data or []
    except Exception:
        return []
def carregar_fc_cenarios(ano): return _q_fc_cen(*_tok(), ano)

@st.cache_data(ttl=30, show_spinner=False)
def _q_fc_fat(tok, rtok, cid):
    cc = _cli_tok(tok, rtok)
    try:
        return cc.table("forecast_fator").select("*").eq("cenario_id", cid).order("id").execute().data or []
    except Exception:
        return []
def carregar_fc_fatores(cid): return _q_fc_fat(*_tok(), cid)

FC_FONTES = [("despesa", "Despesa (conta/CR)"), ("receita", "Receita Bruta"),
             ("cmv", "CMV"), ("deducao", "Deduções")]
FC_FONTE_LABEL = dict(FC_FONTES)
FC_BASES = [("orcado", "Orçado (planejado)"), ("real_proj", "Realizado até o corte + orçado no restante"),
            ("ano_anterior", "Ano anterior (realizado)")]
FC_BASE_LABEL = dict(FC_BASES)

def _fc_is_receita_conta(conta_cod):
    return str(conta_cod)[:1] in ("3", "6")

def _fc_fator_mult(fatores, fonte, mes, uni, cr=None, conta_cod=None, conta_nome=None):
    """Produto de (1 + pct/100) de todos os fatores que se aplicam à célula."""
    f = 1.0
    for ft in fatores:
        if ft.get("fonte") != fonte:
            continue
        if not (int(ft.get("mes_ini", 1) or 1) <= mes <= int(ft.get("mes_fim", 12) or 12)):
            continue
        fu = ft.get("uni_cod")
        if fu is not None and int(fu) != int(uni):
            continue
        if fonte == "despesa":
            fc = ft.get("cr_cod"); fk = ft.get("conta_cod")
            if fc is not None and int(fc) != int(cr or 0):
                continue
            if fk is not None and int(fk) != int(conta_cod or 0):
                continue
        if fonte == "deducao":
            fn = ft.get("conta")
            if fn not in (None, "", conta_nome):
                continue
        try:
            f *= (1.0 + float(ft.get("pct") or 0) / 100.0)
        except (TypeError, ValueError):
            pass
    return f

def _fc_calc(ano, cen, fatores, emp, meses):
    """Calcula os totais (base, cenário, realizado) das linhas do comparativo.
       Base e cenário respeitam a base escolhida; 'realizado' é sempre o realizado
       do ano corrente (referência). Nada é gravado."""
    base_tipo = cen.get("base_tipo", "orcado")
    corte = int(cen.get("corte_mes") or 0)
    ano_base = ano - 1 if base_tipo == "ano_anterior" else ano

    def val_base(row, mes):
        if base_tipo == "ano_anterior":
            return float(row.get("valor_realizado") or 0)
        if base_tipo == "real_proj":
            return float(row.get("valor_realizado") or 0) if (corte and mes <= corte) else float(row.get("valor_planejado") or 0)
        return float(row.get("valor_planejado") or 0)

    acc = {k: {"base": 0.0, "cen": 0.0, "real": 0.0} for k in ("receita", "deducao", "cmv", "despesa")}

    def somar(fonte, rows_base, rows_real, is_desp=False, is_ded=False):
        for r in (rows_base or []):
            mes = int(r.get("mes", 0) or 0)
            if not (1 <= mes <= 12) or mes not in meses:
                continue
            uni = int(r.get("uni_cod", 0) or 0)
            if emp and uni != emp:
                continue
            if is_desp and _fc_is_receita_conta(r.get("conta_cod")):
                continue
            b = val_base(r, mes)
            cr = int(r.get("cr_cod", 0) or 0) if is_desp else None
            conta_cod = int(r.get("conta_cod", 0) or 0) if is_desp else None
            conta_nome = r.get("conta") if is_ded else None
            mult = _fc_fator_mult(fatores, fonte, mes, uni, cr, conta_cod, conta_nome)
            acc[fonte]["base"] += b
            acc[fonte]["cen"] += b * mult
        for r in (rows_real or []):
            mes = int(r.get("mes", 0) or 0)
            if not (1 <= mes <= 12) or mes not in meses:
                continue
            uni = int(r.get("uni_cod", 0) or 0)
            if emp and uni != emp:
                continue
            if is_desp and _fc_is_receita_conta(r.get("conta_cod")):
                continue
            acc[fonte]["real"] += float(r.get("valor_realizado") or 0)

    somar("receita", carregar_receita(ano_base), carregar_receita(ano))
    somar("deducao", carregar_deducao(ano_base), carregar_deducao(ano), is_ded=True)
    somar("cmv", carregar_cmv(ano_base), carregar_cmv(ano))
    somar("despesa", carregar_orc(ano_base), carregar_orc(ano), is_desp=True)

    def trio(k): return acc[k]
    def diff(a, b): return {s: a[s] - b[s] for s in ("base", "cen", "real")}
    rb, ded, cmv, desp = trio("receita"), trio("deducao"), trio("cmv"), trio("despesa")
    rl = diff(rb, ded)
    lb = diff(rl, cmv)
    res = diff(lb, desp)
    linhas = [
        ("Receita Bruta", rb, "rev", False),
        ("(−) Deduções", ded, "cost", False),
        ("(=) Receita Líquida", rl, "rev", True),
        ("(−) CMV", cmv, "cost", False),
        ("(=) Lucro Bruto", lb, "rev", True),
        ("(−) Despesas (orçamento)", desp, "cost", False),
        ("(=) Resultado (antes de pessoal e financeiro)", res, "rev", True),
    ]
    return linhas

def tela_forecast(c, prof, ano):
    EMP = {1: "PISA", 2: "KING"}
    st.markdown("<div class='modtag'>Forecast / Cenários</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Projeções paralelas ao orçamento. Um cenário parte dos dados atuais e aplica "
                "ajustes em %. O orçamento oficial permanece intocado — nada aqui grava sobre ele.</div>", unsafe_allow_html=True)
    st.info("🔒 Este módulo é isolado: os cenários leem os números do orçamento apenas como base de cálculo e guardam "
            "somente os seus próprios ajustes. Receita, CMV, deduções e despesas originais não são alterados.")

    # ---------- criar novo cenário ----------
    with st.expander("➕ Novo cenário", expanded=False):
        nome = st.text_input("Nome do cenário", key="fc_novo_nome", placeholder="Ex.: Revisão jun, Otimista, Corte 10%")
        desc = st.text_input("Descrição (opcional)", key="fc_novo_desc")
        base_tipo = st.radio("Base de partida", [b[0] for b in FC_BASES],
                             format_func=lambda k: FC_BASE_LABEL[k], key="fc_novo_base")
        corte = None
        if base_tipo == "real_proj":
            corte = int(st.selectbox("Mês de corte (realizado até aqui; orçado a partir do mês seguinte)",
                                     list(range(1, 13)), index=5, format_func=lambda m: MESES[m], key="fc_novo_corte"))
        if st.button("Criar cenário", type="primary", key="fc_criar", disabled=(not nome.strip())):
            try:
                c.table("forecast_cenario").insert({
                    "ano": int(ano), "nome": nome.strip(), "descricao": desc.strip() or None,
                    "base_tipo": base_tipo, "corte_mes": corte, "criado_por": prof.get("nome", "")}).execute()
                limpar_cache(); st.success("Cenário criado."); st.rerun()
            except Exception as e:
                st.error(f"Não foi possível criar o cenário: {e}")

    cenarios = carregar_fc_cenarios(ano)
    if not cenarios:
        st.caption(f"Nenhum cenário para {ano} ainda. Crie o primeiro em “Novo cenário”.")
        return

    # ---------- escolher cenário ----------
    opt = {int(x["id"]): x for x in cenarios}
    cid = st.selectbox("Cenário", list(opt.keys()),
                       format_func=lambda i: f"{opt[i]['nome']}  ·  base: {FC_BASE_LABEL.get(opt[i].get('base_tipo'),'')}"
                                             + (f" (corte {MESES[int(opt[i]['corte_mes'])]})" if opt[i].get('corte_mes') else ""),
                       key="fc_sel")
    cen = opt[cid]
    fatores = carregar_fc_fatores(cid)

    csub = st.columns([4, 1])
    if cen.get("descricao"):
        csub[0].caption(f"📝 {cen['descricao']}")
    if csub[1].button("🗑️ Excluir cenário", key="fc_del_cen"):
        try:
            c.table("forecast_cenario").delete().eq("id", cid).execute()  # fatores caem por cascade
            limpar_cache(); st.success("Cenário excluído (o orçamento não foi afetado)."); st.rerun()
        except Exception as e:
            st.error(f"Não foi possível excluir: {e}")

    # ---------- adicionar fator de ajuste ----------
    st.divider()
    st.markdown("#### Ajustes em massa (fatores %)")
    st.caption("Cada ajuste multiplica a base pelo percentual, no escopo e período escolhidos. Ex.: “Despesa · KING · "
               "jul–dez · −10%”. Percentual negativo reduz; positivo aumenta. Vários ajustes se acumulam.")
    orc_rows = carregar_orc(ano) or []
    crs = sorted({(int(r["cr_cod"]), r.get("cr_nome", "")) for r in orc_rows if not _fc_is_receita_conta(r.get("conta_cod"))})
    contas = sorted({(int(r["conta_cod"]), r.get("conta_desc", "")) for r in orc_rows if not _fc_is_receita_conta(r.get("conta_cod"))})

    fc = st.columns([1.3, 1.1, 1.1])
    fonte = fc[0].selectbox("Fonte", [f[0] for f in FC_FONTES], format_func=lambda k: FC_FONTE_LABEL[k], key="fc_ff")
    empf = fc[1].selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="fc_femp")
    pct = fc[2].number_input("Percentual (%)", value=-10.0, step=1.0, format="%.2f", key="fc_fpct")
    cr_sel = conta_sel = 0; ded_sel = "Todas"
    if fonte == "despesa":
        fc2 = st.columns([2, 2])
        cr_sel = fc2[0].selectbox("Centro de resultado", [0] + [x[0] for x in crs],
                                  format_func=lambda x: "Todos" if x == 0 else f"{x} · {dict(crs).get(x,'')}", key="fc_fcr")
        conta_sel = fc2[1].selectbox("Conta", [0] + [x[0] for x in contas],
                                     format_func=lambda x: "Todas" if x == 0 else f"{x} · {dict(contas).get(x,'')}", key="fc_fconta")
    elif fonte == "deducao":
        ded_sel = st.selectbox("Dedução", ["Todas"] + DEDUCOES, key="fc_fded")
    fc3 = st.columns([1.1, 1.1, 2.6])
    mi = int(fc3[0].selectbox("Mês inicial", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="fc_fmi"))
    mf = int(fc3[1].selectbox("Mês final", list(range(1, 13)), index=11, format_func=lambda m: MESES[m], key="fc_fmf"))
    fdesc = fc3[2].text_input("Descrição do ajuste (opcional)", key="fc_fdesc")
    if st.button("Adicionar ajuste", type="primary", key="fc_add_fat", disabled=(mi > mf or pct == 0)):
        payload = {"cenario_id": cid, "fonte": fonte, "mes_ini": mi, "mes_fim": mf,
                   "pct": float(pct), "descricao": (fdesc.strip() or None), "criado_por": prof.get("nome", ""),
                   "uni_cod": (empf or None), "cr_cod": None, "conta_cod": None, "conta": None}
        if fonte == "despesa":
            payload["cr_cod"] = (cr_sel or None); payload["conta_cod"] = (conta_sel or None)
        elif fonte == "deducao":
            payload["conta"] = (None if ded_sel == "Todas" else ded_sel)
        try:
            c.table("forecast_fator").insert(payload).execute()
            limpar_cache(); st.success("Ajuste adicionado."); st.rerun()
        except Exception as e:
            st.error(f"Não foi possível adicionar o ajuste: {e}")
    if mi > mf:
        st.warning("O mês inicial não pode ser maior que o mês final.")

    # ---------- lista de fatores ----------
    if fatores:
        st.markdown("###### Ajustes deste cenário")
        for ft in fatores:
            escopo = [FC_FONTE_LABEL.get(ft.get("fonte"), ft.get("fonte"))]
            if ft.get("uni_cod"): escopo.append(EMP.get(int(ft["uni_cod"]), str(ft["uni_cod"])))
            else: escopo.append("Todas as empresas")
            if ft.get("cr_cod"): escopo.append(f"CR {ft['cr_cod']}")
            if ft.get("conta_cod"): escopo.append(f"conta {ft['conta_cod']}")
            if ft.get("conta"): escopo.append(ft["conta"])
            pcttxt = f"{float(ft.get('pct') or 0):+.2f}%".replace(".", ",")
            per = f"{MABREV[int(ft.get('mes_ini',1))]}–{MABREV[int(ft.get('mes_fim',12))]}"
            cor = VERDE if float(ft.get("pct") or 0) < 0 else VERMELHO
            row = st.columns([7, 1])
            row[0].markdown(f"• <b style='color:{cor}'>{pcttxt}</b> · {' · '.join(escopo)} · {per}"
                            + (f" — <i>{ft['descricao']}</i>" if ft.get("descricao") else ""), unsafe_allow_html=True)
            if row[1].button("Remover", key=f"fc_delfat_{ft['id']}"):
                try:
                    c.table("forecast_fator").delete().eq("id", int(ft["id"])).execute()
                    limpar_cache(); st.rerun()
                except Exception as e:
                    st.error(f"Não foi possível remover: {e}")
    else:
        st.caption("Nenhum ajuste ainda — sem ajustes, o cenário é igual à base.")

    # ---------- comparativo Base × Cenário × Realizado ----------
    st.divider()
    st.markdown("#### Comparativo")
    cco = st.columns([1.3, 2.4])
    emp_cmp = cco[0].selectbox("Empresa", [0, 1, 2], format_func=lambda x: "Todas" if x == 0 else EMP[x], key="fc_cmp_emp")
    visao = cco[1].radio("Período", ["Ano inteiro", "Até um mês", "Intervalo"], horizontal=True, key="fc_cmp_visao")
    if visao == "Até um mês":
        m_ate = int(st.selectbox("Até", list(range(1, 13)), index=11, format_func=lambda m: MESES[m], key="fc_cmp_ate"))
        meses = list(range(1, m_ate + 1))
    elif visao == "Intervalo":
        ci = st.columns(2)
        m1 = int(ci[0].selectbox("De", list(range(1, 13)), index=0, format_func=lambda m: MESES[m], key="fc_cmp_de"))
        m2 = int(ci[1].selectbox("Até", list(range(1, 13)), index=11, format_func=lambda m: MESES[m], key="fc_cmp_ateb"))
        meses = list(range(min(m1, m2), max(m1, m2) + 1))
    else:
        meses = list(range(1, 13))

    linhas = _fc_calc(ano, cen, fatores, emp_cmp, meses)
    corpo = ""
    for nome, tri, tipo, forte in linhas:
        b0, b1 = ("<b>", "</b>") if forte else ("", "")
        base_v = tri["base"]; cen_v = tri["cen"]; real_v = tri["real"]
        dcen = cen_v - base_v
        dpct = (dcen / base_v * 100) if base_v else 0.0
        cor = CINZA_TXT if abs(dpct) < 0.05 else ((VERDE if dcen >= 0 else VERMELHO) if tipo == "rev" else (VERMELHO if dcen > 0 else VERDE))
        cls = " class='mark'" if forte else ""
        corpo += (f"<tr{cls}><td style='text-align:left'>{b0}{nome}{b1}</td>"
                  f"<td>{b0}{brl(base_v)}{b1}</td><td>{b0}{brl(cen_v)}{b1}</td>"
                  f"<td style='color:{cor}'>{b0}{brl(dcen)}{b1}</td>"
                  f"<td style='color:{cor}'>{b0}{pct_txt(dpct)}{b1}</td>"
                  f"<td style='color:{CINZA_TXT}'>{brl(real_v)}</td></tr>")
    th = ("<th style='text-align:left'>Linha</th><th>Base</th><th>Cenário</th>"
          "<th>Δ Cenário</th><th>Δ %</th><th>Realizado (ref.)</th>")
    st.markdown(f"<div class='scroll'><table class='lle matrix'><tr>{th}</tr>{corpo}</table></div>", unsafe_allow_html=True)
    st.caption(f"Base: {FC_BASE_LABEL.get(cen.get('base_tipo'),'')}"
               + (f" · corte em {MESES[int(cen['corte_mes'])]}" if cen.get('corte_mes') else "")
               + ". “Realizado (ref.)” é o realizado do ano corrente, apenas para comparação. Pessoal e resultado "
               "financeiro não entram neste cenário (fora do escopo escolhido).")

SECOES_CTRL = [
    ("Acompanhamento", [("acomp", "📊 Acompanhamento (orçado x realizado)"),
                        ("justif", "📥 Justificativas recebidas")]),
    ("Demonstrativos", [("dre", "📈 DRE"), ("forecast", "🔮 Forecast / Cenários")]),
    ("Orçamento", [("receita", "💰 Receita de Vendas"), ("deducao", "➖ Deduções de Vendas"),
                   ("cmv", "🧾 CMV"), ("investimento", "🏗️ Investimentos"),
                   ("plan", "🧭 Planejamento (orçamento)"), ("manut", "✏️ Manutenção Orçamento")]),
    ("Pessoal", [("pessoal", "👥 Gastos com Pessoal"),
                 ("qlp", "🧭 Planejamento de Pessoal (QLP)")]),
    ("Administração", [("importar", "⬆️ Importar dados"), ("crconfig", "🏢 Unidades de negócio / Config DRE"),
                       ("admin", "🔑 Administração de acessos")]),
]
SECOES_GESTOR = [
    ("Acompanhamento", [("acomp", "📊 Acompanhamento (orçado x realizado)"),
                        ("justif", "📝 Justificativas")]),
    ("Orçamento", [("plan", "🧭 Planejamento (orçamento)")]),
    ("Pessoal", [("pessoal", "👥 Gastos com Pessoal (meu CR)"),
                 ("qlp", "🧭 Planejamento de Pessoal (QLP)")]),
]

def tela_cr_corporativo(c, prof, ano):
    st.markdown("<div class='modtag'>Unidades de negócio / Configuração da DRE</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Configurações da DRE em um só lugar: os centros de resultado que compõem a "
                "unidade Corporativo e o mapeamento das contas do orçamento para as linhas da DRE.</div>", unsafe_allow_html=True)
    orc = carregar_orc(ano) or []
    EMPn = {1: "PISA", 2: "KING"}

    # ---- 1) CRs corporativos ----
    st.markdown("#### 1) Centros de resultado da unidade Corporativo")
    st.caption("Marque os CRs que atendem ao corporativo. Na DRE “Por unidade”, as despesas e o pessoal desses CRs "
               "aparecem numa coluna Corporativo, saindo de PISA/KING — apenas na visualização. Nenhum valor é alterado.")
    crmap = {}
    for x in orc:
        cd = x.get("cr_cod")
        if cd is None:
            continue
        cd = int(cd)
        crmap.setdefault(cd, {"nome": x.get("cr_nome", "") or "", "unis": set()})
        u = int(x.get("uni_cod", 0) or 0)
        if u:
            crmap[cd]["unis"].add(u)
    atuais = get_cr_corporativos(c)
    if not crmap:
        st.info("Nenhum centro de resultado no orçamento deste ano ainda. Importe o orçamento primeiro.")
    else:
        opts = sorted(crmap.keys())
        def _lbl(cd):
            info = crmap[cd]
            emps = "/".join(EMPn.get(u, str(u)) for u in sorted(info["unis"])) or "—"
            return f"{cd} · {info['nome']} ({emps})"
        sel = st.multiselect("CRs corporativos", opts, default=[cd for cd in opts if cd in atuais],
                             format_func=_lbl, key="crcorp_ms")
        if st.button("Salvar CRs corporativos", type="primary", key="crcorp_save"):
            set_cr_corporativos(c, sel); limpar_cache()
            st.success(f"{len(sel)} CR(s) marcado(s) como corporativo. A DRE por unidade já reflete a mudança."); st.rerun()
        fora = sorted(cd for cd in atuais if cd not in crmap)
        if fora:
            st.caption("Também marcados, mas sem lançamento no orçamento deste ano: " + ", ".join(str(x) for x in fora))

    st.divider()
    # ---- 2) Mapeamento de contas para a DRE ----
    st.markdown("#### 2) Mapeamento de contas do orçamento para a DRE")
    mapa = {int(x["conta_cod"]): x.get("linha") for x in (carregar_dre_mapa() or []) if x.get("conta_cod") is not None}
    desc_por_cod = {}
    for x in orc:
        cod = x.get("conta_cod")
        if cod is None:
            continue
        cod = int(cod); d = str(x.get("conta_desc", "") or "").strip()
        desc_por_cod.setdefault(cod, [])
        if d and d not in desc_por_cod[cod]:
            desc_por_cod[cod].append(d)
    contas = sorted((cod, " / ".join(ds) if ds else "") for cod, ds in desc_por_cod.items())
    colididos = {cod: ds for cod, ds in desc_por_cod.items() if len(ds) > 1}
    if not contas:
        st.caption("Nenhuma conta de orçamento carregada para este ano.")
    else:
        if colididos:
            aviso = "; ".join(f"{cod} ({' / '.join(ds)})" for cod, ds in sorted(colididos.items()))
            st.warning("Estes códigos aparecem no orçamento com **mais de uma descrição** — o sistema os trata como "
                       "uma única conta (soma tudo sob o código). Se forem contas diferentes, corrija o código na "
                       f"origem (ERP/plano de contas): {aviso}.")
        _mapa_dre_frag(contas, mapa, c, prof)

def barra_lateral(prof, secoes):
    """Menu lateral setorizado com a logo LLE. Retorna a chave da tela ativa."""
    allowed = [k for _, itens in secoes for k, _ in itens]
    if st.session_state.get("nav") not in allowed:
        st.session_state["nav"] = allowed[0]
    with st.sidebar:
        st.markdown(f"<div class='side-logo'>{LOGO}<div class='side-logo-txt'>GRUPO LLE</div></div>",
                    unsafe_allow_html=True)
        for titulo, itens in secoes:
            st.markdown(f"<div class='navsec'>{titulo}</div>", unsafe_allow_html=True)
            for key, label in itens:
                tipo = "primary" if st.session_state["nav"] == key else "secondary"
                if st.button(label, key=f"nav_{key}", use_container_width=True, type=tipo):
                    st.session_state["nav"] = key
                    st.rerun()
        papel = "Controladoria" if prof["papel"] == "controladoria" else "Gestor"
        st.markdown("<div class='navsec'>Acesso</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='side-user'>👤 <b>{prof['nome']}</b><br><span>{papel}</span></div>",
                    unsafe_allow_html=True)
    return st.session_state["nav"]

def tela_justif_gestor(c, prof, banda, df_orc, ano, mes):
    st.markdown(f"<div class='modtag'>Justificativas · {MESES[mes]}/{ano}</div>", unsafe_allow_html=True)
    st.markdown("<div class='modsub'>Responda as contas com desvio desfavorável do mês selecionado.</div>", unsafe_allow_html=True)
    if df_orc.empty:
        st.info("Nenhum dado carregado ainda. Fale com a controladoria.")
        return
    if mes < get_cobranca(c):
        st.info(f"{MESES[mes]}/{ano} não está sujeito à cobrança de justificativa.")
        return
    d_mes = df_orc[df_orc["mes"] == mes]
    secao_justificativas(c, prof, d_mes, mes, False, banda, ano)

def render_app(c, prof):
    """Renderiza o app autenticado (cabeçalho, barra, abas/telas e rodapé)."""
    is_ctrl = prof["papel"] == "controladoria"
    banda = get_faixa(c)

    header(prof)

    # barra superior: faixa vigente + sair
    tb = st.columns([6, 2.2, 1.1])
    tb[1].markdown(
        f"<div style='text-align:right; padding-top:6px; color:{CINZA_TXT}; font-size:12px;'>"
        f"Faixa neutra vigente: <b style='color:{AZUL_PROFUNDO}'>±{banda:.1f}%</b>".replace(".", ",")
        + (f" · Cobrança desde {MESES[get_cobranca(c)]}" if is_ctrl else "") + "</div>",
        unsafe_allow_html=True)
    if tb[2].button("Sair", key="sair_top", use_container_width=True):
        for k in ("access_token", "refresh_token", "email"): st.session_state.pop(k, None)
        st.rerun()

    # menu lateral primeiro (define a tela ativa)
    secoes = SECOES_CTRL if is_ctrl else SECOES_GESTOR
    nav = barra_lateral(prof, secoes)

    # ----- seletor global de período -----
    # Ano vale para TODAS as telas. O Mês só é usado por Acompanhamento, Justificativas,
    # Pessoal e Manutenção — nas demais (DRE e telas anuais) ele seria inerte, então nem aparece.
    # Ano vale para TODAS as telas. Histórico desde 2020; o teto acompanha os anos
    # habilitados para planejamento (hoje 2027) — nada além disso aparece até ser liberado.
    anos_hab = get_plan_anos(c)
    teto = max([2026] + list(anos_hab))
    ANOS = list(range(2020, teto + 1))
    usa_mes = nav in ("acomp", "justif", "pessoal", "manut")
    if usa_mes:
        pc = st.columns([1.1, 1.4, 6])
        ano = int(pc[0].selectbox("Ano", ANOS, index=(ANOS.index(2026) if 2026 in ANOS else 0), key="g_ano"))
        mes = int(pc[1].selectbox("Mês", list(range(1, 13)), index=5, format_func=lambda m: MESES[m], key="g_mes"))
    else:
        pc = st.columns([1.1, 7])
        ano = int(pc[0].selectbox("Ano", ANOS, index=(ANOS.index(2026) if 2026 in ANOS else 0), key="g_ano"))
        mes = int(st.session_state.get("g_mes", 6) or 6)

    # dados compartilhados (cacheados por token — releituras idênticas ficam instantâneas)
    orc = carregar_orc(ano)
    df_orc = pd.DataFrame(orc) if orc else pd.DataFrame()
    cg = {}
    try:
        for r in carregar_cr_gestor():
            cg[(int(r["uni_cod"]), int(r["cr_cod"]))] = ((r.get("gestor") or {}).get("nome", "—"), r.get("cr_nome", ""))
    except Exception:
        cg = {}

    if is_ctrl:
        if nav == "acomp":        tela_acompanhamento(c, prof, banda, df_orc, cg, is_ctrl, ano, mes)
        elif nav == "justif":     tela_painel(c, prof, banda, df_orc, cg, ano, mes)
        elif nav == "dre":        tela_dre(c, prof, ano)
        elif nav == "forecast":   tela_forecast(c, prof, ano)
        elif nav == "receita":    tela_receita(c, prof, ano)
        elif nav == "deducao":    tela_deducao(c, prof, ano)
        elif nav == "cmv":        tela_cmv(c, prof, ano)
        elif nav == "investimento": tela_investimento(c, prof, ano)
        elif nav == "pessoal":    tela_headcount(c, prof, ano, mes)
        elif nav == "manut":      tela_editar_orcado(c, prof, df_orc, ano, mes)
        elif nav == "plan":       tela_planejamento_ctrl(c, prof, ano)
        elif nav == "qlp":        tela_qlp_ctrl(c, prof, ano)
        elif nav == "admin":      tela_admin(c, prof, ano)
        elif nav == "crconfig":   tela_cr_corporativo(c, prof, ano)
        elif nav == "importar":   tela_importar(c, ano)
    else:
        if nav == "justif":
            tela_justif_gestor(c, prof, banda, df_orc, ano, mes)
        elif nav == "plan":
            tela_planejamento_gestor(c, prof, ano)
        elif nav == "pessoal":
            tela_headcount(c, prof, ano, mes, somente_leitura=True)
        elif nav == "qlp":
            tela_qlp_gestor(c, prof, ano)
        else:
            tela_acompanhamento(c, prof, banda, df_orc, cg, is_ctrl, ano, mes, mostrar_justif=False)

    rodape()

def main():
    inject_css()
    if not URL or not ANON:
        st.error("Faltam os segredos SUPABASE_URL e SUPABASE_ANON_KEY."); return

    # Âncora única: login e app ocupam o MESMO espaço na árvore.
    # Assim, ao autenticar, o app substitui a tela de login sem deixar "fantasma".
    slot = st.empty()

    if "access_token" not in st.session_state:
        with slot.container():
            tela_login()
        return

    c = client(); prof = perfil(c, st.session_state.get("email", ""))
    if not prof:
        with slot.container():
            st.error("Seu e-mail não está cadastrado. Fale com a controladoria.")
            if st.button("Sair"): st.session_state.clear(); st.rerun()
        return
    if prof.get("senha_provisoria"):
        with slot.container():
            tela_trocar_senha(c, st.session_state.get("email", ""))
        return

    with slot.container():
        render_app(c, prof)

main()
